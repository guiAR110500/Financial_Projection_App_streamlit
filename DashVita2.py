import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
import random
from abc import ABC, abstractmethod 
import math
from scipy import stats


# Load custom CSS         
def load_css():
    with open('style.css') as f:
        st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)
        
# Call this at the start
load_css()  

def styled_title(text, level=1):
    """Display a styled title with rounded corners and custom background color
    
    Args:
        text (str): The title text
        level (int): Heading level (1, 2, or 3 for h1, h2, or h3)
    """
    tag = f"h{level}"
    st.markdown(f"""
        <div class="styled-title">
            <{tag}>{text}</{tag}>
        </div>
    """, unsafe_allow_html=True)

class DataHandler:
    @staticmethod
    def load_data(file_path):
        try:
            return pd.read_csv(file_path, sep=';')
        except Exception as e:
            st.error(f"Erro ao carregar o arquivo {file_path}: {e}")
            return None    

class PlotManager:
    @staticmethod
    def create_bar_plot(data, x_column, y_column, title):
        fig = px.bar(
            data,
            x=x_column,
            y=y_column,
            title=title,   
            labels={y_column: y_column, x_column: x_column}
        )
        fig.update_layout(
            title_x=0.5,
            xaxis_title=x_column,
            yaxis_title=y_column,
            xaxis_tickangle=-45
        )
        return fig
    
    @staticmethod
    def create_pie_plot(data, values_column, labels_column, title):
        # Use absolute values for pie chart
        data_for_pie = data.copy()
        data_for_pie[values_column] = data_for_pie[values_column].abs()
        
        fig = px.pie(
            data_for_pie,
            values=values_column,
            names=labels_column,
            title=title,
            hole=0.3
        )
        fig.update_layout(title_x=0.5)
        return fig

class Page(ABC):
    @abstractmethod
    def render(self):
        pass
    
    @property
    @abstractmethod
    def title(self) -> str:
        pass
    
    @property
    @abstractmethod
    def icon(self) -> str:
        pass

class PremissasInvestimentos(Page):                
    def __init__(self):
        # Default empty lists for investments, partner investments, and future investments
        default_params = {
            'investimentos_iniciais': [],
            'investimentos_socios': [],   # Added for partner investments
            'investimentos_futuros': []   # Added for future investments
        }
        
        # Initialize or update session state
        if 'premissas_investimentos' not in st.session_state:
            st.session_state['premissas_investimentos'] = default_params
        else:
            # Check for any missing keys and add them with default values
            for key, value in default_params.items():
                if key not in st.session_state['premissas_investimentos']:
                    st.session_state['premissas_investimentos'][key] = value
    
    @property
    def title(self) -> str:
        return "Premissas Investimentos"
    
    @property
    def icon(self) -> str:
        return "💰"
        
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Create tabs for different types of investments
        tab1, tab2, tab3 = st.tabs(["Investimento Inicial", "Investimentos dos Sócios", "Investimentos Futuros"])
        
        with tab1:
            st.write("### Configuração de Investimentos Iniciais")
            st.write("Defina os itens que compõem o investimento inicial do projeto.")
            
            # Display current investments
            if st.session_state['premissas_investimentos']['investimentos_iniciais']:
                st.write("#### Investimentos configurados:")
                
                # Create a DataFrame for display
                data = []
                for item in st.session_state['premissas_investimentos']['investimentos_iniciais']:
                    total = item['quantidade'] * item['valor_unitario']
                    data.append({
                        "Descrição": item['descricao'],
                        "Quantidade": item['quantidade'],
                        "Valor Unitário": f"R$ {item['valor_unitario']:,.2f}",
                        "Total": f"R$ {total:,.2f}"
                    })
                
                df_display = pd.DataFrame(data)
                st.dataframe(df_display, use_container_width=True)
                
                # Calculate and display the total investment
                total_investment = sum(item['quantidade'] * item['valor_unitario'] 
                                      for item in st.session_state['premissas_investimentos']['investimentos_iniciais'])
                st.metric("Investimento Total", f"R$ {total_investment:,.2f}")
                
                # Option to remove all items
                if st.button("Remover Todos os Itens", key="remove_initial_investments"):
                    st.session_state['premissas_investimentos']['investimentos_iniciais'] = []
                    st.success("Todos os itens foram removidos.")
                    st.rerun()
            
            # Form to add new investment item
            st.write("#### Adicionar Novo Item de Investimento")
            with st.form("novo_investimento"):
                descricao = st.text_input("Descrição", key="inv_descricao")
                
                col1, col2 = st.columns(2)
                with col1:
                    quantidade = st.number_input("Quantidade", min_value=1, value=1, step=1, key="inv_quantidade")
                
                with col2:
                    valor_unitario = st.number_input("Valor Unitário (R$)", min_value=0.0, value=0.0, step=100.0, format="%.2f", key="inv_valor")
                
                # Calculate and display the total for this item
                total = quantidade * valor_unitario
                st.write(f"Total para este item: R$ {total:,.2f}")
                
                submitted = st.form_submit_button("Adicionar Item")
                
                if submitted and descricao and valor_unitario > 0:
                    # Add the new investment to the list
                    novo_investimento = {
                        'descricao': descricao,
                        'quantidade': quantidade,
                        'valor_unitario': valor_unitario
                    }
                    
                    st.session_state['premissas_investimentos']['investimentos_iniciais'].append(novo_investimento)
                    st.success(f"Item '{descricao}' adicionado com sucesso!")
                    st.rerun()
        
        # New tab for Partner Investments
        with tab2:
            st.write("### Investimentos dos Sócios")
            st.write("Configure os aportes de capital dos sócios ao longo do tempo.")
            
            # Display existing investments
            if st.session_state['premissas_investimentos']['investimentos_socios']:
                st.write("#### Aportes configurados:")
                
                # Create a DataFrame for display
                data = []
                for idx, inv in enumerate(st.session_state['premissas_investimentos']['investimentos_socios']):
                    data.append({
                        "Valor": f"R$ {inv['valor']:,.2f}",
                        "Mês Inflow": inv['mes_inflow'],
                        "Periodicidade": "A cada " + str(inv['periodicidade']) + " meses" if inv['periodicidade_ativa'] else "Única"
                    })
                
                df_display = pd.DataFrame(data)
                st.dataframe(df_display, use_container_width=True)
                
                # Option to remove all items
                if st.button("Remover Todos os Aportes", key="remove_partner_investments"):
                    st.session_state['premissas_investimentos']['investimentos_socios'] = []
                    st.success("Todos os aportes foram removidos.")
                    st.rerun()
            
            # Form to add new partner investment
            st.write("#### Adicionar Novo Aporte de Capital")
            with st.form("novo_aporte"):
                col1, col2 = st.columns(2)
                with col1:
                    valor_inv = st.number_input("Valor (R$)", min_value=0.0, step=1000.0, format="%.2f", key="aporte_valor", help="Valor do aporte dos sócios.")
                    mes_inflow_inv = st.number_input("Mês Inflow", min_value=0, max_value=59, step=1, key="aporte_mes", help="Mês em que o aporte entra no fluxo de caixa (0 a 59).")
                
                with col2:
                    periodicidade_ativa_inv = st.checkbox("Possui Periodicidade?", key="aporte_per_ativa", help="Se o aporte possui periodicidade ou não.")
                    periodicidade_inv = st.number_input(
                        "Periodicidade (meses)",
                        min_value=1,
                        max_value=12,
                        value=3,
                        step=1,
                        key="aporte_periodicidade",
                        help="Periodicidade do aporte. Deve ser maior que 0 se a periodicidade for ativada."
                    )
                
                submitted = st.form_submit_button("Adicionar Aporte")
                
                if submitted and valor_inv > 0:
                    # Create new investment and add to list
                    novo_aporte = {
                        'valor': valor_inv,
                        'mes_inflow': mes_inflow_inv,
                        'periodicidade_ativa': periodicidade_ativa_inv,
                        'periodicidade': periodicidade_inv if periodicidade_ativa_inv else 1
                    }
                    st.session_state['premissas_investimentos']['investimentos_socios'].append(novo_aporte)
                    st.success(f"Aporte de R$ {valor_inv:,.2f} adicionado com sucesso!")
                    st.rerun()
        
        # New tab for Future Investments
        with tab3:
            st.write("### Investimentos Futuros")
            st.write("Configure investimentos futuros para ampliações e melhorias.")
            
            # Display existing future investments
            if st.session_state['premissas_investimentos']['investimentos_futuros']:
                st.write("#### Investimentos futuros configurados:")
                
                # Create a DataFrame for display
                data = []
                for idx, fut in enumerate(st.session_state['premissas_investimentos']['investimentos_futuros']):
                    data.append({
                        "Descrição": fut['descricao'],
                        "Valor": f"R$ {fut['valor']:,.2f}",
                        "Mês Outflow": fut['mes_outflow'],
                        "Periodicidade": "A cada " + str(fut['periodicidade']) + " meses" if fut['periodicidade_ativa'] else "Única"
                    })
                
                df_display = pd.DataFrame(data)
                st.dataframe(df_display, use_container_width=True)
                
                # Option to remove all items
                if st.button("Remover Todos os Investimentos Futuros", key="remove_future_investments"):
                    st.session_state['premissas_investimentos']['investimentos_futuros'] = []
                    st.success("Todos os investimentos futuros foram removidos.")
                    st.rerun()
            
            # Form to add new future investment
            st.write("#### Adicionar Novo Investimento Futuro")
            with st.form("novo_inv_futuro"):
                descricao_fut = st.text_input("Descrição do Investimento", key="fut_desc")
                
                col1, col2 = st.columns(2)
                with col1:
                    valor_fut = st.number_input("Valor (R$)", min_value=0.0, step=1000.0, format="%.2f", key="fut_valor", help="Valor do investimento futuro.")
                    mes_outflow_fut = st.number_input("Mês Outflow", min_value=0, max_value=59, step=1, key="fut_mes", help="Mês em que o investimento sai do fluxo de caixa (0 a 59).")
                
                with col2:
                    periodicidade_ativa_fut = st.checkbox("Possui Periodicidade?", key="fut_per_ativa", help="Se o investimento futuro possui periodicidade ou não.")
                    periodicidade_fut = st.number_input(
                        "Periodicidade (meses)",
                        min_value=1,
                        max_value=12,
                        value=6,
                        step=1,
                        key="fut_periodicidade",
                        help="Periodicidade do investimento futuro. Deve ser maior que 0 se a periodicidade for ativada."
                    )
                
                submitted = st.form_submit_button("Adicionar Investimento Futuro")
                
                if submitted and descricao_fut and valor_fut > 0:
                    # Create new future investment and add to list
                    novo_investimento_futuro = {
                        'descricao': descricao_fut,
                        'valor': valor_fut,
                        'mes_outflow': mes_outflow_fut,
                        'periodicidade_ativa': periodicidade_ativa_fut,
                        'periodicidade': periodicidade_fut if periodicidade_ativa_fut else 1
                    }
                    st.session_state['premissas_investimentos']['investimentos_futuros'].append(novo_investimento_futuro)
                    st.success(f"Investimento futuro '{descricao_fut}' adicionado com sucesso!")
                    st.rerun()

class Investimentos(Page):
    def __init__(self):
        self.graph_types = ["Gráfico de Barras", "Gráfico de Linhas", "Gráfico de Pizza"]
        self.time_frames = ["Mensal", "Trimestral", "Semestral", "Anual"]
    
    @property
    def title(self):
        return "Investimentos"
    
    @property
    def icon(self):
        return "💰"
    
    def get_total_investimento(self):
        """Retorna o valor total do investimento inicial"""
        # Check if premissas_investimentos exists
        if 'premissas_investimentos' not in st.session_state:
            return 0.0
        
        # Get the investments data
        investimentos = st.session_state['premissas_investimentos']['investimentos_iniciais']
        
        if not investimentos:
            return 0.0
        
        # Calculate total investment
        total_investment = sum(item['quantidade'] * item['valor_unitario'] 
                            for item in investimentos)
        
        return total_investment
    
    def _gerar_dataframe_investimentos(self):
        """Gera o dataframe de investimentos para 60 meses (0 a 59)"""
        meses = range(60)
        
        # Criar DataFrame com os investimentos
        df = pd.DataFrame(0.0, index=pd.Index(["Investimento Inicial", "Investimentos dos Sócios", 
                                               "Investimentos Futuros", "Total"]), columns=meses)
        
        # Investimento Inicial (apenas no mês 0)
        total_investimento_inicial = self.get_total_investimento()
        df.loc["Investimento Inicial", 0] = -total_investimento_inicial  # Negativo pois é saída
        
        # Investimentos dos Sócios
        if 'premissas_investimentos' in st.session_state:
            for inv in st.session_state['premissas_investimentos']['investimentos_socios']:
                valor = inv['valor']
                mes_inicial = inv['mes_inflow']
                periodicidade = inv['periodicidade'] if inv['periodicidade_ativa'] else 0
                
                if periodicidade > 0:
                    # Investimento periódico
                    for mes in range(mes_inicial, 60, periodicidade):
                        df.loc["Investimentos dos Sócios", mes] += valor
                else:
                    # Investimento único
                    if mes_inicial < 60:
                        df.loc["Investimentos dos Sócios", mes_inicial] += valor
        
        # Investimentos Futuros
        if 'premissas_investimentos' in st.session_state:
            for inv in st.session_state['premissas_investimentos']['investimentos_futuros']:
                valor = -inv['valor']  # Negativo pois é saída
                mes_inicial = inv['mes_outflow']
                periodicidade = inv['periodicidade'] if inv['periodicidade_ativa'] else 0
                
                if periodicidade > 0:
                    # Investimento periódico
                    for mes in range(mes_inicial, 60, periodicidade):
                        df.loc["Investimentos Futuros", mes] += valor
                else:
                    # Investimento único
                    if mes_inicial < 60:
                        df.loc["Investimentos Futuros", mes_inicial] += valor
        
        # Calcular total para cada mês
        for mes in meses:
            inv_inicial = pd.to_numeric(df.loc["Investimento Inicial", mes], errors='coerce')
            inv_socios = pd.to_numeric(df.loc["Investimentos dos Sócios", mes], errors='coerce')
            inv_futuros = pd.to_numeric(df.loc["Investimentos Futuros", mes], errors='coerce')
            
            # Use .fillna(0) to replace NaN values with 0 before summing
            total = pd.Series([inv_inicial, inv_socios, inv_futuros]).fillna(0).sum()
            df.loc["Total", mes] = total
        
        return df
    
    def _agrupar_por_periodo(self, df, periodo):
        """Agrupa o dataframe por período (trimestral, semestral ou anual)"""
        if periodo == "Mensal":
            return df
        
        # Definir o tamanho do grupo
        if periodo == "Trimestral":
            tamanho_grupo = 3
        elif periodo == "Semestral":
            tamanho_grupo = 6
        else:  # Anual
            tamanho_grupo = 12
        
        # Criar novo DataFrame
        periodos = [f"{periodo} {i//tamanho_grupo + 1}" for i in range(0, 60, tamanho_grupo)]
        df_agrupado = pd.DataFrame(index=df.index, columns=pd.Index(periodos))
        
        # Agrupar os dados
        for i, nome_periodo in enumerate(periodos):
            inicio = i * tamanho_grupo
            fim = min((i + 1) * tamanho_grupo, 60)
            
            # Calcular a soma para o período
            for idx in df.index:
                df_agrupado.loc[idx, nome_periodo] = df.loc[idx, inicio:fim-1].sum()
        
        return df_agrupado

    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Check if premissas_investimentos exists
        if 'premissas_investimentos' not in st.session_state:
            st.error("Premissas de investimentos não definidas. Por favor, defina as premissas na página 'Premissas Investimentos'.")
            return
        
        # Get the investments data
        investimentos_iniciais = st.session_state['premissas_investimentos']['investimentos_iniciais']
        investimentos_socios = st.session_state['premissas_investimentos'].get('investimentos_socios', [])
        investimentos_futuros = st.session_state['premissas_investimentos'].get('investimentos_futuros', [])
        
        if not investimentos_iniciais and not investimentos_socios and not investimentos_futuros:
            st.warning("Nenhum investimento configurado. Utilize a página 'Premissas Investimentos' para adicionar itens.")
            return
        
        # Gerar o DataFrame de investimentos
        df_investimentos = self._gerar_dataframe_investimentos()
        
        # Opções de visualização
        col1, col2 = st.columns(2)
        
        with col1:
            time_frame = st.selectbox("Período", self.time_frames, index=0)
        
        with col2:
            plot_type = st.selectbox("Tipo de Gráfico", self.graph_types, index=0)
        
        # Agrupar dados conforme o período selecionado
        df_display = self._agrupar_por_periodo(df_investimentos, time_frame)
        
        # Mostrar o dataframe
        st.write("### Fluxo de Investimentos")
        st.dataframe(df_display.style.format("{:.2f}"), use_container_width=True)
        
        # Seleção de categorias para visualização
        categorias = df_display.index.tolist()
        
        # Seleção múltipla de categorias
        selected_categories = st.multiselect(
            "Selecione as categorias para visualizar", 
            categorias,
            default=["Total"]
        )
        
        # Se nenhuma categoria foi selecionada, mostrar o total
        if not selected_categories:
            selected_categories = ["Total"]
        
        # Visualização gráfica
        st.write("### Visualização Gráfica")
        
        # Preparar dados para o gráfico
        df_plot = df_display.loc[selected_categories].T  # Transpor para ter períodos como índice
        
        # Para gráfico de pizza, o eixo X será a categoria e o valor será a soma total
        if plot_type == "Gráfico de Pizza":
            # Calcular a soma total de cada categoria (valor absoluto para gráfico de pizza)
            data_pie = {
                'Categoria': selected_categories,
                'Valor': [abs(df_display.loc[cat].sum()) for cat in selected_categories]
            }
            df_pie = pd.DataFrame(data_pie)
            
            fig = px.pie(
                df_pie,
                values='Valor',
                names='Categoria',
                title=f"Distribuição de Investimentos - Total ({time_frame})"
            )
            st.plotly_chart(fig, use_container_width=True)
            
        elif plot_type == "Gráfico de Barras":
            fig = px.bar(
                df_plot,
                title=f"Fluxo de Investimentos ({time_frame})"
            )
            fig.update_layout(
                xaxis_title="Período",
                yaxis_title="Valor (R$)",
                legend_title="Categoria"
            )
            st.plotly_chart(fig, use_container_width=True)
            
        else:  # Gráfico de Linhas
            fig = px.line(
                df_plot,
                title=f"Fluxo de Investimentos ({time_frame})",
                markers=True
            )
            fig.update_layout(
                xaxis_title="Período",
                yaxis_title="Valor (R$)",
                legend_title="Categoria"
            )
            st.plotly_chart(fig, use_container_width=True)

        # Mostrar detalhes do investimento inicial
        if investimentos_iniciais:
            with st.expander("Detalhes do Investimento Inicial"):
                st.write("#### Itens do Investimento Inicial")
                
                # Create DataFrame for display
                data = []
                for item in investimentos_iniciais:
                    total = item['quantidade'] * item['valor_unitario']
                    data.append({
                        "Descrição": item['descricao'],
                        "Quantidade": item['quantidade'],
                        "Valor Unitário": item['valor_unitario'],
                        "Total": total
                    })
                
                df = pd.DataFrame(data)
                
                # Calculate total investment
                total_investment = df["Total"].sum()
                
                # Display the DataFrame
                st.dataframe(df.style.format({
                    "Valor Unitário": "R$ {:.2f}",
                    "Total": "R$ {:.2f}"
                }), use_container_width=True)
                
                # Display the total
                st.metric("Investimento Inicial Total", f"R$ {total_investment:,.2f}")

class PremissasComissao(Page):
    def __init__(self):
        # Default values for commission parameters
        default_params = {
            'cargos_comissao': []
        }
        
        # Initialize or update session state
        if 'premissas_comissao' not in st.session_state:
            st.session_state['premissas_comissao'] = default_params
        else:
            # Check for any missing keys and add them with default values
            for key, value in default_params.items():
                if key not in st.session_state['premissas_comissao']:
                    st.session_state['premissas_comissao'][key] = value        
    
    @property
    def title(self) -> str:
        return "Premissas Comissões"
    
    @property
    def icon(self) -> str:
        return "💼"
    
    def render(self):
        styled_title(f"{self.icon} Sistema de Comissões")
        
        st.write("### Adicionar Comissão Por Cargo")     
        
        with st.form("novo_cargo_comissao"):
            # Get the cargos marked as "Sujeito ao Regime de Comissões" from PremissasDespesas
            cargos_comissoes = []
            if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
                cargos_comissoes = [
                    cargo['nome'] 
                    for cargo in st.session_state['premissas_despesas']['equipe_propria'] 
                    if cargo.get('sujeito_comissoes', False)
                ]
        
            # If no cargos are marked for commissions, show a text input
            if not cargos_comissoes:
                nome_cargo = st.text_input("Nome do Cargo", key="cargo_nome")
                st.info("Nenhum cargo marcado como 'Sujeito ao Regime de Comissões'. Configure-os na página 'Premissas Despesas'.")
            else:
                nome_cargo = st.selectbox(
                    "Nome do Cargo",
                    options=cargos_comissoes,
                    key="cargo_nome"
                )
            
            niveis = st.multiselect(
                "Níveis do Cargo",
                options=["A", "B", "C", "D", "E"],
                default=["A"],
                help="Selecione os níveis de carreira para este cargo"
            )
            
            custo_unitario = st.number_input(
                "Custo Unitário Mercadoria / Serviço Vendido (R$)",
                min_value=0.0,
                value=100.0,
                step=10.0,
                format="%.2f",
                help="Valor unitário do produto ou serviço vendido"
            )
            
            meta_em_numero = st.checkbox(
                "Meta em número de Vendas/mês",
                value=True,
                help="Se marcado, a meta será em número de vendas. Se desmarcado, a meta será em valor (R$). " +
                     "No caso de meta em número, o valor será somado diretamente a 'Numero de Vendas Líquidas'. " +
                     "No caso de meta em valor (R$), o sistema calculará o número equivalente de vendas dividindo " +
                     "pelo custo unitário para adicionar ao 'Numero de Vendas Líquidas'."
            )
            
            # Create inputs for each selected level
            nivel_inputs = {}
            for nivel in niveis:
                st.write(f"#### Parâmetros para Nível {nivel}")
                col1, col2 = st.columns(2)
                with col1:
                    inicial_comissao = st.number_input(
                        f"Percentual Inicial Comissão - Level {nivel} (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=0.5,
                        step=0.1,
                        format="%.2f",
                        key=f"inicial_comissao_{nivel}"
                    )
                    
                    inicial_meta = st.number_input(
                        f"Percentual Inicial da Meta - Level {nivel} (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=80.0,
                        step=1.0,
                        format="%.2f",
                        key=f"inicial_meta_{nivel}"
                    )
                
                with col2:
                    final_comissao = st.number_input(
                        f"Percentual Final Comissão - Level {nivel} (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=5.0,
                        step=0.1,
                        format="%.2f",
                        key=f"final_comissao_{nivel}"
                    )
                    
                    final_meta = st.number_input(
                        f"Percentual Final da Meta - Level {nivel} (%)",
                        min_value=0.0,
                        max_value=200.0,
                        value=120.0,
                        step=1.0,
                        format="%.2f",
                        key=f"final_meta_{nivel}"
                    )
                    
                quantidade = st.number_input(
                    f"Quantidade de {nome_cargo} no Level {nivel}",
                    min_value=0,
                    value=1,
                    step=1,
                    key=f"quantidade_{nivel}"
                )
                
                nivel_inputs[nivel] = {
                    'inicial_comissao': inicial_comissao,
                    'inicial_meta': inicial_meta,
                    'final_comissao': final_comissao,
                    'final_meta': final_meta,
                    'quantidade': quantidade
                }
            
            segmentos = st.selectbox(
                "Segmentos de Comissão",
                options=list(range(1, 11)),
                index=2,
                help="Número de segmentos de comissão entre o percentual inicial e final"
            )
            
            crescimento_meta = st.selectbox(
                "Crescimento da Meta",
                options=["Linear", "Exponencial", "Fixa"],
                index=0,
                help="Tipo de crescimento da meta ao longo do tempo"
            )
            
            if crescimento_meta == "Linear":
                taxa_crescimento = st.number_input(
                    "Tx. de crescimento linear (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=5.0,
                    step=0.5,
                    format="%.2f"
                )
                parametros_crescimento = {"tipo": "Linear", "taxa": taxa_crescimento}
            elif crescimento_meta == "Exponencial":
                taxa_inicial = st.number_input(
                    "Taxa inicial de crescimento exponencial (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=2.0,
                    step=0.5,
                    format="%.2f"
                )
                taxa_final = st.number_input(
                    "Taxa final de crescimento exponencial (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=10.0,
                    step=0.5,
                    format="%.2f"
                )
                parametros_crescimento = {"tipo": "Exponencial", "taxa_inicial": taxa_inicial, "taxa_final": taxa_final}
            else:  # Fixa
                taxas = {}
                for i in range(segmentos if segmentos is not None else 0):
                    taxas[i+1] = st.number_input(
                        f"Taxa para segmento {i+1} (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=float(i+1),
                        step=0.5,
                        format="%.2f"
                    )
                parametros_crescimento = {"tipo": "Fixa", "taxas": taxas}
            
            aleatorio = st.checkbox(
                "Aleatório",
                value=False,
                help="Se marcado, a quantidade de profissionais em cada nível mudará aleatoriamente periodicamente"
            )
            
            if aleatorio:
                meses_mudanca = st.number_input(
                    "Intervalo de meses para mudança aleatória",
                    min_value=1,
                    value=3,
                    step=1,
                    help="A cada quantos meses a quantidade de profissionais deve mudar aleatoriamente"
                )
                parametros_aleatorio = {"ativo": True, "intervalo_meses": meses_mudanca}
            else:
                parametros_aleatorio = {"ativo": False}
            
            submitted = st.form_submit_button("Adicionar Cargo")
            
            if submitted and nome_cargo:
                novo_cargo = {
                    'nome': nome_cargo,
                    'niveis': niveis,
                    'custo_unitario': custo_unitario,
                    'meta_em_numero': meta_em_numero,
                    'nivel_inputs': nivel_inputs,
                    'segmentos': segmentos,
                    'parametros_crescimento': parametros_crescimento,
                    'parametros_aleatorio': parametros_aleatorio
                }
                
                # Add new cargo to the list
                st.session_state['premissas_comissao']['cargos_comissao'].append(novo_cargo)
                st.success(f"Cargo '{nome_cargo}' adicionado com sucesso!")
                st.rerun()
        
        # Display existing cargos
        if st.session_state['premissas_comissao']['cargos_comissao']:
            st.write("### Cargos e Comissões Configurados")
            
            for idx, cargo in enumerate(st.session_state['premissas_comissao']['cargos_comissao']):
                with st.expander(f"{cargo['nome']} - {', '.join(cargo['niveis'])} Níveis"):
                    st.write(f"**Custo Unitário:** R$ {cargo['custo_unitario']:.2f}")
                    st.write(f"**Meta em número de vendas:** {'Sim' if cargo['meta_em_numero'] else 'Não'}")
                    st.write(f"**Segmentos de comissão:** {cargo['segmentos']}")
                    st.write(f"**Crescimento da meta:** {cargo['parametros_crescimento']['tipo']}")
                    st.write(f"**Mudança aleatória:** {'Sim' if cargo['parametros_aleatorio']['ativo'] else 'Não'}")
                    
                    # Display table of levels
                    data = []
                    for nivel, params in cargo['nivel_inputs'].items():
                        data.append({
                            "Nível": nivel,
                            "Qtd": params['quantidade'],
                            "% Inicial Meta": f"{params['inicial_meta']:.2f}%",
                            "% Inicial Comissão": f"{params['inicial_comissao']:.2f}%",
                            "% Final Meta": f"{params['final_meta']:.2f}%",
                            "% Final Comissão": f"{params['final_comissao']:.2f}%"
                        })
                    
                    df_levels = pd.DataFrame(data)
                    st.dataframe(df_levels, use_container_width=True)
                    
                    if st.button(f"Remover {cargo['nome']}", key=f"remove_{idx}"):
                        st.session_state['premissas_comissao']['cargos_comissao'].pop(idx)
                        st.success(f"Cargo '{cargo['nome']}' removido com sucesso!")
                        st.rerun()
            

            # Visual representation of commission system
            st.write("### Visualização do Sistema de Comissões")
            
            cargo_selecionado = st.selectbox(
                "Selecione um cargo para visualizar",
                options=[cargo['nome'] for cargo in st.session_state['premissas_comissao']['cargos_comissao']]
            )
            
            # Find the selected cargo
            cargo = next((c for c in st.session_state['premissas_comissao']['cargos_comissao'] if c['nome'] == cargo_selecionado), None)
            
            if cargo:
                nivel_selecionado = st.selectbox(
                    "Selecione um nível",
                    options=cargo['niveis']
                )
                
                if nivel_selecionado in cargo['nivel_inputs']:
                    params = cargo['nivel_inputs'][nivel_selecionado]
                    
                    # Create data for the chart
                    segmentos = cargo['segmentos']
                    metas = np.linspace(params['inicial_meta'], params['final_meta'], segmentos+1)
                    comissoes = np.linspace(params['inicial_comissao'], params['final_comissao'], segmentos+1)
                    
                    df_chart = pd.DataFrame({
                        "Meta (%)": metas,
                        "Comissão (%)": comissoes
                    })
                    
                    fig = px.line(
                        df_chart, 
                        x="Meta (%)", 
                        y="Comissão (%)", 
                        title=f"Comissões para {cargo['nome']} - Nível {nivel_selecionado}",
                        markers=True
                    )
                    
                    st.plotly_chart(fig, use_container_width=True)

class ComissaoVendas(Page):
    def __init__(self):
        pass
    
    @property
    def title(self) -> str:
        return "Comissão Vendas"
    
    @property
    def icon(self) -> str:
        return "💰"
    
    def _calcular_comissoes(self):
        """Calcula as comissões mensais para cada cargo com base nas premissas"""
        if 'premissas_comissao' not in st.session_state or not st.session_state['premissas_comissao']['cargos_comissao']:
            return None
            
        # Get data from Receitas class instead of GeracaoDeCaixa
        try:
            receitas = Receitas()
            df_faturamento = receitas._gerar_dataframe_faturamento()
            df_arrecadacao = receitas._gerar_dataframe_arrecadacao(df_faturamento)
            
            if df_arrecadacao is None or df_arrecadacao.empty:
                return None
            
            # Extract necessary data - use Arrecadação Líquida as our primary metric
            arrecadacao_liquida = df_arrecadacao.loc["Arrecadação Líquida"] if "Arrecadação Líquida" in df_arrecadacao.index else pd.Series(0, index=range(61))
            
            # For number of sales, we'll estimate from the faturamento if available
            if df_faturamento is not None and "Total" in df_faturamento.index:
                faturamento = df_faturamento.loc["Total"]
                # Estimate number of sales if we have a ticket médio in premissas
                ticket_medio = st.session_state.get('premissas_receitas', {}).get('ticket_medio', 2400.0)
                vendas_liquidas = faturamento / ticket_medio if ticket_medio > 0 else pd.Series(0, index=faturamento.index)
            else:
                faturamento = pd.Series(0, index=arrecadacao_liquida.index)
                vendas_liquidas = pd.Series(0, index=arrecadacao_liquida.index)
            
        except Exception as e:
            st.warning(f"Erro ao obter dados de arrecadação: {e}")
            # Create empty series with 61 months
            arrecadacao_liquida = pd.Series(0, index=range(61))
            faturamento = pd.Series(0, index=range(61))
            vendas_liquidas = pd.Series(0, index=range(61))
        
        # Create DataFrame to store commissions
        df_comissoes = pd.DataFrame(index=range(61))  # 0 to 60 months
        
        # Calculate commissions for each cargo
        for cargo in st.session_state['premissas_comissao']['cargos_comissao']:
            # For each level in the cargo
            for nivel, params in cargo['nivel_inputs'].items():
                
                # Initialize column for this cargo and level
                col_name = f"{cargo['nome']} - {nivel}"
                df_comissoes[col_name] = 0.0
                
                # Determine if quantities change randomly over time 
                quantidade_base = params['quantidade']
                aleatorio = cargo['parametros_aleatorio']['ativo']
                intervalo_meses = cargo['parametros_aleatorio'].get('intervalo_meses', 1) if aleatorio else 0
                
                # Calculate parameters for each month
                for mes in df_comissoes.index:

                    # Determine quantity for this month
                    if aleatorio and mes > 0 and mes % intervalo_meses == 0:
                        # Random change in quantity (±50%)
                        quantidade = max(0, int(quantidade_base * (0.5 + random.random())))
                    else:
                        quantidade = quantidade_base
                    
                    # Get base values to calculate commission
                    meta_em_numero = cargo['meta_em_numero']
                    
                    # Apply growth to meta based on growth type
                    crescimento = cargo['parametros_crescimento']
                    crescimento_mes = 1.0
                    
                    if crescimento['tipo'] == 'Linear':
                        # Linear growth: meta increases by X% each month
                        crescimento_mes = 1.0 + (mes * crescimento['taxa'] / 100)
                    elif crescimento['tipo'] == 'Exponencial':
                        # Exponential growth: starts with initial rate, grows to final rate
                        taxa_inicial = crescimento['taxa_inicial'] / 100
                        taxa_final = crescimento['taxa_final'] / 100
                        taxa_mes = taxa_inicial + (taxa_final - taxa_inicial) * (mes / 60)
                        crescimento_mes = (1 + taxa_mes) ** mes
                    elif crescimento['tipo'] == 'Fixa':
                        # Fixed rates for each segment
                        taxas = crescimento['taxas']
                        # Use the tax for the closest segment (based on month)
                        segmento = min(len(taxas), max(1, int(mes * len(taxas) / 60) + 1))
                        crescimento_mes = 1.0 + (taxas[segmento] / 100 * mes)
                    
                    # Calculate meta for this month
                    if meta_em_numero:
                        # Meta is in number of sales
                        meta_base = quantidade  # Each person should sell at least 1 unit
                    else:
                        # Meta is in monetary value
                        meta_base = quantidade * cargo['custo_unitario'] * 10  # Arbitrary multiple of unit cost
                    
                    meta = meta_base * crescimento_mes
                    
                    # Get actual performance
                    if meta_em_numero:
                        # For month 0, use the current month's data; otherwise use previous month
                        index_to_use = mes if mes == 0 else mes - 1
                        performance = vendas_liquidas.loc[index_to_use] if index_to_use in vendas_liquidas.index else 0
                    else:
                        # For month 0, use the current month's data; otherwise use previous month
                        index_to_use = mes if mes == 0 else mes - 1
                        performance = faturamento.loc[index_to_use] if index_to_use in faturamento.index else 0
                    
                    # Calculate performance percentage relative to meta
                    performance_pct = (performance / meta * 100) if meta > 0 else 0
                    
                    # Map performance to commission percentage using segments
                    inicial_meta = params['inicial_meta']
                    final_meta = params['final_meta']
                    inicial_comissao = params['inicial_comissao']
                    final_comissao = params['final_comissao']
                    
                    # No commission if below initial meta percentage
                    if performance_pct < inicial_meta:
                        comissao_pct = 0
                    # Max commission if above final meta percentage
                    elif performance_pct >= final_meta:
                        comissao_pct = final_comissao
                    # Linear interpolation between initial and final commission
                    else:
                        # Calculate which segment performance falls into
                        segmentos = cargo['segmentos']
                        segmento_tamanho = (final_meta - inicial_meta) / segmentos
                        
                        for i in range(segmentos):
                            segmento_min = inicial_meta + i * segmento_tamanho
                            segmento_max = inicial_meta + (i + 1) * segmento_tamanho
                            
                            if segmento_min <= performance_pct < segmento_max:
                                # Calculate commission percentage within this segment
                                segmento_comissao_min = inicial_comissao + i * (final_comissao - inicial_comissao) / segmentos
                                segmento_comissao_max = inicial_comissao + (i + 1) * (final_comissao - inicial_comissao) / segmentos
                                
                                # Linear interpolation
                                comissao_pct = segmento_comissao_min + (performance_pct - segmento_min) * \
                                              (segmento_comissao_max - segmento_comissao_min) / segmento_tamanho
                                break
                        else:
                            # Fallback if no segment was found (shouldn't happen)
                            comissao_pct = 0
                    
                    # Calculate commission value
                    if meta_em_numero:
                        comissao_valor = performance * cargo['custo_unitario'] * (comissao_pct / 100) * quantidade
                    else:
                        comissao_valor = performance * (comissao_pct / 100) * quantidade
                    
                    # Store in DataFrame (negative value because it's an expense)
                    df_comissoes.loc[mes, col_name] = -comissao_valor
        
        # Add Total column
        df_comissoes['Total'] = df_comissoes.sum(axis=1)
        
        return df_comissoes
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        if 'premissas_comissao' not in st.session_state or not st.session_state['premissas_comissao']['cargos_comissao']:
            st.warning("Nenhuma comissão configurada. Por favor, defina as premissas na página 'Premissas Comissões'.")
            return
        
        # Calculate commissions
        df_comissoes = self._calcular_comissoes()
        
        if df_comissoes is None or df_comissoes.empty:
            st.error("Não foi possível calcular as comissões. Verifique se há dados de vendas disponíveis.")
            return
        
        # Store in session state for use in GeracaoDeCaixa
        st.session_state['comissoes_calculadas'] = df_comissoes['Total']
        
        # Display commissions DataFrame
        st.write("### Comissões Calculadas")
        st.dataframe(df_comissoes.style.format("{:.2f}"), use_container_width=True)
        
        # Visualization options
        st.write("### Visualização Gráfica")
        
        col1, col2 = st.columns(2)
        
        with col1:
            time_frame = st.selectbox(
                "Período",
                options=["Mensal", "Trimestral", "Anual"],
                index=0
            )
        
        with col2:
            plot_type = st.selectbox(
                "Tipo de Gráfico",
                options=["Linha", "Barra", "Área"],
                index=0
            )
        
        # Process data based on selected time frame
        df_plot = df_comissoes.copy()
        
        if time_frame == "Trimestral":
            # Group by quarter
            df_grouped = pd.DataFrame()
            for i in range(0, 60, 3):
                quarter = f"T{i//3 + 1}"
                df_grouped[quarter] = df_plot.iloc[i:i+3].mean()
            df_plot = df_grouped
        elif time_frame == "Anual":
            # Group by year
            df_grouped = pd.DataFrame()
            for i in range(0, 60, 12):
                year = f"Ano {i//12 + 1}"
                df_grouped[year] = df_plot.iloc[i:i+12].mean()
            df_plot = df_grouped
        
        # Create plot
        if "Total" in df_plot.columns:
            # Remove 'Total' from the columns for plotting
            plot_cols = [col for col in df_plot.columns if col != 'Total']
            
            # Only plot if there are columns to plot
            if plot_cols:
                # Transpose for plotting
                df_plot_t = df_plot[plot_cols].T
                
                if plot_type == "Linha":
                    fig = px.line(
                        df_plot_t,
                        title=f"Comissões por Cargo - {time_frame}",
                        labels={"value": "Valor (R$)", "variable": "Período", "index": "Cargo"}
                    )
                elif plot_type == "Barra":
                    fig = px.bar(
                        df_plot_t,
                        title=f"Comissões por Cargo - {time_frame}",
                        labels={"value": "Valor (R$)", "variable": "Período", "index": "Cargo"}
                    )
                else:  # Área
                    fig = px.area(
                        df_plot_t,
                        title=f"Comissões por Cargo - {time_frame}",
                        labels={"value": "Valor (R$)", "variable": "Período", "index": "Cargo"}
                    )
                
                st.plotly_chart(fig, use_container_width=True)
        
        # Total commission metrics
        st.write("### Métricas de Comissão")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            comissao_total = df_comissoes['Total'].sum()
            st.metric("Comissão Total (60 meses)", f"R$ {abs(comissao_total):,.2f}")
        
        with col2:
            comissao_media = df_comissoes['Total'].mean()
            st.metric("Comissão Média Mensal", f"R$ {abs(comissao_media):,.2f}")
        
        with col3:
            # Get total revenue if available
            try:
                receitas = Receitas()
                df_faturamento = receitas._gerar_dataframe_faturamento()
                df_arrecadacao = receitas._gerar_dataframe_arrecadacao(df_faturamento)
                
                if df_arrecadacao is not None and "Arrecadação Líquida" in df_arrecadacao.index:
                    arrecadacao_total = df_arrecadacao.loc["Arrecadação Líquida"].sum()
                    
                    if isinstance(arrecadacao_total, pd.Series):
                        if arrecadacao_total.sum() > 0:
                            percentual_comissao = abs(comissao_total) / arrecadacao_total.sum() * 100
                            st.metric("% do Faturamento", f"{percentual_comissao:.2f}%")
                        else:
                            st.metric("% do Faturamento", "N/A")
                    else:
                        if arrecadacao_total > 0:
                            percentual_comissao = abs(comissao_total) / arrecadacao_total * 100
                            st.metric("% do Faturamento", f"{percentual_comissao:.2f}%")
                        else:
                            st.metric("% do Faturamento", "N/A")
                else:
                    st.metric("% do Faturamento", "N/A")
            except Exception as e:
                st.warning(f"Erro ao obter dados de arrecadação: {e}")
                st.metric("% do Faturamento", "N/A")

class PremissasDespesas(Page):
    def __init__(self):
        # Default values for all expense parameters
        default_params = {
            # Parâmetros gerais
            'ipca_medio_anual': 4.5,
            'igpm_medio_anual': 5.0,  # Added IGP-M parameter
            'modo_calculo': 'Percentual',  # 'Percentual' ou 'Nominal'
            'budget_mensal': 30000.0,  # Orçamento mensal total para o modo percentual
            
            # Água e Luz
            'modo_energia': 'Constante',  # 'Constante', 'Estressado' ou 'Extremamente Conservador'
            'consumo_mensal_kwh': 2000.0,
            
            # Rest of the existing parameters remain unchanged
            'aluguel': 8000.0,
            'condominio': 1500.0,
            'iptu': 1000.0,
            
            # Outras despesas
            'marketing_publicidade': 5000.0,
            'internet': 350.0,
            'material_escritorio': 800.0,
            'treinamentos': 2000.0,
            'manutencao_conservacao': 1200.0,
            'seguros_funcionarios': 2000.0,
            'licencas_telefonia': 500.0,
            'licencas_crm': 1000.0,
            'telefonica': 500.0,
            
            # Percentuais para modo percentual
            'perc_agua_luz': 5.0,
            'perc_aluguel_condominio_iptu': 35.0,
            'perc_marketing_publicidade': 16.7,
            'perc_internet': 1.2,
            'perc_material_escritorio': 2.7,
            'perc_treinamentos': 6.7,
            'perc_manutencao_conservacao': 4.0,
            'perc_seguros_funcionarios': 6.7,
            'perc_licencas_telefonia': 1.7,
            'perc_licencas_crm': 3.3,
            'perc_telefonica': 1.7,

            # Parâmetros da Equipe
            # Equipe Própria
            'ceo_salario': 15000.0,
            'ceo_qtd': 1,
            'cfo_salario': 12000.0,
            'cfo_qtd': 1,
            'head_vendas_salario': 8000.0,
            'head_vendas_qtd': 1,
            'sdr_salario': 2800.0,
            'sdr_qtd': 3,
            'closer_salario': 3500.0,
            'closer_qtd': 2,
            'social_media_salario': 2500.0,
            'social_media_qtd': 1,
            'ti_salario': 5000.0,
            'ti_qtd': 1,
            'closer_taxa_cancelamento': 5.0,
            
            # Encargos Sociais
            'encargos_sociais_perc': 68.0,  # Percentual de encargos sobre a folha
            
            # Despesas com Alimentação e Transporte
            'vale_alimentacao': 30.0,  # Valor diário
            'vale_transporte': 12.0,   # Valor diário
            'roles_com_beneficios': ['ceo', 'cfo', 'head_vendas', 'sdr', 'closer', 'social_media', 'ti'],
            'terceiros': [],
            
            # # Terceiros - Prestadores de Serviços
            'agencia_trafego_valor': 0.0,
            'agencia_trafego_qtd': 0,
            'contabilidade_valor': 0.0,
            'contabilidade_qtd': 0,
            'juridico_valor': 0.0,
            'juridico_qtd': 0,
            'agencia_criativos_valor': 0.0,
            'agencia_criativos_qtd': 0, 
            
            # Bônus dos Lucros
            'benchmark_anual_bonus': 10.0,  # Percentual de crescimento do lucro líquido para bônus
            'percentual_bonus_ceo': 5.0,    # Percentual do valor excedente
            'percentual_bonus_cfo': 3.0,    # Percentual do valor excedente
            'percentual_bonus_head_vendas': 2.0,  # Percentual do valor excedente
            
            # Dummy para lucro líquido (temporário)
            'lucro_liquido_inicial': 100000.0,  # Lucro líquido inicial
            'crescimento_lucro': 15.0,  # Percentual de crescimento anual do lucro líquido

            # Custos de Tecnologia
            'desenvolvimento_ferramenta': 0.0,
            'manutencao_ferramenta': 0.0,
            'inovacao': 0.0,
            'licencas_software': 2513.0,
                
        }
            
        # Initialize or update session state
        if 'premissas_despesas' not in st.session_state:
            st.session_state['premissas_despesas'] = default_params
        else:
            # Check for any missing keys and add them with default values
            for key, value in default_params.items():
                if key not in st.session_state['premissas_despesas']:
                    st.session_state['premissas_despesas'][key] = value
    
    @property
    def title(self) -> str:
        return "Premissas Despesas"
    
    @property
    def icon(self) -> str:
        return "📝"
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Create tabs for better organization
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["Configurações Gerais", "Despesas Administrativas", "Equipe", "Tecnologia", "Reajustes"])
        
        with tab1:
            st.write("### Configurações Gerais")
            st.write("Defina os parâmetros gerais para cálculo das despesas.")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Replace number_input with slider for IPCA
                st.session_state['premissas_despesas']['ipca_medio_anual'] = st.slider(
                    "IPCA Médio Anual (%)",
                    min_value=0.0,
                    max_value=20.0,
                    value=st.session_state['premissas_despesas']['ipca_medio_anual'],
                    step=0.1,
                    format="%.1f",
                    help="Taxa média de inflação anual utilizada para correção monetária."
                )
                
                st.session_state['premissas_despesas']['modo_calculo'] = st.selectbox(
                    "Modo de Cálculo",
                    ["Percentual", "Nominal"],
                    index=0 if st.session_state['premissas_despesas']['modo_calculo'] == "Percentual" else 1,
                    help="Escolha entre cálculo percentual (baseado no orçamento mensal total, utilizando % para cada despesa) ou nominal (valores fixos em R$ para cada despesa)."
                )
                
                # Show budget slider only if Percentual is selected
                if st.session_state['premissas_despesas']['modo_calculo'] == "Percentual":
                    st.session_state['premissas_despesas']['budget_mensal'] = st.slider(
                        "Orçamento Mensal Total (R$)",
                        min_value=1000.0,
                        max_value=100000.0,
                        value=st.session_state['premissas_despesas']['budget_mensal'],
                        step=1000.0,
                        format="%.2f",
                        help="Orçamento mensal total disponível para despesas administrativas. Este valor será usado como base para calcular as porcentagens de cada despesa."
                    )
            
            with col2:
                # Add IGP-M slider
                st.session_state['premissas_despesas']['igpm_medio_anual'] = st.slider(
                    "IGP-M Médio Anual (%)",
                    min_value=0.0,
                    max_value=20.0,
                    value=st.session_state['premissas_despesas']['igpm_medio_anual'],
                    step=0.1,
                    format="%.1f",
                    help="Taxa média do IGP-M anual utilizada para correção monetária."
                )
                
                # Update selectbox to include the new extreme option
                st.session_state['premissas_despesas']['modo_energia'] = st.selectbox(
                    "Modo de Cálculo de Energia",
                    ["Constante", "Estressado", "Extremamente Conservador"],
                    index=0 if st.session_state['premissas_despesas']['modo_energia'] == "Constante" 
                          else 1 if st.session_state['premissas_despesas']['modo_energia'] == "Estressado"
                          else 2,
                    help="Escolha entre cálculo constante (valores fixos para consumo de energia), estressado (valores variam com o tempo, incidindo de maneira randômica, bandeiras tarifárias), ou extremamente conservador (bandeira vermelha em todos os meses)."
                )

        # In PremissasDespesas.render, inside tab2 ("Despesas Administrativas")
        
        with tab2:
            st.write("### Despesas Administrativas")
            st.write("Defina os valores para cada tipo de despesa administrativa.")
            
            # Add slider for "Mês de Início"
            st.write("#### Mês de Início das Despesas")
            if 'mes_inicio_despesas' not in st.session_state['premissas_despesas']:
                st.session_state['premissas_despesas']['mes_inicio_despesas'] = 0
                
            st.session_state['premissas_despesas']['mes_inicio_despesas'] = st.slider(
                "Mês de Início",
                min_value=0,
                max_value=24,
                value=st.session_state['premissas_despesas']['mes_inicio_despesas'],
                step=1,
                help="Mês a partir do qual as despesas administrativas serão aplicadas. Antes desse mês, os valores serão zero."
            )
            
            # Get calculation mode from tab1
            modo_calculo = st.session_state['premissas_despesas']['modo_calculo']
            
            # Rest of the existing code remains unchanged
            # Água e Luz section
            st.write("#### Água e Luz")
            if modo_calculo == "Nominal":
                st.session_state['premissas_despesas']['consumo_mensal_kwh'] = st.slider(
                    "Consumo Mensal (kWh)",
                    min_value=0.0,
                    max_value=10000.0,
                    value=st.session_state['premissas_despesas']['consumo_mensal_kwh'],
                    step=100.0,
                    format="%.2f",
                    help="Consumo mensal de energia elétrica em kWh"
                )
            else:  # Percentual mode
                st.session_state['premissas_despesas']['perc_agua_luz'] = st.slider(
                    "Água e Luz (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=st.session_state['premissas_despesas']['perc_agua_luz'],
                    step=0.1,
                    format="%.1f",
                    help="Percentual do orçamento mensal total destinado a água e luz."
                )
            
            # Aluguéis, Condomínios e IPTU
            st.write("#### Aluguéis, Condomínios e IPTU")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if modo_calculo == "Nominal":
                    st.session_state['premissas_despesas']['aluguel'] = st.slider(
                        "Aluguel (R$)",
                        min_value=0.0,
                        max_value=50000.0,
                        value=st.session_state['premissas_despesas']['aluguel'],
                        step=100.0,
                        format="%.2f",
                        help="Valor mensal do aluguel do imóvel onde a empresa opera."
                    )
                else:  # No need to show individual percentages for Aluguel in Percentual mode
                    st.info("No modo percentual, use o valor total para Aluguéis, Condomínios e IPTU abaixo.")
            
            with col2:
                if modo_calculo == "Nominal":
                    st.session_state['premissas_despesas']['condominio'] = st.slider(
                        "Condomínio (R$)",
                        min_value=0.0,
                        max_value=10000.0,
                        value=st.session_state['premissas_despesas']['condominio'],
                        step=100.0,
                        format="%.2f",
                        help="Valor mensal do condomínio."
                    )
            
            with col3:
                if modo_calculo == "Nominal":
                    st.session_state['premissas_despesas']['iptu'] = st.slider(
                        "IPTU (R$)",
                        min_value=0.0,
                        max_value=10000.0,
                        value=st.session_state['premissas_despesas']['iptu'],
                        step=100.0,
                        format="%.2f",
                        help="Valor mensal do IPTU (Imposto Predial e Territorial Urbano)."
                    )
            
            # Display the percentage slider for Aluguel/Condominio/IPTU when in Percentual mode
            if modo_calculo == "Percentual":
                st.session_state['premissas_despesas']['perc_aluguel_condominio_iptu'] = st.slider(
                    "Aluguéis, Condomínios e IPTU (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=st.session_state['premissas_despesas']['perc_aluguel_condominio_iptu'],
                    step=0.1,
                    format="%.1f",
                    help="Percentual do orçamento mensal total destinado a aluguéis, condomínios e IPTU."
                )
            
            # Outras despesas
            st.write("#### Outras Despesas")
            
            col1, col2 = st.columns(2)
            
            with col1:
                if modo_calculo == "Nominal":
                    # Nominal mode - left column sliders
                    st.session_state['premissas_despesas']['internet'] = st.slider(
                        "Internet (R$)",
                        min_value=0.0,
                        max_value=5000.0,
                        value=st.session_state['premissas_despesas']['internet'],
                        step=10.0,
                        format="%.2f",
                        help="Valor mensal da conexão de internet."
                    )
                    
                    st.session_state['premissas_despesas']['material_escritorio'] = st.slider(
                        "Material de Escritório (R$)",
                        min_value=0.0,
                        max_value=5000.0,
                        value=st.session_state['premissas_despesas']['material_escritorio'],
                        step=100.0,
                        format="%.2f",
                        help="Valor mensal gasto com material de escritório."
                    )
                    
                    st.session_state['premissas_despesas']['treinamentos'] = st.slider(
                        "Treinamentos (R$)",
                        min_value=0.0,
                        max_value=10000.0,
                        value=st.session_state['premissas_despesas']['treinamentos'],
                        step=100.0,
                        format="%.2f",
                        help="Valor mensal destinado a treinamentos e capacitações."
                    )
                    
                    st.session_state['premissas_despesas']['manutencao_conservacao'] = st.slider(
                        "Manutenção & Conservação (R$)",
                        min_value=0.0,
                        max_value=10000.0,
                        value=st.session_state['premissas_despesas']['manutencao_conservacao'],
                        step=100.0,
                        format="%.2f",
                        help="Valor mensal destinado à manutenção e conservação do espaço físico."
                    )
                else:
                    # Percentual mode - left column sliders
                    st.session_state['premissas_despesas']['perc_internet'] = st.slider(
                        "Internet (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=st.session_state['premissas_despesas']['perc_internet'],
                        step=0.1,
                        format="%.1f",
                        help="Percentual do orçamento mensal total destinado à internet."
                    )
                    
                    st.session_state['premissas_despesas']['perc_material_escritorio'] = st.slider(
                        "Material de Escritório (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=st.session_state['premissas_despesas']['perc_material_escritorio'],
                        step=0.1,
                        format="%.1f",
                        help="Percentual do orçamento mensal total destinado a material de escritório."
                    )
                    
                    st.session_state['premissas_despesas']['perc_treinamentos'] = st.slider(
                        "Treinamentos (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=st.session_state['premissas_despesas']['perc_treinamentos'],
                        step=0.1,
                        format="%.1f",
                        help="Percentual do orçamento mensal total destinado a treinamentos e capacitações."
                    )
                    
                    st.session_state['premissas_despesas']['perc_manutencao_conservacao'] = st.slider(
                        "Manutenção & Conservação (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=st.session_state['premissas_despesas']['perc_manutencao_conservacao'],
                        step=0.1,
                        format="%.1f",
                        help="Percentual do orçamento mensal total destinado à manutenção e conservação do espaço físico."
                    )
            
            with col2:
                if modo_calculo == "Nominal":
                    # Nominal mode - right column sliders
                    st.session_state['premissas_despesas']['seguros_funcionarios'] = st.slider(
                        "Seguros Funcionários (R$)",
                        min_value=0.0,
                        max_value=10000.0,
                        value=st.session_state['premissas_despesas']['seguros_funcionarios'],
                        step=100.0,
                        format="%.2f",
                        help="Valor mensal destinado a seguros para funcionários (saúde, acidentes, etc.)."
                    )
                    
                    st.session_state['premissas_despesas']['licencas_telefonia'] = st.slider(
                        "Licenças de Telefonia (R$)",
                        min_value=0.0,
                        max_value=5000.0,
                        value=st.session_state['premissas_despesas']['licencas_telefonia'],
                        step=50.0,
                        format="%.2f",
                        help="Valor mensal gasto com licenças de telefonia (como VoIP, etc.)."
                    )
                    
                    st.session_state['premissas_despesas']['licencas_crm'] = st.slider(
                        "Licenças CRM (R$)",
                        min_value=0.0,
                        max_value=10000.0,
                        value=st.session_state['premissas_despesas']['licencas_crm'],
                        step=100.0,
                        format="%.2f",
                        help="Valor mensal gasto com licenças de software CRM (Customer Relationship Management)."
                    )
                    
                    st.session_state['premissas_despesas']['telefonica'] = st.slider(
                        "Telefônica (R$)",
                        min_value=0.0,
                        max_value=5000.0,
                        value=st.session_state['premissas_despesas']['telefonica'],
                        step=50.0,
                        format="%.2f",
                        help="Valor mensal gasto com serviços de telefonia."
                    )
                else:
                    # Percentual mode - right column sliders
                    st.session_state['premissas_despesas']['perc_seguros_funcionarios'] = st.slider(
                        "Seguros Funcionários (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=st.session_state['premissas_despesas']['perc_seguros_funcionarios'],
                        step=0.1,
                        format="%.1f",
                        help="Percentual do orçamento mensal total destinado a seguros para funcionários (saúde, acidentes, etc.)."
                    )
                    
                    st.session_state['premissas_despesas']['perc_licencas_telefonia'] = st.slider(
                        "Licenças de Telefonia (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=st.session_state['premissas_despesas']['perc_licencas_telefonia'],
                        step=0.1,
                        format="%.1f",
                        help="Percentual do orçamento mensal total destinado a licenças de telefonia."
                    )
                    
                    st.session_state['premissas_despesas']['perc_licencas_crm'] = st.slider(
                        "Licenças CRM (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=st.session_state['premissas_despesas']['perc_licencas_crm'],
                        step=0.1,
                        format="%.1f",
                        help="Percentual do orçamento mensal total destinado a licenças de software CRM (Customer Relationship Management)."
                    )
                    
                    st.session_state['premissas_despesas']['perc_telefonica'] = st.slider(
                        "Telefônica (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=st.session_state['premissas_despesas']['perc_telefonica'],
                        step=0.1,
                        format="%.1f",
                        help="Percentual do orçamento mensal total destinado a serviços de telefonia."
                    )
            
            # Verificar se a soma dos percentuais é aproximadamente 100% no modo percentual
            if modo_calculo == "Percentual":
                soma_percentuais = (
                    st.session_state['premissas_despesas']['perc_agua_luz'] +
                    st.session_state['premissas_despesas']['perc_aluguel_condominio_iptu'] +
                    st.session_state['premissas_despesas']['perc_internet'] +
                    st.session_state['premissas_despesas']['perc_material_escritorio'] +
                    st.session_state['premissas_despesas']['perc_treinamentos'] +
                    st.session_state['premissas_despesas']['perc_manutencao_conservacao'] +
                    st.session_state['premissas_despesas']['perc_seguros_funcionarios'] +
                    st.session_state['premissas_despesas']['perc_licencas_telefonia'] +
                    st.session_state['premissas_despesas']['perc_licencas_crm'] +
                    st.session_state['premissas_despesas']['perc_telefonica']
                )
                
                st.write(f"Soma dos percentuais: {soma_percentuais:.1f}%")
                
                if abs(soma_percentuais - 100.0) > 0.1:
                    st.warning("A soma dos percentuais deve ser igual a 100%.")
        with tab3:
            st.write("### Configurações da Equipe")
            st.write("Defina os parâmetros relacionados à equipe da empresa.")
            
            # Add selection box for "Percentual" or "Nominal" mode
            equipe_modo_calculo = st.selectbox(
                "Modo de Cálculo da Equipe",
                ["Percentual", "Nominal"],
                index=0 if st.session_state['premissas_despesas'].get('equipe_modo_calculo', "Nominal") == "Percentual" else 1,
                help="Escolha entre cálculo percentual (baseado em orçamentos para cada tipo de equipe) ou nominal (valores fixos em R$ para cada cargo)."
            )
            
            # Store the selection in session state
            st.session_state['premissas_despesas']['equipe_modo_calculo'] = equipe_modo_calculo
            
            # If Percentual mode, show budget sliders
            if equipe_modo_calculo == "Percentual":
                col1, col2 = st.columns(2)
                
                with col1:
                    st.session_state['premissas_despesas']['budget_equipe_propria'] = st.slider(
                        "Budget Equipe Própria (R$)",
                        min_value=0.0,
                        max_value=200000.0,
                        value=st.session_state['premissas_despesas'].get('budget_equipe_propria', 50000.0),
                        step=1000.0,
                        format="%.2f",
                        help="Orçamento mensal total disponível para a equipe própria."
                    )
                    
                with col2:
                    st.session_state['premissas_despesas']['budget_terceiros'] = st.slider(
                        "Budget Terceiros - Prestadores de Serviços (R$)",
                        min_value=0.0,
                        max_value=100000.0,
                        value=st.session_state['premissas_despesas'].get('budget_terceiros', 10000.0),
                        step=1000.0,
                        format="%.2f",
                        help="Orçamento mensal total disponível para prestadores de serviços terceirizados."
                    )
            
            # Initialize session state for editing
            if 'editing_team_member' not in st.session_state:
                st.session_state['editing_team_member'] = None
            
            # Equipe Própria section
            st.write("#### Equipe Própria")
            
            # Initialize equipe_propria list if it doesn't exist
            if 'equipe_propria' not in st.session_state['premissas_despesas']:
                st.session_state['premissas_despesas']['equipe_propria'] = []
            
            # Expander for adding new roles
            with st.expander("Adicionar Novo Cargo - Equipe Própria"):
                with st.form("novo_cargo_equipe"):
                    nome_cargo = st.selectbox("Nome do Cargo", ["CEO", "CFO", "Head de Vendas", "SDR", "Closer", "Account Manager", "TI"])
                    
                    # Common fields for all roles except Closer
                    if nome_cargo != "Closer":
                        quantidade = st.number_input("Quantidade de profissionais neste cargo", min_value=0, value=1, step=1)
                    else:
                        # For Closer, we'll calculate the quantity based on appointments per Closer
                        quantidade = 1  # Default value, will be automatically calculated based on comparecimentos_por_closer
                    
                    # Initialize variables with default values to avoid unbound errors
                    taxa_agendamento = 0.0
                    taxa_comparecimento = 0.0
                    estimativa_leads = 0
                    periodicidade = "Mensal"
                    fator_aceleracao = 1.0
                    valor_unitario = 0.0
                    taxa_conversao = 0.0
                    produtos_por_lead = 10  # Default value
                    comparecimentos_por_closer = 90  # Default value
            
                    # Conditional fields based on role
                    if nome_cargo == "SDR":
                        salario = st.number_input("Salário (R$)", min_value=0.0, value=2800.0, step=100.0)
                        taxa_agendamento = st.slider("Taxa de Agendamento (%)", min_value=1.0, max_value=100.0, value=30.0, step=1.0)
                        taxa_comparecimento = st.slider("Taxa de Comparecimento (%)", min_value=1.0, max_value=100.0, value=70.0, step=1.0)
                        estimativa_leads = st.slider("Estimativa Inicial de Leads (mensal)", min_value=50, max_value=1000, value=200, step=50)
                    elif nome_cargo == "Closer":
                        salario = st.number_input("Salário (R$)", min_value=0.0, value=3500.0, step=100.0)
                        periodicidade = st.selectbox("Periodicidade de Aumento de Gasto", ["Mensal", "Trimestral", "Semestral", "Anual"])
                        fator_aceleracao = st.slider("Fator de Aceleração do Crescimento", min_value=0.1, max_value=5.0, value=1.0, step=0.1)
                        valor_unitario = st.number_input("Valor Unitário Médio dos Produtos/Serviços (R$)", min_value=10.0, value=2400.0, step=100.0)
                        taxa_conversao = st.slider("Taxa de Conversão (%)", min_value=1.0, max_value=100.0, value=45.0, step=1.0)
                        produtos_por_lead = st.slider("Número Médio de Produtos/Serviços por Lead", min_value=5, max_value=100, value=10, step=5)
                        comparecimentos_por_closer = st.number_input("Comparecimentos por Closer", min_value=10, max_value=500, value=90, step=10)
                        taxa_de_cancelamento = st.slider("Taxa de Cancelamento (%)", min_value=0.0, max_value=100.0, value=5.0, step=0.5)
                    else:
                        salario = st.number_input("Salário (R$)", min_value=0.0, value=5000.0, step=100.0)
                    
                    submitted = st.form_submit_button("Adicionar Cargo")
                    
                    if submitted and nome_cargo:
                        # Prepare cargo data with role-specific fields
                        cargo_data = {
                            'nome': nome_cargo,
                            'salario': salario,
                            'quantidade': quantidade,
                        }
                        
                        # Add role-specific data
                        if nome_cargo == "SDR":
                            cargo_data.update({
                                'taxa_agendamento': taxa_agendamento,
                                'taxa_comparecimento': taxa_comparecimento,
                                'estimativa_leads': estimativa_leads,
                                'capacidade_leads': 750  # Default capacity
                            })
                        elif nome_cargo == "Closer":
                            cargo_data.update({
                                'valor_unitario': valor_unitario,
                                'taxa_conversao': taxa_conversao,
                                'periodicidade': periodicidade,
                                'fator_aceleracao_crescimento': fator_aceleracao,
                                'capacidade_atendimentos': comparecimentos_por_closer,  # Default capacity
                                'crescimento_vendas': 'Produtividade',  # Set default growth model
                                'produtos_por_lead': produtos_por_lead,
                                'taxa_cancelamento': taxa_de_cancelamento
                            })
                        
                        # Add cargo to session state
                        if 'equipe_propria' not in st.session_state['premissas_despesas']:
                            st.session_state['premissas_despesas']['equipe_propria'] = []
                        
                        st.session_state['premissas_despesas']['equipe_propria'].append(cargo_data)
                        st.success(f"Cargo {nome_cargo} adicionado com sucesso!")
                        st.rerun()
            
            # Display current team members
            if st.session_state['premissas_despesas']['equipe_propria']:
                # Create DataFrame for display
                data_equipe = []
                
                for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                    if equipe_modo_calculo == "Percentual":
                        cargo_info = {
                            "Cargo": cargo['nome'],
                            "Percentual (%)": f"{cargo['percentual']:.1f}",
                            "Quantidade": cargo['quantidade'],
                            "Sujeito a Comissões": "Sim" if cargo.get('sujeito_comissoes', False) else "Não",
                            "Sujeito a Aumento de Receita": "Sim" if cargo.get('sujeito_aumento_receita', False) else "Não"
                        }
                    else:
                        cargo_info = {
                            "Cargo": cargo['nome'],
                            "Salário (R$)": f"{cargo['salario']:,.2f}",
                            "Quantidade": cargo['quantidade'],
                            "Sujeito a Comissões": "Sim" if cargo.get('sujeito_comissoes', False) else "Não",
                            "Sujeito a Aumento de Receita": "Sim" if cargo.get('sujeito_aumento_receita', False) else "Não"
                        }
                    data_equipe.append(cargo_info)
                
                df_equipe = pd.DataFrame(data_equipe)
                st.dataframe(df_equipe, hide_index=True, use_container_width=True)
                
                # Add selectbox to choose a role for editing
                if st.session_state['premissas_despesas']['equipe_propria']:
                    cargos_disponiveis = [cargo['nome'] for cargo in st.session_state['premissas_despesas']['equipe_propria']]
                    cargo_selecionado = st.selectbox(
                        "Selecione um cargo para editar:",
                        options=cargos_disponiveis,
                        key="editar_cargo_selecionado"
                    )
                    
                    # Find the selected cargo in the list
                    idx_selecionado = next((i for i, cargo in enumerate(st.session_state['premissas_despesas']['equipe_propria']) 
                                         if cargo['nome'] == cargo_selecionado), None)
                    
                    if idx_selecionado is not None:
                        cargo_atual = st.session_state['premissas_despesas']['equipe_propria'][idx_selecionado]
                        
                        # Inside the expander for editing the selected role
                        with st.expander(f"Editar Cargo Adicionado - Equipe Própria ({cargo_selecionado})"):
                            with st.form("editar_cargo_form"):
                                # Check if the cargo is a predefined option or custom
                                cargo_options = ["Closer", "SDR", "Account Manager", "CEO", "CFO", "Head de Vendas", "TI", "Outros"]
                                if cargo_atual['nome'] in cargo_options:
                                    cargo_selected = st.selectbox(
                                        "Cargo",
                                        cargo_options,
                                        index=cargo_options.index(cargo_atual['nome']),
                                        key="edit_cargo_select"
                                    )
                                    
                                    # If "Outros" is selected, show text input for custom role
                                    if cargo_selected == "Outros":
                                        cargo_nome = st.text_input(
                                            "Nome do Cargo Customizado",
                                            value=cargo_atual['nome'] if cargo_atual['nome'] not in cargo_options else "",
                                            key="edit_custom_cargo_nome"
                                        )
                                    else:
                                        cargo_nome = cargo_selected
                                else:
                                    # If the current cargo is not in the options, select "Outros" and show the custom input
                                    cargo_selected = st.selectbox(
                                        "Cargo",
                                        cargo_options,
                                        index=cargo_options.index("Outros"),
                                        key="edit_cargo_select"
                                    )
                                    cargo_nome = st.text_input(
                                        "Nome do Cargo Customizado",
                                        value=cargo_atual['nome'],
                                        key="edit_custom_cargo_nome"
                                    )
                                
                                salario = st.number_input(
                                    "Salário (R$)",
                                    min_value=0.0,
                                    value=cargo_atual['salario'],
                                    step=100.0,
                                    format="%.2f",
                                    key="edit_salario",
                                    help="Salário mensal para este cargo."
                                )
                                
                                quantidade = st.number_input(
                                    "Quantidade de profissionais neste cargo",
                                    min_value=0,
                                    value=cargo_atual['quantidade'],
                                    step=1,
                                    key="edit_quantidade",
                                    help="Número de profissionais neste cargo."
                                )
                                
                                percentual = cargo_atual.get('percentual', 0.0)  # Not used in Nominal mode
                                
                                # Initialize variables with current values
                                sujeito_comissoes = cargo_atual.get('sujeito_comissoes', False)
                                produtos_por_lead = cargo_atual.get('produtos_por_lead', 1) # Initialize with current or default value
                                
                                # Special fields based on role
                                if cargo_nome == "SDR":
                                    st.write("### Parâmetros Específicos para SDR")
                                    taxa_agendamento = st.slider(
                                        "Taxa de Agendamento (%)", 
                                        min_value=1.0, 
                                        max_value=100.0, 
                                        value=cargo_atual.get('taxa_agendamento', 30.0), 
                                        step=1.0,
                                        key="edit_taxa_agendamento"
                                    )
                                    taxa_comparecimento = st.slider(
                                        "Taxa de Comparecimento (%)", 
                                        min_value=1.0, 
                                        max_value=100.0, 
                                        value=cargo_atual.get('taxa_comparecimento', 70.0), 
                                        step=1.0,
                                        key="edit_taxa_comparecimento"
                                    )
                                    estimativa_leads = st.slider(
                                        "Estimativa Inicial de Leads (mensal)", 
                                        min_value=50, 
                                        max_value=1000, 
                                        value=cargo_atual.get('estimativa_leads', 200), 
                                        step=50,
                                        key="edit_estimativa_leads"
                                    )
                                    
                                    # Set default values for fields we might need later
                                    periodicidade = "Mensal"
                                    fator_aceleracao = 1.0
                                    valor_unitario = 0.0
                                    taxa_conversao = 0.0
                                    produtos_por_lead = 1 # Default value
                                    
                                elif cargo_nome == "Closer":
                                    st.write("### Parâmetros Específicos para Closer")
                                    periodicidade = st.selectbox(
                                        "Periodicidade de Aumento de Gasto",
                                        options=["Mensal", "Trimestral", "Semestral", "Anual"],
                                        index=["Mensal", "Trimestral", "Semestral", "Anual"].index(cargo_atual.get('periodicidade', "Mensal")),
                                        key="edit_periodicidade"
                                    )
                                    fator_aceleracao = st.slider(
                                        "Fator de Aceleração do Crescimento",
                                        min_value=0.1,
                                        max_value=5.0,
                                        value=cargo_atual.get('fator_aceleracao_crescimento', 1.0),
                                        step=0.1,
                                        key="edit_fator_aceleracao"
                                    )
                                    valor_unitario = st.number_input(
                                        "Valor Unitário Médio dos Produtos/Serviços (R$)",
                                        min_value=10.0,
                                        value=cargo_atual.get('valor_unitario', 2400.0),
                                        step=100.0,
                                        format="%.2f",
                                        key="edit_valor_unitario"
                                    )
                                    taxa_conversao = st.slider(
                                        "Taxa de Conversão (%)",
                                        min_value=1.0,
                                        max_value=100.0,
                                        value=cargo_atual.get('taxa_conversao', 45.0),
                                        step=1.0,
                                        key="edit_taxa_conversao"
                                    )
                                    produtos_por_lead = st.slider(
                                        "Número Médio de Produtos/Serviços por Lead",
                                        min_value=5,
                                        max_value=100,
                                        value=cargo_atual.get('produtos_por_lead', 10),
                                        step=5,
                                        key="edit_produtos_por_lead"
                                    )
                                    comparecimentos_por_closer = st.number_input(
                                        "Comparecimentos por Closer",
                                        min_value=10,
                                        max_value=500,
                                        value=cargo_atual.get('capacidade_atendimentos', 90),
                                        step=10,
                                        key="edit_comparecimentos_por_closer"
                                    )
                                    taxa_de_cancelamento = st.slider(
                                        "Taxa de Cancelamento (%)",
                                        min_value=0.0,
                                        max_value=100.0,
                                        value=cargo_atual.get('taxa_cancelamento', 5.0),
                                        step=0.5,
                                        key="edit_taxa_cancelamento"
                                    )
                                    
                                    # Set default values for fields we might need later
                                    taxa_agendamento = 0.0
                                    taxa_comparecimento = 0.0
                                    estimativa_leads = 0
                                    
                                    # For Closer, don't ask for quantidade as it will be calculated automatically
                                    quantidade = cargo_atual.get('quantidade', 1)  # Keep the existing value for data consistency
                                else:
                                    # For other roles, just set default values
                                    taxa_agendamento = 0.0
                                    taxa_comparecimento = 0.0
                                    estimativa_leads = 0
                                    periodicidade = "Mensal"
                                    valor_unitario = 0.0
                                    valor_unitario = 0.0
                                    taxa_conversao = 0.0
                                    produtos_por_lead = 1 # Default value
                                    fator_aceleracao = 1.0 # Initialize to avoid unbound error
                                    comparecimentos_por_closer = 90 # Initialize to avoid unbound error
                                
                                # Add checkbox for commission eligibility (shown for all roles)
                                col1, _ = st.columns(2)
                                with col1:
                                    sujeito_comissoes = st.checkbox(
                                        "Sujeito ao Regime de Comissões",
                                        value=sujeito_comissoes,
                                        key="edit_sujeito_comissoes",
                                        help="Se marcado, este cargo estará disponível para configuração de comissões na página 'Premissas Comissões'."
                                    )
                                
                                # Submit button inside the form    
                                submitted = st.form_submit_button("Salvar Alterações")
                                
                                # Process form submission
                                if submitted and cargo_nome:
                                    # Update the cargo with new values
                                    cargo_atualizado = {
                                        'nome': cargo_nome,
                                        'salario': salario,
                                        'quantidade': quantidade,
                                        'percentual': percentual,
                                        'sujeito_comissoes': sujeito_comissoes,
                                        'sujeito_aumento_receita': cargo_nome == "Closer" or cargo_nome == "SDR",  # Always true for Closer and SDR
                                        'taxa_agendamento': taxa_agendamento,
                                        'taxa_comparecimento': taxa_comparecimento,
                                        'estimativa_leads': estimativa_leads,
                                        'valor_unitario': valor_unitario,
                                        'taxa_conversao': taxa_conversao,
                                        'periodicidade': periodicidade,
                                        'fator_aceleracao_crescimento': fator_aceleracao,
                                        'crescimento_vendas': 'Produtividade',  # Always set to Produtividade
                                        'produtos_por_lead': produtos_por_lead if cargo_nome == "Closer" else 1,
                                        'capacidade_leads': cargo_atual.get('capacidade_leads', 750) if cargo_nome == "SDR" else 0,
                                        'capacidade_atendimentos': comparecimentos_por_closer if cargo_nome == "Closer" else 0,
                                        'taxa_cancelamento': taxa_de_cancelamento if cargo_nome == "Closer" else 0
                                    }
                                    
                                    # Update the cargo in the list
                                    st.session_state['premissas_despesas']['equipe_propria'][idx_selecionado] = cargo_atualizado
                                    st.success(f"Cargo '{cargo_nome}' atualizado com sucesso!")
                                    st.rerun()
                
                # Add removal options for team members
                st.write("#### Remover Membros da Equipe")
                
                col1, col2 = st.columns([3, 1])
                with col1:
                    cargo_para_remover = st.selectbox(
                        "Selecione um cargo para remover",
                        options=[cargo['nome'] for cargo in st.session_state['premissas_despesas']['equipe_propria']],
                        key="cargo_para_remover"
                    )
                
                with col2:
                    if st.button("Remover Cargo", key="btn_remover_cargo"):
                        # Find the index of the cargo to remove
                        idx_to_remove = next((i for i, cargo in enumerate(st.session_state['premissas_despesas']['equipe_propria']) 
                                            if cargo['nome'] == cargo_para_remover), None)
                        
                        if idx_to_remove is not None:
                            # Remove the cargo
                            st.session_state['premissas_despesas']['equipe_propria'].pop(idx_to_remove)
                            
                            # Also remove from roles_com_beneficios if present
                            if cargo_para_remover in st.session_state['premissas_despesas']['roles_com_beneficios']:
                                st.session_state['premissas_despesas']['roles_com_beneficios'].remove(cargo_para_remover)
                            
                            st.success(f"Cargo '{cargo_para_remover}' removido com sucesso!")
                            st.rerun()
                
                # Add button to remove all team members
                if st.button("Remover Todos os Cargos", key="btn_remover_todos_cargos"):
                    if st.session_state['premissas_despesas']['equipe_propria']:
                        st.session_state['premissas_despesas']['equipe_propria'] = []
                        st.session_state['premissas_despesas']['roles_com_beneficios'] = []
                        st.success("Todos os cargos foram removidos!")
                        st.rerun()
        
                # Add ability to view details of a selected role
                if st.session_state['premissas_despesas']['equipe_propria']:
                    cargo_para_detalhes = st.selectbox(
                        "Selecione um cargo para ver detalhes",
                        options=[cargo['nome'] for cargo in st.session_state['premissas_despesas']['equipe_propria']],
                        key="detalhar_cargo_equipe"
                    )
                    
                    # Find the selected cargo
                    cargo_selecionado = None
                    for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                        if cargo['nome'] == cargo_para_detalhes:
                            cargo_selecionado = cargo
                            break
                    
                    if cargo_selecionado and cargo_selecionado.get('sujeito_aumento_receita', False):
                        with st.expander(f"Detalhes de Marketing para {cargo_para_detalhes}", expanded=True):
                            st.write("### Parâmetros de Crescimento")
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write(f"**Gasto Mensal Base:** R$ {cargo_selecionado.get('gasto_mensal', 0):,.2f}")
                                st.write(f"**Custo por Lead Base:** R$ {cargo_selecionado.get('cpl_base', 0):,.2f}")
                                st.write(f"**Modelo de Crescimento:** {cargo_selecionado.get('crescimento_vendas', 'Linear')}")
                                st.write(f"**Periodicidade:** {cargo_selecionado.get('periodicidade', 'Mensal')}")
                                
                            with col2:
                                if cargo_selecionado.get('crescimento_vendas') == "Linear":
                                    st.write(f"**Taxa de Crescimento Mensal:** {cargo_selecionado.get('tx_cresc_mensal', 0):,.1f}%")
                                elif cargo_selecionado.get('crescimento_vendas') in ["Não Linear S/ Downside", "Não Linear C/ Downside"]:
                                    st.write(f"**Média de Crescimento Anual:** {cargo_selecionado.get('media_cresc_anual', 0):,.1f}%")
                                    st.write(f"**Com Downside:** {'Sim' if cargo_selecionado.get('crescimento_vendas') == 'Não Linear C/ Downside' else 'Não'}")
                                elif cargo_selecionado.get('crescimento_vendas') == "Produtividade":
                                    st.write(f"**RPE Anual:** R$ {cargo_selecionado.get('rpe_anual', 125000):,.2f}")
                                    st.write(f"**Salário Médio Anual:** R$ {cargo_selecionado.get('salario_medio', 60000):,.2f}")
                                    st.write(f"**Taxa de Depreciação:** {cargo_selecionado.get('depreciacao', 1.5):,.1f}%")
                                
                                st.write(f"**Fator de Aceleração:** {cargo_selecionado.get('fator_aceleracao_crescimento', 1):,.1f}")
                            
                            # Add new section showing parameters for _calcular_crescimento
                            st.write("### Parâmetros para Cálculo de Crescimento")
                            
                            tipo_crescimento = cargo_selecionado.get('crescimento_vendas', 'Linear')
                            growth_params = {
                                "valor_base": cargo_selecionado.get('gasto_mensal', 0),
                                "tipo_crescimento": tipo_crescimento,
                                "periodicidade": cargo_selecionado.get('periodicidade', 'Mensal'),
                                "fator_aceleracao": cargo_selecionado.get('fator_aceleracao_crescimento', 1.0)
                            }
                            
                            # Add specific parameters based on growth type
                            if tipo_crescimento == "Linear":
                                growth_params["tx_cresc_mensal"] = cargo_selecionado.get('tx_cresc_mensal', 0.0)
                            elif tipo_crescimento in ["Não Linear S/ Downside", "Não Linear C/ Downside"]:
                                growth_params["media_cresc_anual"] = cargo_selecionado.get('media_cresc_anual', 0.0)
                                growth_params["com_downside"] = tipo_crescimento == "Não Linear C/ Downside"
                            elif tipo_crescimento == "Produtividade":
                                growth_params["rpe_anual"] = cargo_selecionado.get('rpe_anual', 125000.0)
                                growth_params["salario_medio"] = cargo_selecionado.get('salario_medio', 60000.0)
                                growth_params["depreciacao"] = cargo_selecionado.get('depreciacao', 1.5)
                            
                            # Create a sample of projected values for months 0, 3, 6, 12
                            sample_months = [0, 3, 6, 12]
                            
                            try:
                                # Create a temporary instance of Receitas class to use _calcular_crescimento
                                receitas_temp = Receitas()
                                
                                # Calculate values for sample months
                                sample_values = {}
                                for mes in sample_months:
                                    value = growth_params.get("valor_base", 0)  # Initialize value
                                    try:
                                        if tipo_crescimento == "Produtividade":
                                            # Ensure necessary parameters exist and are valid before calling _calcular_crescimento
                                            if all(key in growth_params for key in ["valor_base", "periodicidade", "fator_aceleracao", "rpe_anual", "salario_medio", "depreciacao"]):
                                                # Make sure no zero division will occur
                                                payroll_data = receitas_temp._get_payroll_data(months=61)
                                                if payroll_data is not None and payroll_data[0] > 0:  # Check if payroll data exists and first value is positive
                                                    value = receitas_temp._calcular_crescimento(
                                                        growth_params["valor_base"], 
                                                        mes, 
                                                        tipo_crescimento,
                                                        periodicidade=growth_params["periodicidade"],
                                                        fator_aceleracao=growth_params["fator_aceleracao"],
                                                        rpe_anual=growth_params["rpe_anual"],
                                                        salario_medio=growth_params["salario_medio"],
                                                        depreciacao=growth_params["depreciacao"]
                                                    )
                                                else:
                                                    value = growth_params["valor_base"]  # Fallback if invalid payroll data
                                    except Exception as calc_error:
                                        st.warning(f"Erro no cálculo para mês {mes}: {str(calc_error)}")
                                        value = growth_params["valor_base"]  # Fallback on calculation error
                                    
                                    sample_values[mes] = value
                                
                                # Display the sample projected values
                                st.write("#### Projeção de Valores")
                                st.write("Valores projetados para gastos de marketing nos meses:")
                                
                                df_samples = pd.DataFrame({
                                    "Mês": sample_months,
                                    "Gasto Projetado (R$)": [sample_values[m] for m in sample_months],
                                    "Variação (%)": ["Base" if m == 0 else f"{((sample_values[m]/sample_values[0])-1)*100:.1f}%" for m in sample_months]
                                })
                                
                                st.dataframe(df_samples, hide_index=True, use_container_width=True)
                                
                            except Exception as e:
                                st.error(f"Erro ao calcular projeções: {str(e)}")
                                st.info("Por favor, verifique se os parâmetros de crescimento estão configurados corretamente.")
                            
                            st.write("### Parâmetros de Conversão")
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write(f"**Fator de Elasticidade:** {cargo_selecionado.get('fator_elasticidade', 1):,.1f}")
                                st.write(f"**Taxa de Agendamento:** {cargo_selecionado.get('taxa_agendamento', 30):,.1f}%")
                                st.write(f"**Taxa de Comparecimento:** {cargo_selecionado.get('taxa_comparecimento', 70):,.1f}%")
                            
                            with col2:
                                st.write(f"**Taxa de Conversão:** {cargo_selecionado.get('taxa_conversao', 45):,.1f}%")
                                st.write(f"**Ticket Médio:** R$ {cargo_selecionado.get('ticket_medio', 2400):,.2f}")
                            
                            # Add funnel sample calculation
                            st.write("#### Projeção de Funil de Vendas")
                            if 'gasto_mensal' in cargo_selecionado and 'cpl_base' in cargo_selecionado:
                                gasto_mensal = cargo_selecionado.get('gasto_mensal', 0)
                                cpl_base = cargo_selecionado.get('cpl_base', 10)
                                fator_elasticidade = cargo_selecionado.get('fator_elasticidade', 1)
                                
                                # Create temporary Receitas instance for CPL calculation
                                receitas_temp = Receitas()
                                cpl_ajustado = receitas_temp._calcular_cpl_ajustado(gasto_mensal, cpl_base, fator_elasticidade)
                                
                                # Calculate funnel
                                leads = gasto_mensal / cpl_ajustado if cpl_ajustado > 0 else 0
                                taxa_agendamento = cargo_selecionado.get('taxa_agendamento', 30) / 100
                                taxa_comparecimento = cargo_selecionado.get('taxa_comparecimento', 70) / 100
                                taxa_conversao = cargo_selecionado.get('taxa_conversao', 45) / 100
                                ticket_medio = cargo_selecionado.get('ticket_medio', 2400)
                                
                                agendamentos = leads * taxa_agendamento
                                comparecimentos = agendamentos * taxa_comparecimento
                                conversoes = comparecimentos * taxa_conversao
                                faturamento = conversoes * ticket_medio
                                
                                # Display funnel metrics
                                st.write("Projeção do funil de vendas para o gasto mensal base:")
                                
                                df_funnel = pd.DataFrame({
                                    "Métrica": ["Gasto Mensal", "CPL Ajustado", "Leads", "Agendamentos", 
                                               "Comparecimentos", "Conversões", "Faturamento"],
                                    "Valor": [
                                        f"R$ {gasto_mensal:,.2f}",
                                        f"R$ {cpl_ajustado:,.2f}",
                                        f"{leads:,.1f}",
                                        f"{agendamentos:,.1f}",
                                        f"{comparecimentos:,.1f}",
                                        f"{conversoes:,.1f}",
                                        f"R$ {faturamento:,.2f}"
                                    ]
                                })
                                
                                st.dataframe(df_funnel, hide_index=True, use_container_width=True)
                                
                                # Calculate and show ROAS
                                roas = faturamento / gasto_mensal if gasto_mensal > 0 else 0
                                st.write(f"**ROAS (Return On Ad Spend):** {roas:,.2f}")
            
            # Terceiros
            st.write("#### Terceiros - Prestadores de Serviços")
            if 'terceiros' in st.session_state['premissas_despesas'] and st.session_state['premissas_despesas']['terceiros']:
                data = []
                for prestador in st.session_state['premissas_despesas']['terceiros']:
                    data.append({
                        'Prestador': prestador['nome'],
                        'Valor (R$)': f"{prestador['valor']:.2f}",
                        'Quantidade': prestador['quantidade']
                    })
                df_terceiros = pd.DataFrame(data)
                st.dataframe(df_terceiros, use_container_width=True)
            else:
                st.info("Nenhum prestador de serviço adicionado.")
            
            # Expander for adding new service providers
            with st.expander("Adicionar Novo Prestador de Serviço"):
                with st.form("novo_prestador_servico"):
                    nome_prestador = st.text_input(
                        "Nome do Prestador de Serviço", 
                        key="novo_prestador_nome",
                        help="Nome do prestador de serviço a ser adicionado (ex: Agência de Tráfego, Contabilidade, etc.)"
                    )
                    
                    # Different input based on calculation mode
                    if equipe_modo_calculo == "Percentual":
                        percentual_terceiro = st.slider(
                            f"Percentual do Budget Total para {nome_prestador or 'este prestador'} (%)",
                            min_value=0.0,
                            max_value=100.0,
                            value=20.0,
                            step=0.1,
                            format="%.1f",
                            help="Percentual do orçamento total de terceiros destinado a este prestador."
                        )
                        
                        valor_terceiro = 0.0  # Will be calculated based on budget and percentage
                        quantidade_terceiro = st.number_input(
                            f"Quantidade de {nome_prestador or 'prestadores'}",
                            min_value=0,
                            value=1,
                            step=1,
                            help="Número de prestadores deste tipo."
                        )
                    else:  # Nominal mode
                        valor_terceiro = st.slider(
                            f"Valor {nome_prestador or 'deste prestador'} (R$)",
                            min_value=0.0,
                            max_value=30000.0,
                            value=2000.0,
                            step=100.0,
                            format="%.2f",
                            help="Valor mensal para este prestador de serviço."
                        )
                        
                        quantidade_terceiro = st.number_input(
                            f"Quantidade de {nome_prestador or 'prestadores'}",
                            min_value=0,
                            value=1,
                            step=1,
                            help="Número de prestadores deste tipo."
                        )
                        
                        percentual_terceiro = 0.0  # Not used in Nominal mode
                    
                    submitted = st.form_submit_button("Adicionar Prestador")
                    
                    if submitted and nome_prestador:
                        novo_prestador = {
                            'nome': nome_prestador,
                            'valor': valor_terceiro,
                            'quantidade': quantidade_terceiro,
                            'percentual': percentual_terceiro
                        }
                        
                        st.session_state['premissas_despesas']['terceiros'].append(novo_prestador)
                        st.success(f"Prestador '{nome_prestador}' adicionado com sucesso!")
                        st.rerun()
                        
            
            # Display current service providers
            if st.session_state['premissas_despesas']['terceiros']:
                # Create DataFrame for display
                data_terceiros = []
                
                if equipe_modo_calculo == "Percentual":
                    budget_terceiros = st.session_state['premissas_despesas']['budget_terceiros']
                    total_percentual_terceiros = sum(prestador['percentual'] for prestador in st.session_state['premissas_despesas']['terceiros'])
                    
                    for i, prestador in enumerate(st.session_state['premissas_despesas']['terceiros']):
                        # Calculate value based on percentage of budget and number of providers
                        if prestador['quantidade'] > 0:
                            valor_calculado = (prestador['percentual'] / 100 * budget_terceiros) / prestador['quantidade']
                        else:
                            valor_calculado = 0
                        
                        data_terceiros.append({
                            "Prestador": prestador['nome'],
                            "Percentual (%)": f"{prestador['percentual']:.1f}",
                            "Quantidade": prestador['quantidade'],
                            "Valor Calculado (R$)": f"R$ {valor_calculado:.2f}",
                            "Total (R$)": f"R$ {(prestador['percentual'] / 100 * budget_terceiros):.2f}",
                            "Ações": i
                        })
                    
                    # Show warning if total percentage exceeds 100%
                    if total_percentual_terceiros > 100:
                        st.warning(f"A soma dos percentuais ({total_percentual_terceiros:.1f}%) excede 100%. Os valores serão ajustados proporcionalmente.")
                else:
                    for i, prestador in enumerate(st.session_state['premissas_despesas']['terceiros']):
                        data_terceiros.append({
                            "Prestador": prestador['nome'],
                            "Valor (R$)": f"R$ {prestador['valor']:.2f}",
                            "Quantidade": prestador['quantidade'],
                            "Total (R$)": f"R$ {prestador['valor'] * prestador['quantidade']:.2f}",
                            "Ações": i
                        })
                
                # Display the DataFrame
                df_terceiros = pd.DataFrame(data_terceiros)
                st.dataframe(df_terceiros, hide_index=True, use_container_width=True)
                
                # Add remove buttons for each service provider
                col1, col2 = st.columns([3, 1])
                with col2:
                    prestador_para_remover = st.selectbox(
                        "Selecione o prestador para remover",
                        options=[prestador['nome'] for prestador in st.session_state['premissas_despesas']['terceiros']],
                        key="remover_prestador"
                    )
                
                if st.button("Remover Prestador Selecionado", key="btn_remover_prestador"):
                    # Find and remove the selected service provider
                    for i, prestador in enumerate(st.session_state['premissas_despesas']['terceiros']):
                        if prestador['nome'] == prestador_para_remover:
                            st.session_state['premissas_despesas']['terceiros'].pop(i)
                            st.success(f"Prestador '{prestador_para_remover}' removido com sucesso!")
                            st.rerun()
                            break
            else:
                st.info("Nenhum prestador de serviços adicionado. Utilize o formulário acima para adicionar prestadores.")
            
            # Encargos Sociais
            st.write("#### Encargos Sociais")
            st.session_state['premissas_despesas']['encargos_sociais_perc'] = st.slider(
                "Encargos Sociais (%)",
                min_value=0.0,
                max_value=100.0,
                value=st.session_state['premissas_despesas']['encargos_sociais_perc'],
                step=1.0,
                format="%.1f",
                help="Percentual de encargos sociais sobre a folha de pagamento. Inclui INSS, FGTS, férias, 13º salário, etc."
            )
            
            # Despesas com Alimentação e Transporte
            st.write("#### Despesas com Alimentação e Transporte")
            col1, col2 = st.columns(2)
            with col1:
                st.session_state['premissas_despesas']['vale_alimentacao'] = st.slider(
                    "Vale Alimentação (R$/dia)",
                    min_value=0.0,
                    max_value=100.0,
                    value=st.session_state['premissas_despesas']['vale_alimentacao'],
                    step=1.0,
                    format="%.2f",
                    help="Valor diário do vale alimentação fornecido aos funcionários."
                )
            with col2:
                st.session_state['premissas_despesas']['vale_transporte'] = st.slider(
                    "Vale Transporte (R$/dia)",
                    min_value=0.0,
                    max_value=50.0,
                    value=st.session_state['premissas_despesas']['vale_transporte'],
                    step=0.5,
                    format="%.2f",
                    help="Valor diário do vale transporte fornecido aos funcionários."
                )
            
            # Cargos com benefícios
            st.write("#### Cargos com Benefícios")
            
            if st.session_state['premissas_despesas']['equipe_propria']:
                # Get all role names
                role_names = [cargo['nome'] for cargo in st.session_state['premissas_despesas']['equipe_propria']]
                
                # Initialize 'roles_com_beneficios' if it doesn't exist or contains outdated roles
                if 'roles_com_beneficios' not in st.session_state['premissas_despesas'] or not all(role in role_names for role in st.session_state['premissas_despesas']['roles_com_beneficios']):
                    st.session_state['premissas_despesas']['roles_com_beneficios'] = role_names.copy()
                
                # Multi-select for roles with benefits
                st.session_state['premissas_despesas']['roles_com_beneficios'] = st.multiselect(
                    "Selecione os cargos que receberão benefícios:",
                    options=role_names,
                    default=st.session_state['premissas_despesas']['roles_com_beneficios'],
                    help="Selecione os cargos que receberão benefícios de vale alimentação e vale transporte."
                )
            else:
                st.info("Adicione cargos à equipe própria para configurar os benefícios.")
            
            # Bônus dos Lucros
            st.write("#### Bônus dos Lucros")
            col1, col2 = st.columns(2)
            
            with col1:
                st.session_state['premissas_despesas']['benchmark_anual_bonus'] = st.slider(
                    "Benchmark Anual para Bônus (%)",
                    min_value=0.0,
                    max_value=50.0,
                    value=st.session_state['premissas_despesas']['benchmark_anual_bonus'],
                    step=0.5,
                    format="%.1f",
                    help="Percentual de benchmark anual utilizado como base para cálculo dos bônus dos lucros."
                )
                
                st.session_state['premissas_despesas']['lucro_liquido_inicial'] = st.slider(
                    "Lucro Líquido Inicial (R$)",
                    min_value=0.0,
                    max_value=1000000.0,
                    value=st.session_state['premissas_despesas']['lucro_liquido_inicial'],
                    step=10000.0,
                    format="%.2f",
                    help="Valor do lucro líquido inicial utilizado para cálculo dos bônus dos lucros."
                )
            
            with col2:
                st.session_state['premissas_despesas']['crescimento_lucro'] = st.slider(
                    "Crescimento Anual do Lucro (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=st.session_state['premissas_despesas']['crescimento_lucro'],
                    step=1.0,
                    format="%.1f",
                    help="Taxa de crescimento anual do lucro líquido, que influencia o cálculo dos bônus dos lucros."
                )
                
            # Bonus percentage for each role
            st.write("#### Percentual de Bônus por Cargo")
            
            if st.session_state['premissas_despesas']['equipe_propria']:
                # Create a form for bonus percentages
                bonus_col1, bonus_col2 = st.columns(2)
                
                # In PremissasDespesas.render, in tab1 where the bonus sliders are created
                
                # Replace this code in tab1 for the bonus percentage sliders:
                for i, cargo in enumerate(st.session_state['premissas_despesas']['equipe_propria']):
                    # Alternate columns
                    current_col = bonus_col1 if i % 2 == 0 else bonus_col2
                    
                    with current_col:
                        bonus_key = f"bonus_{cargo['nome'].lower().replace(' ', '_')}"
                        
                        # Initialize if not exists
                        if bonus_key not in st.session_state['premissas_despesas']:
                            st.session_state['premissas_despesas'][bonus_key] = 1.0
                        
                        # Fix: Add index to key to ensure uniqueness
                        st.session_state['premissas_despesas'][bonus_key] = st.slider(
                            f"Bônus {cargo['nome']} (%)",
                            min_value=0.0,
                            max_value=20.0,
                            value=st.session_state['premissas_despesas'].get(bonus_key, 1.0),
                            step=0.1,
                            format="%.1f",
                            # Add index to make key unique
                            key=f"slider_bonus_{cargo['nome'].lower().replace(' ', '_')}_{i}",
                            help=f"Percentual do bônus para {cargo['nome']} calculado sobre o lucro líquido anual excedente ao benchmark."
                        )
            else:
                st.info("Adicione cargos à equipe própria para configurar os percentuais de bônus.")
        
        with tab4:
            st.write("### Custos de Tecnologia")
            st.write("Defina os valores para os custos relacionados à tecnologia.")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.session_state['premissas_despesas']['desenvolvimento_ferramenta'] = st.slider(
                    "Desenvolvimento da ferramenta/mês (R$)",
                    min_value=0.0,
                    max_value=20000.0,
                    value=st.session_state['premissas_despesas'].get('desenvolvimento_ferramenta', 0.0),
                    step=100.0,
                    format="%.2f",
                    help="Valor mensal destinado ao desenvolvimento da ferramenta de gestão de vendas."
                )
                
                st.session_state['premissas_despesas']['manutencao_ferramenta'] = st.slider(
                    "Manutenção da ferramenta/mês (R$)",
                    min_value=0.0,
                    max_value=10000.0,
                    value=st.session_state['premissas_despesas'].get('manutencao_ferramenta', 0.0),
                    step=100.0,
                    format="%.2f",
                    help="Valor mensal destinado à manutenção da ferramenta de gestão de vendas."
                )
            
            with col2:
                st.session_state['premissas_despesas']['inovacao'] = st.slider(
                    "Inovação/mês (R$)",
                    min_value=0.0,
                    max_value=10000.0,
                    value=st.session_state['premissas_despesas'].get('inovacao', 0.0),
                    step=100.0,
                    format="%.2f",
                    help="Valor mensal destinado a inovações e melhorias na ferramenta de gestão de vendas."
                )
                
                st.session_state['premissas_despesas']['licencas_software'] = st.slider(
                    "Licenças de software/mês (R$)",
                    min_value=0.0,
                    max_value=10000.0,
                    value=st.session_state['premissas_despesas'].get('licencas_software', 2513.0),
                    step=100.0,
                    format="%.2f",
                    help="Valor mensal gasto com licenças de software necessários para a operação da empresa."
                )
            
            # Add equipments expander
            with st.expander("Novos Equipamentos"):
                st.write("### Gestão de Equipamentos e Depreciação")
                
                # Initialize equipments list if it doesn't exist
                if 'equipamentos' not in st.session_state['premissas_despesas']:
                    st.session_state['premissas_despesas']['equipamentos'] = []
                
                # Form to add new equipment
                with st.form("novo_equipamento"):
                    nome_equip = st.text_input("Nome do Equipamento", key="equip_nome")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        valor_equip = st.slider(
                            "Valor (R$)",
                            min_value=100.0,
                            max_value=10000.0,
                            value=5000.0,
                            step=100.0,
                            format="%.2f"
                        )
                        # Add a month selector
                        mes_aquisicao = st.slider(
                            "Mês de Aquisição",
                            min_value=0,
                            max_value=59,
                            value=0,
                            step=1,
                            help="Em qual mês este equipamento será adquirido (0-59)"
                        )
                        
                    with col2:
                        quantidade_equip = st.slider(
                            "Quantidade",
                            min_value=1,
                            max_value=100,
                            value=1,
                            step=1
                        )
                    
                    metodo_depreciacao = st.selectbox(
                        "Método de Depreciação",
                        ["Método da Linha Reta", "Método da Soma dos Dígitos"],
                        help="Método da Linha Reta: depreciação constante ao longo da vida útil.\n\n"
                             "Método da Soma dos Dígitos: depreciação acelerada, maior nos primeiros anos."
                    )
                    
                    if metodo_depreciacao == "Método da Linha Reta":
                        vida_util = st.slider(
                            f"Vida Útil de {nome_equip if nome_equip else 'equipamento'} (Anos)",
                            min_value=1,
                            max_value=20,
                            value=5,
                            step=1
                        )
                        
                        valor_residual = st.slider(
                            f"Valor Residual Estimado de {nome_equip if nome_equip else 'equipamento'} (R$)",
                            min_value=0.0,
                            max_value=valor_equip * 0.5,  # Maximum 50% of original value
                            value=valor_equip * 0.1,  # Default 10% of original value
                            step=100.0,
                            format="%.2f"
                        )
                        
                        # Calculate depreciation
                        depreciacao_anual = (valor_equip - valor_residual) / vida_util
                        st.write(f"Depreciação Anual: R$ {depreciacao_anual:.2f} por unidade")
                        st.write(f"Depreciação Anual Total: R$ {depreciacao_anual * quantidade_equip:.2f}")
                        
                        # Additional parameters for this method
                        metodo_params = {
                            "vida_util": vida_util,
                            "valor_residual": valor_residual,
                            "depreciacao_anual": depreciacao_anual
                        }
                        
                    else:  # Método da Soma dos Dígitos
                        anos_restantes = st.slider(
                            f"Anos Restantes de {nome_equip if nome_equip else 'equipamento'} (Anos)",
                            min_value=1,
                            max_value=20,
                            value=5,
                            step=1
                        )
                        
                        valor_residual = st.slider(
                            f"Valor Residual Estimado de {nome_equip if nome_equip else 'equipamento'} (R$)",
                            min_value=0.0,
                            max_value=valor_equip * 0.5,  # Maximum 50% of original value
                            value=valor_equip * 0.1,  # Default 10% of original value
                            step=100.0,
                            format="%.2f"
                        )
                        
                        # Calculate sum of years digits
                        sum_of_years = sum(range(1, anos_restantes + 1))
                        
                        # Show depreciation for each year
                        st.write("Depreciação Anual do Equipamento:")
                        depreciacao_anos = {}
                        valor_depreciavel = valor_equip - valor_residual
                        
                        for i in range(1, anos_restantes + 1):
                            fator = (anos_restantes - i + 1) / sum_of_years
                            dep_ano = valor_depreciavel * fator
                            depreciacao_anos[i] = dep_ano
                            st.write(f"Ano {i}: R$ {dep_ano:.2f} por unidade (R$ {dep_ano * quantidade_equip:.2f} total)")
                        
                        # Additional parameters for this method
                        metodo_params = {
                            "anos_restantes": anos_restantes,
                            "valor_residual": valor_residual,
                            "sum_of_years": sum_of_years,
                            "depreciacao_anos": depreciacao_anos
                        }
                    
                    submitted = st.form_submit_button("Adicionar Equipamento")
                    
                    if submitted and nome_equip:
                        # Create equipment dictionary
                        novo_equipamento = {
                            "nome": nome_equip,
                            "valor": valor_equip,
                            "quantidade": quantidade_equip,
                            "metodo": metodo_depreciacao,
                            "metodo_params": metodo_params,
                            "mes_aquisicao": mes_aquisicao  
                        }
                        
                        # Add to session state
                        st.session_state['premissas_despesas']['equipamentos'].append(novo_equipamento)
                        st.success(f"Equipamento '{nome_equip}' adicionado com sucesso!")
                        st.rerun()
                
                # Display existing equipment
                if st.session_state['premissas_despesas']['equipamentos']:
                    st.write("### Equipamentos Cadastrados")
                    
                    data = []
                    for i, equip in enumerate(st.session_state['premissas_despesas']['equipamentos']):
                        # Calculate total value
                        valor_total = equip['valor'] * equip['quantidade']
                        
                        # Get depreciation info
                        if equip['metodo'] == "Método da Linha Reta":
                            depreciacao_anual = equip['metodo_params']['depreciacao_anual'] * equip['quantidade']
                            info_depreciacao = f"R$ {depreciacao_anual:.2f}/ano"
                        else:
                            # Get first year depreciation for display
                            dep_ano1 = equip['metodo_params']['depreciacao_anos'][1] * equip['quantidade']
                            info_depreciacao = f"R$ {dep_ano1:.2f} (1º ano)"
                        
                        data.append({
                            "Nome": equip['nome'],
                            "Valor Unitário": f"R$ {equip['valor']:,.2f}",
                            "Quantidade": equip['quantidade'],
                            "Valor Total": f"R$ {valor_total:,.2f}",
                            "Método": "Linha Reta" if equip['metodo'] == "Método da Linha Reta" else "Soma Dígitos",
                            "Depreciação": info_depreciacao,
                            "Ações": i
                        })
                    
                    # Display equipment table
                    df_equip = pd.DataFrame(data)
                    st.dataframe(df_equip, hide_index=True, use_container_width=True)
                    
                    # Add option to remove equipment
                    col1, col2 = st.columns([3, 1])
                    with col2:
                        equip_to_remove = st.selectbox(
                            "Selecione para remover",
                            options=[equip['nome'] for equip in st.session_state['premissas_despesas']['equipamentos']]
                        )
                    
                    if st.button("Remover Equipamento Selecionado"):
                        for i, equip in enumerate(st.session_state['premissas_despesas']['equipamentos']):
                            if equip['nome'] == equip_to_remove:
                                st.session_state['premissas_despesas']['equipamentos'].pop(i)
                                st.success(f"Equipamento '{equip_to_remove}' removido com sucesso!")
                                st.rerun()
                                break
        
        with tab5:
            st.write("### Configuração de Reajustes")
            st.write("Defina quais despesas serão reajustadas automaticamente e com qual índice.")
            
            # Initialize reajustes configuration in session state if it doesn't exist
            if 'reajustes' not in st.session_state['premissas_despesas']:
                st.session_state['premissas_despesas']['reajustes'] = {
                    'mensal': {},
                    'trimestral': {},
                    'semestral': {},
                    'anual': {}
                }
            
            # Ensure each reajuste has the correct structure
            for period in ['mensal', 'trimestral', 'semestral', 'anual']:
                if period not in st.session_state['premissas_despesas']['reajustes']:
                    st.session_state['premissas_despesas']['reajustes'][period] = {}
                    
                if not isinstance(st.session_state['premissas_despesas']['reajustes'][period], dict):
                    st.session_state['premissas_despesas']['reajustes'][period] = {}
                    
                if 'expenses' not in st.session_state['premissas_despesas']['reajustes'][period]:
                    st.session_state['premissas_despesas']['reajustes'][period]['expenses'] = []
            
            # List of all available expense categories
            all_expenses = [
                "Água e Luz",
                "Aluguéis, Condomínios e IPTU",
                "Internet",
                "Material de Escritório",
                "Treinamentos",
                "Manutenção & Conservação",
                "Seguros Funcionários",
                "Licenças de Telefonia",
                "Licenças CRM",
                "Telefônica",
                "Equipe Própria",
                "Terceiros",
                "Alimentação e Transporte",
                "Desenvolvimento da ferramenta",
                "Manutenção da ferramenta",
                "Inovação",
                "Licenças de software"
            ]
            
            # Create a column for each timeframe
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("#### Reajustes Mensais e Trimestrais")
                
                # Mensal reajustes - Fix the default value issue
                mensal_expenses = st.multiselect(
                    "Despesas com Reajuste Mensal",
                    options=all_expenses,
                    default=st.session_state['premissas_despesas']['reajustes']['mensal'].get('expenses', []),
                    key="mensal_expenses"
                )
                
                mensal_index = st.radio(
                    "Índice para Reajuste Mensal",
                    options=["IPCA Médio Anual (%)", "IGP-M Médio Anual (%)", "Sem reajuste"],
                    index=0,
                    key="mensal_index"
                )
                
                st.session_state['premissas_despesas']['reajustes']['mensal'] = {
                    'expenses': mensal_expenses,
                    'index': mensal_index
                }
                
                # Trimestral reajustes
                trimestral_expenses = st.multiselect(
                    "Despesas com Reajuste Trimestral",
                    options=all_expenses,
                    default=st.session_state['premissas_despesas']['reajustes']['trimestral'].get('expenses', []),
                    key="trimestral_expenses"
                )
                
                trimestral_index = st.radio(
                    "Índice para Reajuste Trimestral",
                    options=["IPCA Médio Anual (%)", "IGP-M Médio Anual (%)", "Sem reajuste"],
                    index=0,
                    key="trimestral_index"
                )
                
                st.session_state['premissas_despesas']['reajustes']['trimestral'] = {
                    'expenses': trimestral_expenses,
                    'index': trimestral_index
                }
            
            with col2:
                st.write("#### Reajustes Semestrais e Anuais")
                
                # Semestral reajustes
                semestral_expenses = st.multiselect(
                    "Despesas com Reajuste Semestral",
                    options=all_expenses,
                    default=st.session_state['premissas_despesas']['reajustes']['semestral'].get('expenses', []),
                    key="semestral_expenses"
                )
                
                semestral_index = st.radio(
                    "Índice para Reajuste Semestral",
                    options=["IPCA Médio Anual (%)", "IGP-M Médio Anual (%)", "Sem reajuste"],
                    index=0,
                    key="semestral_index"
                )
                
                st.session_state['premissas_despesas']['reajustes']['semestral'] = {
                    'expenses': semestral_expenses,
                    'index': semestral_index
                }
                
                # Anual reajustes
                anual_expenses = st.multiselect(
                    "Despesas com Reajuste Anual",
                    options=all_expenses,
                    default=st.session_state['premissas_despesas']['reajustes']['anual'].get('expenses', []),
                    key="anual_expenses"
                )
                
                anual_index = st.radio(
                    "Índice para Reajuste Anual",
                    options=["IPCA Médio Anual (%)", "IGP-M Médio Anual (%)", "Sem reajuste"],
                    index=0,
                    key="anual_index"
                )
                
                st.session_state['premissas_despesas']['reajustes']['anual'] = {
                    'expenses': anual_expenses,
                    'index': anual_index
                }
            
            # Display a summary of the reajustes
            st.write("### Resumo dos Reajustes")
            
            reajustes_data = []
            for periodo, config in st.session_state['premissas_despesas']['reajustes'].items():
                if isinstance(config, dict) and 'expenses' in config and 'index' in config:
                    for expense in config['expenses']:
                        reajustes_data.append({
                            "Despesa": expense,
                            "Periodicidade": periodo.capitalize(),
                            "Índice": config['index']
                        })
            
            if reajustes_data:
                df_reajustes = pd.DataFrame(reajustes_data)
                st.dataframe(df_reajustes, use_container_width=True)
            else:
                st.info("Nenhum reajuste configurado.")
    
        # Resumo das premissas
        st.write("### Resumo das Premissas de Despesas")
        
        # Create tabs for summary display
        tab_geral, tab_valores, tab_equipe = st.tabs(["Configurações Gerais", "Valores das Despesas", "Equipe"])
        
        with tab_geral:
            df_premissas_geral = pd.DataFrame({
                'Parâmetro': [
                    'IPCA Médio Anual',
                    'Modo de Cálculo',
                    'Orçamento Mensal Total',
                    'Modo de Cálculo de Energia'
                ],
                'Valor': [
                    f"{st.session_state['premissas_despesas']['ipca_medio_anual']}%",
                    st.session_state['premissas_despesas']['modo_calculo'],
                    f"R$ {st.session_state['premissas_despesas']['budget_mensal']:.2f}" if st.session_state['premissas_despesas']['modo_calculo'] == "Percentual" else "N/A",
                    st.session_state['premissas_despesas']['modo_energia']
                ]
            })
            
            st.dataframe(df_premissas_geral, use_container_width=True)
        
        with tab_valores:
            if st.session_state['premissas_despesas']['modo_calculo'] == "Nominal":
                # Exibir valores nominais
                aluguel_condominio_iptu = (
                    st.session_state['premissas_despesas']['aluguel'] +
                    st.session_state['premissas_despesas']['condominio'] +
                    st.session_state['premissas_despesas']['iptu']
                )
                
                # Calcular valor de Água e Luz baseado no modo selecionado
                agua_luz_valor = st.session_state['premissas_despesas']['consumo_mensal_kwh']
                agua_luz_valor += (0.2 * agua_luz_valor)  # 20% de impostos
                agua_luz_valor += (0.0412 * agua_luz_valor)  # 4.12% de impostos
                
                total_despesas = (
                    agua_luz_valor +
                    aluguel_condominio_iptu +
                    st.session_state['premissas_despesas']['marketing_publicidade'] +
                    st.session_state['premissas_despesas']['internet'] +
                    st.session_state['premissas_despesas']['material_escritorio'] +
                    st.session_state['premissas_despesas']['treinamentos'] +
                    st.session_state['premissas_despesas']['manutencao_conservacao'] +
                    st.session_state['premissas_despesas']['seguros_funcionarios'] +
                    st.session_state['premissas_despesas']['licencas_telefonia'] +
                    st.session_state['premissas_despesas']['licencas_crm'] +
                    st.session_state['premissas_despesas']['telefonica']
                )
                
                df_premissas_valores = pd.DataFrame({
                    'Despesa': [
                        'Água e Luz',
                        'Aluguéis, Condomínios e IPTU',
                        'Marketing & Publicidade',
                        'Internet',
                        'Material de Escritório',
                        'Treinamentos',
                        'Manutenção & Conservação',
                        'Seguros Funcionários',
                        'Licenças de Telefonia',
                        'Licenças CRM',
                        'Telefônica',
                        'Total'
                    ],
                    'Valor (R$)': [
                        f"{agua_luz_valor:.2f}",
                        f"{aluguel_condominio_iptu:.2f}",
                        f"{st.session_state['premissas_despesas']['marketing_publicidade']:.2f}",
                        f"{st.session_state['premissas_despesas']['internet']:.2f}",
                        f"{st.session_state['premissas_despesas']['material_escritorio']:.2f}",
                        f"{st.session_state['premissas_despesas']['treinamentos']:.2f}",
                        f"{st.session_state['premissas_despesas']['manutencao_conservacao']:.2f}",
                        f"{st.session_state['premissas_despesas']['seguros_funcionarios']:.2f}",
                        f"{st.session_state['premissas_despesas']['licencas_telefonia']:.2f}",
                        f"{st.session_state['premissas_despesas']['licencas_crm']:.2f}",
                        f"{st.session_state['premissas_despesas']['telefonica']:.2f}",
                        f"{total_despesas:.2f}"
                    ],
                    'Percentual (%)': [
                        f"{(agua_luz_valor / total_despesas * 100):.1f}%",
                        f"{(aluguel_condominio_iptu / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['marketing_publicidade'] / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['internet'] / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['material_escritorio'] / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['treinamentos'] / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['manutencao_conservacao'] / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['seguros_funcionarios'] / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['licencas_telefonia'] / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['licencas_crm'] / total_despesas * 100):.1f}%",
                        f"{(st.session_state['premissas_despesas']['telefonica'] / total_despesas * 100):.1f}%",
                        "100.0%"
                    ]
                })
            else:
                # Exibir valores percentuais
                budget = st.session_state['premissas_despesas']['budget_mensal']
                
                # Recalcular a soma dos percentuais para garantir que está definido
                soma_percentuais = (
                    st.session_state['premissas_despesas']['perc_agua_luz'] +
                    st.session_state['premissas_despesas']['perc_aluguel_condominio_iptu'] +
                    st.session_state['premissas_despesas']['perc_marketing_publicidade'] +
                    st.session_state['premissas_despesas']['perc_internet'] +
                    st.session_state['premissas_despesas']['perc_material_escritorio'] +
                    st.session_state['premissas_despesas']['perc_treinamentos'] +
                    st.session_state['premissas_despesas']['perc_manutencao_conservacao'] +
                    st.session_state['premissas_despesas']['perc_seguros_funcionarios'] +
                    st.session_state['premissas_despesas']['perc_licencas_telefonia'] +
                    st.session_state['premissas_despesas']['perc_licencas_crm'] +
                    st.session_state['premissas_despesas']['perc_telefonica']
                )
                
                df_premissas_valores = pd.DataFrame({
                    'Despesa': [
                        'Água e Luz',
                        'Aluguéis, Condomínios e IPTU',
                        'Marketing & Publicidade',
                        'Internet',
                        'Material de Escritório',
                        'Treinamentos',
                        'Manutenção & Conservação',
                        'Seguros Funcionários',
                        'Licenças de Telefonia',
                        'Licenças CRM',
                        'Telefônica',
                        'Total'
                    ],
                    'Percentual (%)': [
                        f"{st.session_state['premissas_despesas']['perc_agua_luz']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_aluguel_condominio_iptu']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_marketing_publicidade']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_internet']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_material_escritorio']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_treinamentos']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_manutencao_conservacao']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_seguros_funcionarios']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_licencas_telefonia']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_licencas_crm']:.1f}%",
                        f"{st.session_state['premissas_despesas']['perc_telefonica']:.1f}%",
                        f"{soma_percentuais:.1f}%"
                    ],
                    'Valor (R$)': [
                        f"{budget * st.session_state['premissas_despesas']['perc_agua_luz'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_aluguel_condominio_iptu'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_marketing_publicidade'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_internet'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_material_escritorio'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_treinamentos'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_manutencao_conservacao'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_seguros_funcionarios'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_licencas_telefonia'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_licencas_crm'] / 100:.2f}",
                        f"{budget * st.session_state['premissas_despesas']['perc_telefonica'] / 100:.2f}",
                        f"{budget:.2f}"
                    ]
                })
            
            st.dataframe(df_premissas_valores, use_container_width=True)
        
        # Add the new tab_equipe content
        with tab_equipe:
            
            # Equipe Própria
            st.write("#### Equipe Própria")
            df_equipe_propria = pd.DataFrame({
                'Cargo': ['CEO', 'CFO', 'Head de Vendas', 'SDR', 'Closer', 'Account Manager', 'TI'],
                'Salário (R$)': [
                    f"{st.session_state['premissas_despesas']['ceo_salario']:.2f}",
                    f"{st.session_state['premissas_despesas']['cfo_salario']:.2f}",
                    f"{st.session_state['premissas_despesas']['head_vendas_salario']:.2f}",
                    f"{st.session_state['premissas_despesas']['sdr_salario']:.2f}",
                    f"{st.session_state['premissas_despesas']['closer_salario']:.2f}",
                    f"{st.session_state['premissas_despesas']['social_media_salario']:.2f}",
                    f"{st.session_state['premissas_despesas']['ti_salario']:.2f}",
                ],
                'Quantidade': [
                    st.session_state['premissas_despesas']['ceo_qtd'],
                    st.session_state['premissas_despesas']['cfo_qtd'],
                    st.session_state['premissas_despesas']['head_vendas_qtd'],
                    st.session_state['premissas_despesas']['sdr_qtd'],
                    st.session_state['premissas_despesas']['closer_qtd'],
                    st.session_state['premissas_despesas']['social_media_qtd'],
                    st.session_state['premissas_despesas']['ti_qtd'],
                ]
            })
            st.dataframe(df_equipe_propria, use_container_width=True)
            
            # Encargos e Benefícios
            st.write("#### Encargos e Benefícios")
            df_encargos = pd.DataFrame({
                'Parâmetro': [
                    'Encargos Sociais',
                    'Vale Alimentação',
                    'Vale Transporte',
                    'Cargos com Benefícios'
                ],
                'Valor': [
                    f"{st.session_state['premissas_despesas']['encargos_sociais_perc']}%",
                    f"R$ {st.session_state['premissas_despesas']['vale_alimentacao']:.2f}/dia",
                    f"R$ {st.session_state['premissas_despesas']['vale_transporte']:.2f}/dia",
                    ", ".join([cargo.capitalize() for cargo in st.session_state['premissas_despesas']['roles_com_beneficios']])
                ]
            })
            st.dataframe(df_encargos, use_container_width=True)
            
            # Terceiros
            st.write("#### Terceiros - Prestadores de Serviços")
            df_terceiros = pd.DataFrame({
                'Prestador': [
                    'Agência de Tráfego',
                    'Contabilidade',
                    'Jurídico',
                    'Agência de Criativos'
                ],
                'Valor (R$)': [
                    f"{st.session_state['premissas_despesas']['agencia_trafego_valor']:.2f}",
                    f"{st.session_state['premissas_despesas']['contabilidade_valor']:.2f}",
                    f"{st.session_state['premissas_despesas']['juridico_valor']:.2f}",
                    f"{st.session_state['premissas_despesas']['agencia_criativos_valor']:.2f}"
                ],
                'Quantidade': [
                    st.session_state['premissas_despesas']['agencia_trafego_qtd'],
                    st.session_state['premissas_despesas']['contabilidade_qtd'],
                    st.session_state['premissas_despesas']['juridico_qtd'],
                    st.session_state['premissas_despesas']['agencia_criativos_qtd']
                ]
            })
            st.dataframe(df_terceiros, use_container_width=True)
            
            # Bônus
            st.write("#### Bônus dos Lucros")
            df_bonus = pd.DataFrame({
                'Parâmetro': [
                    'Benchmark Anual para Bônus',
                    'Percentual Bônus CEO',
                    'Percentual Bônus CFO',
                    'Percentual Bônus Head de Vendas',
                    'Lucro Líquido Inicial',
                    'Crescimento Anual do Lucro'
                ],
                'Valor': [
                    f"{st.session_state['premissas_despesas']['benchmark_anual_bonus']}%",
                    f"{st.session_state['premissas_despesas']['percentual_bonus_ceo']}%",
                    f"{st.session_state['premissas_despesas']['percentual_bonus_cfo']}%",
                    f"{st.session_state['premissas_despesas']['percentual_bonus_head_vendas']}%",
                    f"R$ {st.session_state['premissas_despesas']['lucro_liquido_inicial']:,.2f}",
                    f"{st.session_state['premissas_despesas']['crescimento_lucro']}%"
                ]
            })
            st.dataframe(df_bonus, use_container_width=True)
            
class DespesasAdm(Page):
    def __init__(self):
        # Verificar se as premissas de despesas existem
        if 'premissas_despesas' not in st.session_state:
            # Inicializar premissas de despesas
            PremissasDespesas()
        
        self.graph_types = ["Gráfico de Linhas", "Gráfico de Pizza"]
        self.time_frames = ["Mensal", "Anual"]
    
    @property
    def title(self) -> str:
        return "Despesas Administrativas"
    
    @property
    def icon(self) -> str:
        return "📊"
    
    def _calcular_valor_agua_luz(self, consumo_kwh, mes, modo_energia):
        """Calcula o valor de água e luz baseado no consumo e no modo de energia"""
        # Taxa base: consumo + 20% de impostos + 4.12% de impostos
        valor_base = consumo_kwh + (0.2 * consumo_kwh) + (0.0412 * consumo_kwh)
        
        if modo_energia == "Constante":
            return valor_base
        else:
            # Modo estressado: escolher uma das bandeiras tarifárias
            # A escolha é feita a cada 2 meses para manter consistência
            bandeiras = [
                "Bandeira Verde",
                "Bandeira Amarela",
                "Bandeira Vermelha - Patamar 1",
                "Bandeira Vermelha - Patamar 2"
            ]
            
            # Usar o mês para determinar a bandeira (mesmo valor por 2 meses consecutivos)
            seed = mes // 2
            random.seed(seed)
            bandeira = bandeiras[random.randint(0, 3)]
            
            # Aplicar o adicional de acordo com a bandeira
            if bandeira == "Bandeira Verde":
                return valor_base
            elif bandeira == "Bandeira Amarela":
                return valor_base + (consumo_kwh * 0.01885)
            elif bandeira == "Bandeira Vermelha - Patamar 1":
                return valor_base + (consumo_kwh * 0.04463)
            else:  # Bandeira Vermelha - Patamar 2
                return valor_base + (consumo_kwh * 0.07877)
    
    # Add this to the _gerar_dataframe_despesas method in DespesasAdm class
    def _gerar_dataframe_despesas(self):
        """Gera o dataframe de despesas para 60 meses (5 anos)"""
        if 'premissas_despesas' not in st.session_state:
            return pd.DataFrame()
        
        p = st.session_state['premissas_despesas']
        meses = 60
        
        # Get the starting month for expenses
        mes_inicio = p.get('mes_inicio_despesas', 0)
        
        # Inicializar DataFrame com os meses
        df = pd.DataFrame(index=range(meses))
        df.index.name = 'Mês'
        
        # Get reajustes configuration
        reajustes = p.get('reajustes', {})
        
        # Calculate base inflation rates
        ipca_anual = p['ipca_medio_anual'] / 100
        igpm_anual = p['igpm_medio_anual'] / 100
        
        # Calculate monthly equivalents
        taxa_ipca_mensal = (1 + ipca_anual) ** (1/12) - 1
        taxa_igpm_mensal = (1 + igpm_anual) ** (1/12) - 1
        
        # Inicializar arrays para armazenar valores de cada despesa por mês
        agua_luz = []
        aluguel_condominio_iptu = []
        internet = []
        material_escritorio = []
        treinamentos = []
        manutencao_conservacao = []
        seguros_funcionarios = []
        licencas_telefonia = []
        licencas_crm = []
        telefonica = []
        
        # Mapping of expense names to arrays
        expense_arrays = {
            "Água e Luz": agua_luz,
            "Aluguéis, Condomínios e IPTU": aluguel_condominio_iptu,
            "Internet": internet,
            "Material de Escritório": material_escritorio,
            "Treinamentos": treinamentos,
            "Manutenção & Conservação": manutencao_conservacao,
            "Seguros Funcionários": seguros_funcionarios,
            "Licenças de Telefonia": licencas_telefonia,
            "Licenças CRM": licencas_crm,
            "Telefônica": telefonica
        }
        
        # Function to check if a expense should be adjusted in a specific month
        def should_adjust(expense_name, month):
            # Check monthly adjustments
            if 'mensal' in reajustes and isinstance(reajustes['mensal'], dict) and expense_name in reajustes['mensal'].get('expenses', []):
                return True, reajustes['mensal']['index']
            
            # Check quarterly adjustments
            if 'trimestral' in reajustes and isinstance(reajustes['trimestral'], dict) and expense_name in reajustes['trimestral'].get('expenses', []) and month % 3 == 0:
                return True, reajustes['trimestral']['index']
            
            # Check semi-annual adjustments
            if 'semestral' in reajustes and isinstance(reajustes['semestral'], dict) and expense_name in reajustes['semestral'].get('expenses', []) and month % 6 == 0:
                return True, reajustes['semestral']['index']
            
            # Check annual adjustments
            if 'anual' in reajustes and isinstance(reajustes['anual'], dict) and expense_name in reajustes['anual'].get('expenses', []) and month % 12 == 0:
                return True, reajustes['anual']['index']
            
            return False, None
        
        if p['modo_calculo'] == "Percentual":
            # Process for Percentual mode
            budget_base = p['budget_mensal']
            
            for i in range(meses):
                # Apply zeros for months before mes_inicio
                if i < mes_inicio:
                    for expense_name, array in expense_arrays.items():
                        array.append(0.0)
                    continue
                    
                # Start with base budget
                budget_atual = budget_base
                
                # Apply global inflation to base budget if necessary
                ano_atual = i // 12
                if ano_atual > 0:
                    budget_atual = budget_base * ((1 + taxa_ipca_mensal) ** i)
                
                # Process each expense category
                for expense_name, array in expense_arrays.items():
                    # Get base value
                    base_value = None
                    if expense_name == "Água e Luz":
                        base_value = budget_atual * p['perc_agua_luz'] / 100
                    elif expense_name == "Aluguéis, Condomínios e IPTU":
                        base_value = budget_atual * p['perc_aluguel_condominio_iptu'] / 100
                    elif expense_name == "Internet":
                        base_value = budget_atual * p['perc_internet'] / 100
                    elif expense_name == "Material de Escritório":
                        base_value = budget_atual * p['perc_material_escritorio'] / 100
                    elif expense_name == "Treinamentos":
                        base_value = budget_atual * p['perc_treinamentos'] / 100
                    elif expense_name == "Manutenção & Conservação":
                        base_value = budget_atual * p['perc_manutencao_conservacao'] / 100
                    elif expense_name == "Seguros Funcionários":
                        base_value = budget_atual * p['perc_seguros_funcionarios'] / 100
                    elif expense_name == "Licenças de Telefonia":
                        base_value = budget_atual * p['perc_licencas_telefonia'] / 100
                    elif expense_name == "Licenças CRM":
                        base_value = budget_atual * p['perc_licencas_crm'] / 100
                    elif expense_name == "Telefônica":
                        base_value = budget_atual * p['perc_telefonica'] / 100
                    
                    # Apply specific reajuste if needed
                    adjust, index_type = should_adjust(expense_name, i)
                    
                    if adjust and base_value is not None:
                        if index_type == "IPCA Médio Anual (%)":
                            # Calculate the adjustment factor based on IPCA
                            factor = (1 + taxa_ipca_mensal)
                            base_value *= factor
                        elif index_type == "IGP-M Médio Anual (%)":
                            # Calculate the adjustment factor based on IGP-M
                            factor = (1 + taxa_igpm_mensal)
                            base_value *= factor
                    
                    array.append(base_value)
        else:
            # Process for Nominal mode
            # Base values
            consumo_kwh_base = p['consumo_mensal_kwh']
            aluguel_base = p['aluguel'] + p['condominio'] + p['iptu']
            internet_base = p['internet']
            material_escritorio_base = p['material_escritorio']
            treinamentos_base = p['treinamentos']
            manutencao_conservacao_base = p['manutencao_conservacao']
            seguros_funcionarios_base = p['seguros_funcionarios']
            licencas_telefonia_base = p['licencas_telefonia']
            licencas_crm_base = p['licencas_crm']
            telefonica_base = p['telefonica']
            
            # Mapping of expense names to base values
            base_values = {
                "Água e Luz": consumo_kwh_base,
                "Aluguéis, Condomínios e IPTU": aluguel_base,
                "Internet": internet_base,
                "Material de Escritório": material_escritorio_base,
                "Treinamentos": treinamentos_base,
                "Manutenção & Conservação": manutencao_conservacao_base,
                "Seguros Funcionários": seguros_funcionarios_base,
                "Licenças de Telefonia": licencas_telefonia_base,
                "Licenças CRM": licencas_crm_base,
                "Telefônica": telefonica_base
            }
            
            for i in range(meses):
                # Apply zeros for months before mes_inicio
                if i < mes_inicio:
                    for expense_name, array in expense_arrays.items():
                        array.append(0.0)
                    continue
                    
                # Process each expense category
                for expense_name, array in expense_arrays.items():
                    # Get base value
                    base_value = base_values[expense_name]
                    current_value = base_value
                    
                    # Apply standard inflation
                    fator_inflacao = (1 + taxa_ipca_mensal) ** i
                    current_value = base_value * fator_inflacao
                    
                    # Check if we need to apply a specific reajuste
                    adjust, index_type = should_adjust(expense_name, i)
                    
                    if adjust:
                        if index_type == "IPCA Médio Anual (%)":
                            # Already applied the IPCA, nothing more to do
                            pass
                        elif index_type == "IGP-M Médio Anual (%)":
                            # Replace the IPCA adjustment with IGP-M
                            fator_igpm = (1 + taxa_igpm_mensal) ** i
                            current_value = base_value * fator_igpm
                        elif index_type == "Sem reajuste":
                            # No adjustment, use the base value
                            current_value = base_value
                    
                    # Special case for Água e Luz to apply bandeira tarifária
                    if expense_name == "Água e Luz":
                        current_value = self._calcular_valor_agua_luz(current_value, i, p['modo_energia'])
                    
                    array.append(current_value)
        
        # Adicionar arrays ao DataFrame
        df['Água, Luz'] = agua_luz
        df['Aluguéis, Condomínios e IPTU'] = aluguel_condominio_iptu
        df['Internet'] = internet
        df['Material de Escritório'] = material_escritorio
        df['Treinamentos'] = treinamentos
        df['Manutenção & Conservação'] = manutencao_conservacao
        df['Seguros Funcionarios'] = seguros_funcionarios
        df['Lincenças de Telefonia'] = licencas_telefonia
        df['Lincenças CRM'] = licencas_crm
        df['Telefonica'] = telefonica
        
        # Calcular o total
        df['Total'] = df.sum(axis=1)
        
        return df
    
    def _gerar_dataframe_anual(self, df_mensal):
        """Converte o dataframe mensal para anual"""
        if df_mensal.empty:
            return pd.DataFrame()
        
        # Criar coluna de ano
        df_mensal['Ano'] = df_mensal.index // 12 + 1
        
        # Agrupar por ano e somar
        df_anual = df_mensal.groupby('Ano').sum()
        df_anual.index.name = 'Ano'
        
        return df_anual
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        st.write("### Fluxo de Despesas Administrativas")
        st.write("Visualização das despesas administrativas ao longo do tempo.")
        
        # Gerar o dataframe de despesas
        df_despesas = self._gerar_dataframe_despesas()
        
        # Verificar se o dataframe foi gerado corretamente
        if not df_despesas.empty:
            # Opções de visualização
            col1, col2 = st.columns(2)
            
            with col1:
                time_frame = st.selectbox("Período", self.time_frames, index=0)
            
            with col2:
                plot_type = st.selectbox("Tipo de Gráfico", self.graph_types, index=0)
            
            # Converter para anual se selecionado
            df_display = df_despesas.copy()
            if time_frame == "Anual":
                df_display = self._gerar_dataframe_anual(df_despesas)
                
            # Mostrar o dataframe de despesas
            st.write(f"#### Tabela de Despesas - {time_frame}")
            st.dataframe(df_display.style.format("{:.2f}"), use_container_width=True)
            
            # Plotar gráficos
            st.write(f"#### Visualização Gráfica - {time_frame}")
            
            # Seleção de despesas para visualização
            all_columns = df_display.columns.tolist()
            
            # Não remover 'Total' das opções, apenas garantir que apareça como primeira opção por default
            selected_items = st.multiselect(
                "Selecione os itens para visualizar", 
                all_columns,
                default=["Total"] if "Total" in all_columns else []
            )
            
            # Se nenhum item foi selecionado, mostrar apenas o total se disponível
            if not selected_items:
                if "Total" in all_columns:
                    selected_items = ["Total"]
                elif len(all_columns) > 0:
                    selected_items = [all_columns[0]]  # Selecionar o primeiro item disponível
            
            # Preparar dados para o gráfico
            df_plot = df_display[selected_items].reset_index()
            
            # Obter o nome da coluna do índice após o reset (será 'Mês' ou 'Ano')
            index_col = df_plot.columns[0]  # A primeira coluna após reset_index é a coluna do índice original
                
            if plot_type == "Gráfico de Linhas":
                # Gráfico de linha para evolução temporal
                fig = px.line(
                    df_plot,
                    x=index_col,  # Usar o nome real da coluna do índice
                    y=selected_items,
                    title=f"Evolução das Despesas por {time_frame}",
                    markers=True
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                # Gráfico de pizza para composição
                # Usar o último período para o gráfico de pizza
                last_period = df_plot.iloc[-1]
                data_pie = {
                    'Categoria': selected_items,
                    'Valor': [last_period[item] for item in selected_items]
                }
                df_pie = pd.DataFrame(data_pie)
                
                periodo_singular = "Mês" if time_frame == "Mensal" else "Ano"
                fig = px.pie(
                    df_pie,
                    values='Valor',
                    names='Categoria',
                    title=f"Composição das Despesas - Último {periodo_singular}"
                )
                st.plotly_chart(fig, use_container_width=True)
            
            # Gráfico adicional mostrando a evolução do total
            if "Total" in df_display.columns and len(selected_items) > 1:
                st.write(f"#### Evolução do Total de Despesas - {time_frame}")
                
                # Reset index e pegar o nome da coluna do índice
                df_total = df_display.reset_index()
                index_col_total = df_total.columns[0]  # A primeira coluna após reset_index
                
                fig_total = px.line(
                    df_total,
                    x=index_col_total,  # 'mês' ou 'ano'
                    y=["Total"],
                    title=f"Evolução do Total de Despesas por {time_frame}",
                    markers=True
                )
                st.plotly_chart(fig_total, use_container_width=True)
        else:
            st.error("Não foi possível gerar as despesas. Verifique os parâmetros nas premissas.")

class Equipe(Page):
    def __init__(self):
        pass
    
    @property
    def title(self) -> str:
        return "Equipe"
    
    @property
    def icon(self) -> str:
        return "👥"
    
    def _calcular_inflacao_anual(self, valor_base, mes, ipca_anual):
        """Calcula o valor corrigido pela inflação anual"""
        ano = (mes + 1) // 12  # +1 porque o mês 0 já é o primeiro mês
        # Só aplica inflação a partir do segundo ano (após 12 meses)
        if ano >= 1:
            fator_inflacao = (1 + ipca_anual / 100) ** ano
            return valor_base * fator_inflacao
        return valor_base
    
    # Add this to the _gerar_dataframe_custos_equipe method in Equipe class
    def _gerar_dataframe_custos_equipe(self):
        """Gera o dataframe completo com os custos da equipe ao longo de 5 anos"""
        if 'premissas_despesas' not in st.session_state:
            st.error("Premissas de despesas não definidas")
            return pd.DataFrame()
        
        p = st.session_state['premissas_despesas']
        ipca = p['ipca_medio_anual']
        igpm = p['igpm_medio_anual']
        reajustes = p.get('reajustes', {})
        
        # Monthly rates
        taxa_ipca_mensal = (1 + ipca / 100) ** (1/12) - 1
        taxa_igpm_mensal = (1 + igpm / 100) ** (1/12) - 1
        
        # Function to check if a expense should be adjusted in a specific month
        def should_adjust(expense_name, month):
            # Check monthly adjustments
            if 'mensal' in reajustes and isinstance(reajustes['mensal'], dict) and expense_name in reajustes['mensal'].get('expenses', []):
                return True, reajustes['mensal']['index']
            
            # Check quarterly adjustments
            if 'trimestral' in reajustes and isinstance(reajustes['trimestral'], dict) and expense_name in reajustes['trimestral'].get('expenses', []) and month % 3 == 0:
                return True, reajustes['trimestral']['index']
            
            # Check semi-annual adjustments
            if 'semestral' in reajustes and isinstance(reajustes['semestral'], dict) and expense_name in reajustes['semestral'].get('expenses', []) and month % 6 == 0:
                return True, reajustes['semestral']['index']
            
            # Check annual adjustments - Fix: Check for annual adjustments at the beginning of each year (month 12, 24, etc.)
            if 'anual' in reajustes and isinstance(reajustes['anual'], dict) and expense_name in reajustes['anual'].get('expenses', []):
                # Apply at the start of each year (months 12, 24, 36, 48)
                if month > 0 and month % 12 == 0:  # Changed condition to catch annual adjustments
                    return True, reajustes['anual']['index']
            
            return False, None
        
        # Criar dataframe para 5 anos (60 meses)
        total_meses = 60
        meses = list(range(1, total_meses + 1))
        
        # Get roles from session state
        equipe_propria = p.get('equipe_propria', [])
        terceiros = p.get('terceiros', [])
        
        # Create dynamic indices based on available roles
        indices = []
        
        # Add Equipe Própria roles
        for cargo in equipe_propria:
            indices.append(("Equipe Própria", f"Salário {cargo['nome']}"))
        
        # Add fixed categories
        indices.append(("Encargos Sociais", "Encargos Sociais"))
        indices.append(("Despesas com Alimentação e Transporte", "Alimentação e Transporte"))
        
        # Add Terceiros roles
        for prestador in terceiros:
            indices.append(("Terceiros - Prestadores de Serviços", prestador['nome']))
        
        # Add Bônus dos Lucros for Equipe Própria
        for cargo in equipe_propria:
            indices.append(("Bônus dos Lucros", f"Bônus {cargo['nome']}"))
        
        # Add total
        indices.append(("TOTAL", "Total Custos de Equipe"))
        
        # Create DataFrame with zeros
        idx = pd.MultiIndex.from_tuples(indices)
        df = pd.DataFrame(0, index=idx, columns=pd.Index(meses))
        
        # Apply cumulative inflation factors based on adjustments
        equipe_propria_inflation = [1.0] * total_meses  # Base factor starts at 1.0
        alimentacao_inflation = [1.0] * total_meses
        terceiros_inflation = [1.0] * total_meses
        
        # Pre-calculate inflation factors for each category and month
        for mes in range(1, total_meses + 1):
            # Get the previous month's factor (or 1.0 for the first month)
            prev_equipe = equipe_propria_inflation[mes-1] if mes > 1 else 1.0
            prev_alimentacao = alimentacao_inflation[mes-1] if mes > 1 else 1.0
            prev_terceiros = terceiros_inflation[mes-1] if mes > 1 else 1.0
            
            # Check for adjustments and apply them
            adjust_equipe, index_equipe = should_adjust("Equipe Própria", mes-1)
            adjust_alim, index_alim = should_adjust("Alimentação e Transporte", mes-1)
            adjust_terc, index_terc = should_adjust("Terceiros", mes-1)
            
            # Apply the appropriate inflation factor
            if adjust_equipe:
                if index_equipe == "IPCA Médio Anual (%)":
                    equipe_propria_inflation[mes-1] = prev_equipe * (1 + taxa_ipca_mensal)
                elif index_equipe == "IGP-M Médio Anual (%)":
                    equipe_propria_inflation[mes-1] = prev_equipe * (1 + taxa_igpm_mensal)
            
            if adjust_alim:
                if index_alim == "IPCA Médio Anual (%)":
                    alimentacao_inflation[mes-1] = prev_alimentacao * (1 + taxa_ipca_mensal)
                elif index_alim == "IGP-M Médio Anual (%)":
                    alimentacao_inflation[mes-1] = prev_alimentacao * (1 + taxa_igpm_mensal)
            
            if adjust_terc:
                if index_terc == "IPCA Médio Anual (%)":
                    terceiros_inflation[mes-1] = prev_terceiros * (1 + taxa_ipca_mensal)
                elif index_terc == "IGP-M Médio Anual (%)":
                    terceiros_inflation[mes-1] = prev_terceiros * (1 + taxa_igpm_mensal)
            
            # For annual adjustments at the beginning of each year (month 12, 24, etc.)
            if mes > 1 and (mes - 1) % 12 == 0:
                
                # Annual adjustment for Equipe Própria if configured
                if "Equipe Própria" in reajustes.get('anual', {}).get('expenses', []):
                    index_type = reajustes['anual']['index']
                    if index_type == "IPCA Médio Anual (%)":
                        equipe_propria_inflation[mes-1] = prev_equipe * (1 + ipca/100)
                    elif index_type == "IGP-M Médio Anual (%)":
                        equipe_propria_inflation[mes-1] = prev_equipe * (1 + igpm/100)
                
                # Annual adjustment for Alimentação e Transporte if configured
                if "Alimentação e Transporte" in reajustes.get('anual', {}).get('expenses', []):
                    index_type = reajustes['anual']['index']
                    if index_type == "IPCA Médio Anual (%)":
                        alimentacao_inflation[mes-1] = prev_alimentacao * (1 + ipca/100)
                    elif index_type == "IGP-M Médio Anual (%)":
                        alimentacao_inflation[mes-1] = prev_alimentacao * (1 + igpm/100)
                
                # Annual adjustment for Terceiros if configured
                if "Terceiros" in reajustes.get('anual', {}).get('expenses', []):
                    index_type = reajustes['anual']['index']
                    if index_type == "IPCA Médio Anual (%)":
                        terceiros_inflation[mes-1] = prev_terceiros * (1 + ipca/100)
                    elif index_type == "IGP-M Médio Anual (%)":
                        terceiros_inflation[mes-1] = prev_terceiros * (1 + igpm/100)
            
            # Copy the inflation factor to the next month if not adjusted
            if mes < total_meses:
                equipe_propria_inflation[mes] = equipe_propria_inflation[mes-1]
                alimentacao_inflation[mes] = alimentacao_inflation[mes-1]
                terceiros_inflation[mes] = terceiros_inflation[mes-1]
        
        # Preencher valores para Equipe Própria
        for mes in meses:
            equipe_propria_total = 0
            
            # Process each role from equipe_propria
            for cargo in equipe_propria:
                nome_cargo = cargo['nome']
                 
                # Calculate salary with inflation adjustment
                if p['equipe_modo_calculo'] == "Percentual":
                    # Calculate salary based on budget and percentage
                    budget = p['budget_equipe_propria']
                    percentual = cargo['percentual']
                    salario_base = (percentual / 100 * budget) / cargo['quantidade'] if cargo['quantidade'] > 0 else 0
                else:
                    # Use nominal salary
                    salario_base = cargo['salario'] 
                 
                # Apply inflation adjustment from our pre-calculated factors
                inflation_factor = equipe_propria_inflation[mes-1] 
                salario = salario_base * inflation_factor
                 
                # Calculate total salary cost for this role      
                custo_salario = salario * cargo['quantidade']
                df.loc[("Equipe Própria", f"Salário {nome_cargo}"), mes] = custo_salario
                
                # Add to team total
                equipe_propria_total += custo_salario      
            
            # Calculate social security charges
            encargos = equipe_propria_total * (p['encargos_sociais_perc'] / 100)
            df.loc[("Encargos Sociais", "Encargos Sociais"), mes] = encargos
            
            # Calculate food and transportation benefits with inflation adjustment
            total_beneficios = 0
            roles_com_beneficios = p.get('roles_com_beneficios', [])
            
            for cargo in equipe_propria:
                if cargo['nome'] in roles_com_beneficios:
                    # 20 dias úteis por mês
                    beneficio_pessoa = (p['vale_alimentacao'] * 20) + (p['vale_transporte'] * 20)
                    
                    # Apply inflation adjustment from pre-calculated factors
                    inflation_factor = alimentacao_inflation[mes-1]
                    beneficio_pessoa *= inflation_factor
                    
                    total_beneficios += beneficio_pessoa * cargo['quantidade']
            
            df.loc[("Despesas com Alimentação e Transporte", "Alimentação e Transporte"), mes] = total_beneficios
            
            # Process each third-party service provider with inflation adjustment
            for prestador in terceiros:
                nome_prestador = prestador['nome']
                
                # Calculate service cost with inflation adjustment
                if p['equipe_modo_calculo'] == "Percentual":
                    # Calculate value based on budget and percentage
                    budget = p['budget_terceiros']
                    percentual = prestador['percentual']
                    valor_base = (percentual / 100 * budget) / prestador['quantidade'] if prestador['quantidade'] > 0 else 0
                else:
                    # Use nominal value
                    valor_base = prestador['valor']
                
                # Apply inflation from pre-calculated factors
                inflation_factor = terceiros_inflation[mes-1]
                valor = valor_base * inflation_factor
                
                # Calculate total cost for this service provider
                custo_servico = valor * prestador['quantidade']
                df.loc[("Terceiros - Prestadores de Serviços", nome_prestador), mes] = custo_servico
            
            # Calculate bonuses - only in January of each year after year 1
            if mes > 12 and mes % 12 == 1:  # January of each year after year 1
                # Calculate previous year's profit growth
                ano_atual = (mes - 1) // 12
                lucro_liquido_ano_anterior = p['lucro_liquido_inicial'] * ((1 + p['crescimento_lucro'] / 100) ** (ano_atual - 1))
                lucro_liquido_atual = p['lucro_liquido_inicial'] * ((1 + p['crescimento_lucro'] / 100) ** ano_atual)
                
                # Calculate percentage growth
                crescimento_percentual = ((lucro_liquido_atual / lucro_liquido_ano_anterior) - 1) * 100
                
                # Check if growth exceeded benchmark
                if crescimento_percentual > p['benchmark_anual_bonus']:
                    # Calculate excess value
                    valor_excedente = lucro_liquido_atual - (lucro_liquido_ano_anterior * (1 + p['benchmark_anual_bonus'] / 100))
                    
                    # Calculate bonus for each role
                    for cargo in equipe_propria:
                        nome_cargo = cargo['nome']
                        bonus_key = f"bonus_{cargo['nome'].lower().replace(' ', '_')}"
                        bonus_perc = p.get(bonus_key, 1.0)
                        
                        # Apply bonus percentage to excess value
                        bonus_valor = valor_excedente * (bonus_perc / 100)
                        
                        # Store the bonus value
                        df.loc[("Bônus dos Lucros", f"Bônus {nome_cargo}"), mes] = bonus_valor
            
            # Calculate total cost for this month
            total_mes = 0
            
            # Sum all categories except TOTAL
            for idx in df.index:
                if idx[0] != "TOTAL":
                    value = pd.to_numeric(df.loc[idx, mes], errors='coerce')
                    if pd.notna(value):
                        total_mes += float(value)
            
            # Store the total
            df.loc[("TOTAL", "Total Custos de Equipe"), mes] = total_mes
        
        return df

    def _gerar_df_anual(self, df_mensal):
        """Converte o dataframe mensal para anual"""
        anos = 5
        df_anual = pd.DataFrame(0, index=df_mensal.index, columns=range(1, anos + 1))
        
        for ano in range(1, anos + 1):
            mes_inicio = (ano - 1) * 12 + 1
            mes_fim = ano * 12
            
            # Para cada item, somar os valores dos meses do ano
            for idx in df_mensal.index:
                df_anual.loc[idx, ano] = df_mensal.loc[idx, mes_inicio:mes_fim].sum()
        
        return df_anual
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        if 'premissas_despesas' not in st.session_state:
            st.error("Premissas de despesas não definidas. Por favor, defina as premissas na página 'Premissas Despesas'.")
            return
        
        # Gerar DataFrame com os custos da equipe
        df_mensal = self._gerar_dataframe_custos_equipe()
        df_anual = self._gerar_df_anual(df_mensal)
        
        # Criar tabs para visualizar os dados de acordo com diferentes timeframes
        tab1, tab2, tab3, tab4 = st.tabs(["Visão Mensal", "Visão Trimestral", "Visão Semestral", "Visão Anual"])
    
        with tab1:
            self._render_timeframe_view(df_mensal, "Mensal", 1)
    
        with tab2:
            self._render_timeframe_view(df_mensal, "Trimestral", 3)
    
        with tab3:
            self._render_timeframe_view(df_mensal, "Semestral", 6)
    
        with tab4:
            self._render_timeframe_view(df_anual, "Anual", 12, is_annual_df=True)
    
    def _render_timeframe_view(self, df, timeframe_name, months_per_period, is_annual_df=False):
        """Render a view for a specific timeframe"""
        st.write(f"### Custos da Equipe - Visão {timeframe_name}")
        
        # For annual df, display as is; for other timeframes, group by period
        if is_annual_df:
            df_display = df
            df_display.columns = [f"Ano {ano}" for ano in df_display.columns]
        else:
            # Group data into periods (trimestres, semestres)
            df_display = self._agrupar_por_periodo(df, months_per_period)
        
        # Display the dataframe
        st.dataframe(df_display.style.format("{:.2f}"), use_container_width=True)
        
        # Gráficos
        st.write(f"### Visualização Gráfica ({timeframe_name})")
        
        # Get available categories for selection
        if is_annual_df:
            available_categories = [(idx[0], idx[1]) for idx in df.index]
        else:
            available_categories = [(idx[0], idx[1]) for idx in df.index]
        
        # Convert to a list of tuples for the multiselect
        category_options = [(cat[0], cat[1]) for cat in available_categories]
        
        # Default selection is the Total category
        default_selection = [cat for cat in category_options if cat[0] == "TOTAL"]
        if not default_selection and category_options:
            default_selection = [category_options[0]]
        
        # Add multiselect for categories
        selected_categories = st.multiselect(
            "Selecione as categorias para visualizar:",
            options=category_options,
            default=default_selection,
            format_func=lambda x: f"{x[0]} - {x[1]}",
            key=f"{timeframe_name.lower()}_categories"
        )
        
        # If no selection, use the total by default
        if not selected_categories:
            selected_categories = [cat for cat in category_options if cat[0] == "TOTAL"]
            if not selected_categories and category_options:
                selected_categories = [category_options[0]]
        
        # Prepare data for plotting
        if selected_categories:
            plot_data = pd.DataFrame()
            for cat in selected_categories:
                try:
                    series = df_display.loc[cat]
                    plot_data[f"{cat[0]} - {cat[1]}"] = series
                except KeyError:
                    st.warning(f"Categoria {cat} não encontrada no DataFrame")
            
            # Add period as column instead of index
            plot_data = plot_data.reset_index()
            if is_annual_df:
                plot_data.rename(columns={'index': 'Ano'}, inplace=True)
            else:
                plot_data.rename(columns={'index': 'Período'}, inplace=True)
            
            # Create line chart with selected categories
            if not plot_data.empty:
                x_axis = 'Ano' if is_annual_df else 'Período'
                fig = px.line(
                    plot_data,
                    x=x_axis,
                    y=plot_data.columns[1:],  # Skip the Ano/Período column
                    title=f"Custos da Equipe - Visão {timeframe_name}",
                    labels={'value': 'Valor (R$)', 'variable': 'Categoria'},
                    markers=True
                )
                st.plotly_chart(fig, use_container_width=True)
    
    def _agrupar_por_periodo(self, df, months_per_period):
        """Agrupa o dataframe mensal por período (trimestral ou semestral)"""
        # Determine the period label based on months_per_period
        if months_per_period == 3:
            periodo_nome = "T"  # Trimestre
        elif months_per_period == 6:
            periodo_nome = "S"  # Semestre
        else:
            periodo_nome = "M"  # Mês (mantendo o padrão original)
        
        # Create a new DataFrame for grouped data
        df_agrupado = pd.DataFrame(index=df.index)
        
        # Group data by period
        for i in range(0, 60, months_per_period):
            end_idx = min(i + months_per_period, 60)
            period_label = f"{periodo_nome}{i//months_per_period + 1}"
            
            # Calculate average for this period
            period_data = df.loc[:, i+1:end_idx].mean(axis=1)
            df_agrupado[period_label] = period_data
        
        return df_agrupado

# Replace the existing CustosTecnologia class with this implementation:
class CustosTecnologia(Page):
    def __init__(self):
        self.graph_types = ["Gráfico de Linhas", "Gráfico de Pizza"]
        self.time_frames = ["Mensal", "Anual"]
    
    @property
    def title(self) -> str:
        return "Custos de Tecnologia"
    
    @property
    def icon(self) -> str:
        return "💻"
    
    def _calcular_inflacao_anual(self, valor_base, mes, ipca_anual):
        """Calcula o valor corrigido pela inflação anual"""
        ano = mes // 12  # Determina em qual ano o mês se encontra
        # Aplica inflação conforme o ano
        if ano >= 1:
            fator_inflacao = (1 + ipca_anual / 100) ** ano
            return valor_base * fator_inflacao
        return valor_base
    
    # Add this to the _gerar_dataframe_custos_tecnologia method in CustosTecnologia class
    def _gerar_dataframe_custos_tecnologia(self):
        """Gera o dataframe com os custos de tecnologia para 60 meses (5 anos)"""
        if 'premissas_despesas' not in st.session_state:
            return pd.DataFrame()
        
        p = st.session_state['premissas_despesas']
        ipca = p.get('ipca_medio_anual', 4.5)  # Valor padrão de 4.5% se não estiver definido
        igpm = p.get('igpm_medio_anual', 5.0)  # Valor padrão de 5.0% se não estiver definido
        reajustes = p.get('reajustes', {})
        
        # Monthly rates
        taxa_ipca_mensal = (1 + ipca / 100) ** (1/12) - 1
        taxa_igpm_mensal = (1 + igpm / 100) ** (1/12) - 1
        
        # Valores base mensais
        desenvolvimento_ferramenta = p.get('desenvolvimento_ferramenta', 0.0)
        manutencao_ferramenta = p.get('manutencao_ferramenta', 0.0)
        inovacao = p.get('inovacao', 0.0)
        licencas_software = p.get('licencas_software', 2513.0)
        
        # Function to check if a expense should be adjusted in a specific month
        def should_adjust(expense_name, month):
            # Check monthly adjustments
            if 'mensal' in reajustes and isinstance(reajustes['mensal'], dict) and expense_name in reajustes['mensal'].get('expenses', []):
                return True, reajustes['mensal']['index']
            
            # Check quarterly adjustments
            if 'trimestral' in reajustes and isinstance(reajustes['trimestral'], dict) and expense_name in reajustes['trimestral'].get('expenses', []) and month % 3 == 0:
                return True, reajustes['trimestral']['index']
            
            # Check semi-annual adjustments
            if 'semestral' in reajustes and isinstance(reajustes['semestral'], dict) and expense_name in reajustes['semestral'].get('expenses', []) and month % 6 == 0:
                return True, reajustes['semestral']['index']
            
            # Check annual adjustments
            if 'anual' in reajustes and isinstance(reajustes['anual'], dict) and expense_name in reajustes['anual'].get('expenses', []) and month % 12 == 0:
                return True, reajustes['anual']['index']
            
            return False, None
        
        # Criar DataFrame para 60 meses (0 a 59)
        meses = range(60)
        index_labels = ["Desenvolvimento da ferramenta", 
                        "Manutenção da ferramenta", 
                        "Inovação", 
                        "Licenças de software",
                        "Aquisição de Equipamentos",
                        "Depreciação de Equipamentos",
                        "Total"]
        df = pd.DataFrame(index=pd.Index(index_labels), 
                          columns=meses)
        df.fillna(0.0, inplace=True)  # Initialize with zeros
        
        # Mapping of expense names to values
        expense_values = {
            "Desenvolvimento da ferramenta": desenvolvimento_ferramenta,
            "Manutenção da ferramenta": manutencao_ferramenta,
            "Inovação": inovacao,
            "Licenças de software": licencas_software
        }
        
        # Preencher valores para cada mês
        for mes in meses:
            # Process each expense
            for expense_name, base_value in expense_values.items():
                current_value = base_value
                
                # Check if we need to apply a specific reajuste
                adjust, index_type = should_adjust(expense_name, mes)
                
                if adjust:
                    if index_type == "IPCA Médio Anual (%)":
                        current_value = base_value * (1 + taxa_ipca_mensal)
                    elif index_type == "IGP-M Médio Anual (%)":
                        current_value = base_value * (1 + taxa_igpm_mensal)
                    # else: "Sem reajuste" - keep the base value
                else:
                    # Default inflation adjustment
                    current_value = self._calcular_inflacao_anual(base_value, mes, ipca)
                
                # Store the value in DataFrame
                df.loc[expense_name, mes] = current_value
            
            # Calculate equipment acquisition costs
            if 'equipamentos' in p:
                equipment_cost = 0
                for equip in p['equipamentos']:
                    mes_aquisicao = equip.get('mes_aquisicao', 0)
                    if mes == mes_aquisicao:
                        equipment_cost += equip['valor'] * equip['quantidade']
                df.loc["Aquisição de Equipamentos", mes] = equipment_cost
            
            # Calculate equipment depreciation
            if 'equipamentos' in p:
                depreciation_value = 0.0
                for equip in p['equipamentos']:
                    mes_aquisicao = equip.get('mes_aquisicao', 0)
                    
                    # Only apply depreciation after acquisition
                    if mes >= mes_aquisicao:
                        # Different depreciation methods
                        if equip['metodo'] == "Método da Linha Reta":
                            # Calculate monthly depreciation (annual / 12)
                            monthly_depreciation = equip['metodo_params']['depreciacao_anual'] * equip['quantidade'] / 12
                            depreciation_value += monthly_depreciation
                        else:  # Método da Soma dos Dígitos
                            # Calculate how many months have passed since acquisition
                            months_since_acquisition = mes - mes_aquisicao
                            # Calculate which year we're in (1-based)
                            year_since_acquisition = (months_since_acquisition // 12) + 1
                            
                            # Only apply depreciation for valid years
                            if year_since_acquisition <= len(equip['metodo_params']['depreciacao_anos']):
                                # Get annual depreciation for this year
                                annual_depreciation = equip['metodo_params']['depreciacao_anos'][year_since_acquisition] * equip['quantidade']
                                # Convert to monthly
                                monthly_depreciation = annual_depreciation / 12
                                depreciation_value += monthly_depreciation
                
                df.loc["Depreciação de Equipamentos", mes] = depreciation_value
            
            # Calculate total for this month - include all columns including new ones
            df.loc["Total", mes] = df.loc[df.index != 'Total', mes].sum()
        
        return df
    
    def _gerar_dataframe_anual(self, df_mensal):
        """Converte o dataframe mensal para anual"""
        if df_mensal.empty:
            return pd.DataFrame()
        
        anos = 5
        df_anual = pd.DataFrame(index=df_mensal.index, columns=range(1, anos + 1))
        
        for ano in range(1, anos + 1):
            mes_inicio = (ano - 1) * 12
            mes_fim = ano * 12 - 1
            
            # Para cada categoria, somar os valores dos meses do ano
            for categoria in df_mensal.index:
                df_anual.loc[categoria, ano] = df_mensal.loc[categoria, mes_inicio:mes_fim].sum()
        
        return df_anual
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        if 'premissas_despesas' not in st.session_state:
            st.error("Premissas de despesas não definidas. Por favor, defina as premissas na página 'Premissas Despesas'.")
            return
        
        # Gerar DataFrame com os custos de tecnologia
        df_mensal = self._gerar_dataframe_custos_tecnologia()
        
        if not df_mensal.empty:
            # Opções de visualização
            col1, col2 = st.columns(2)
            
            with col1:
                time_frame = st.selectbox("Período", self.time_frames, index=0)
            
            with col2:
                plot_type = st.selectbox("Tipo de Gráfico", self.graph_types, index=0)
            
             # Converter para anual se selecionado
            df_display = df_mensal.copy()
            if time_frame == "Anual":
                df_display = self._gerar_dataframe_anual(df_mensal)
                if df_display is not None:
                    df_display.columns = [f"Ano {ano}" for ano in df_display.columns]
            else:
                if df_display is not None:
                    df_display.columns = [f"Mês {mes}" for mes in df_display.columns]
           
            # Mostrar o dataframe de custos de tecnologia
            periodo = "Anual" if time_frame == "Anual" else "Mensal"
            st.write(f"### Custos de Tecnologia - Visão {periodo}")
            st.dataframe(df_display.style.format("{:.2f}"), use_container_width=True)
            
            # Seleção de categorias para visualização
            categorias = df_display.index.tolist()
            
            # Não remover 'Total' das opções, apenas garantir que apareça como primeira opção por default
            selected_categories = st.multiselect(
                "Selecione as categorias para visualizar", 
                categorias,
                default=["Total"] if "Total" in categorias else []
            )
            
            # Se nenhuma categoria foi selecionada, mostrar apenas o total
            if not selected_categories:
                if "Total" in categorias:
                    selected_categories = ["Total"]
                elif len(categorias) > 0:
                    selected_categories = [categorias[0]]
            
            # Preparar dados para o gráfico
            if selected_categories and df_display is not None:
                df_plot = df_display.loc[selected_categories].T  # Transpor para ter meses/anos como índice
                
                if plot_type == "Gráfico de Linhas":
                    # Gráfico de linha para evolução temporal
                    fig = px.line(
                        df_plot,
                        y=selected_categories,
                        title=f"Evolução dos Custos de Tecnologia ({periodo})",
                        markers=True
                    )
                    fig.update_layout(
                        xaxis_title=f"{'Ano' if time_frame == 'Anual' else 'Mês'}",
                        yaxis_title="Valor (R$)",
                        legend_title="Categoria"
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    # Gráfico de pizza para composição
                    # Usar o último período para o gráfico de pizza
                    last_period = df_plot.iloc[-1]
                    data_pie = {
                        'Categoria': selected_categories,
                        'Valor': [last_period[cat] for cat in selected_categories]
                    }
                    df_pie = pd.DataFrame(data_pie)
                    
                    periodo_singular = "Ano" if time_frame == "Anual" else "Mês"
                    fig = px.pie(
                        df_pie,
                        values='Valor',
                        names='Categoria',
                        title=f"Composição dos Custos de Tecnologia - Último {periodo_singular}"
                    )
                    st.plotly_chart(fig, use_container_width=True)
                    
                # Métricas principais
                st.write("### Métricas Principais")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    # Get DRE data for resultado do exercício
                    try:
                        dre = ProjeçãoDRE()
                        df_dre = dre._gerar_dataframe_dre()
                        if df_dre is not None and ("10", "(=) Resultado do Exercício") in df_dre.index:
                            # Get first 12 months and sum
                            resultado_12m = df_dre.loc[("10", "(=) Resultado do Exercício"), :11].sum()
                            st.metric(
                                label="Resultado do Exercício Acumulado (12 meses)",
                                value=f"R$ {resultado_12m:,.2f}"
                            )
                        else:
                            st.metric(
                                label="Resultado do Exercício Acumulado (12 meses)",
                                value="Dados indisponíveis"
                            )
                    except Exception as e:
                        st.metric(
                            label="Resultado do Exercício Acumulado (12 meses)",
                            value="Erro ao calcular"
                        )
                
                with col2:
                    # Get DRE data for EBITDA
                    # Initialize df_dre first
                    df_dre = None
                    try:
                        if df_dre is not None and ("6", "(=) Resultado Operacional (EBITDA/LAJIDA)") in df_dre.index:
                            # Get first 12 months and sum
                            ebitda_12m = df_dre.loc[("6", "(=) Resultado Operacional (EBITDA/LAJIDA)"), :11].sum()
                            st.metric(
                                label="EBITDA Acumulado (12 meses)",
                                value=f"R$ {ebitda_12m:,.2f}"
                            )
                        else:
                            st.metric(
                                label="EBITDA Acumulado (12 meses)",
                                value="Dados indisponíveis"
                            )
                    except Exception as e:
                        st.metric(
                            label="EBITDA Acumulado (12 meses)",
                            value="Erro ao calcular"
                        )
                
                with col3:
                    # Calculate Margem Líquida Média (12 meses)
                    try:
                        if df_dre is not None and ("10", "(=) Resultado do Exercício") in df_dre.index and ("1", "Receita Bruta de Vendas") in df_dre.index:
                            # Get first 12 months of resultado and receita bruta
                            resultado_12m = df_dre.loc[("10", "(=) Resultado do Exercício"), :11].sum()
                            receita_12m = df_dre.loc[("1", "Receita Bruta de Vendas"), :11].sum()
                            
                            if receita_12m > 0:
                                margem_liquida = (resultado_12m / receita_12m) * 100
                                st.metric(
                                    label="Margem Líquida Média (12 meses)",
                                    value=f"{margem_liquida:.2f}%"
                                )
                            else:
                                st.metric(
                                    label="Margem Líquida Média (12 meses)",
                                    value="N/A"
                                )
                        else:
                            st.metric(
                                label="Margem Líquida Média (12 meses)",
                                value="Dados indisponíveis"
                            )
                    except Exception as e:
                        st.metric(
                            label="Margem Líquida Média (12 meses)",
                            value="Erro ao calcular"
                        )
        else:
            st.error("Não foi possível gerar os custos de tecnologia. Verifique os parâmetros nas premissas.")

class PremissasReceitas(Page):
    # In PremissasReceitas.__init__
    # In PremissasReceitas.__init__
    def __init__(self):
        # Default values for revenue parameters
        default_params = {
            # General configuration
            'modelo_marketing': True,
            'repasse_bruto': 85.0,
            'canais_venda': [],
            'outras_receitas': [],
            
            # Financial model parameters
            'receita_inicial': 100000.0,
            'valor_unitario': 2400.0,
            'crescimento_receita': 'Linear',
            'tx_cresc_mensal': 5.0,
            'media_cresc_anual': 15.0,
            'fator_crescimento': 0.5,  # New default for sigmoid growth
            'fator_estabilizacao': 0.8,  # New default for sigmoid growth
            'fontes_primarias': [],
            
            # New parameters for productivity model
            'rpe_anual': 125000.0,  # Revenue per employee per year
            'salario_medio': 60000.0,  # Average annual salary
            'depreciacao': 1.5,  # Monthly depreciation rate (%)
        }
        
        # Initialize or update session state
        if 'premissas_receitas' not in st.session_state:
            st.session_state['premissas_receitas'] = default_params
        else:
            # Check for any missing keys and add them with default values
            for key, value in default_params.items():
                if key not in st.session_state['premissas_receitas']:
                    st.session_state['premissas_receitas'][key] = value
    
    @property
    def title(self) -> str:
        return "Premissas Receitas"
    
    @property
    def icon(self) -> str:
        return "💲"
    
    def _get_all_conversion_parameters_from_equipe_propria(self):
        """Get conversion parameters from all team members in PremissasDespesas"""
        conversion_params_list = []
        
        # Check if we have team members
        if ('premissas_despesas' in st.session_state and 
            'equipe_propria' in st.session_state['premissas_despesas']):
            
            # Get all team members with conversion parameters
            for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                if all(param in cargo for param in ['fator_elasticidade', 'taxa_agendamento', 
                                                   'taxa_comparecimento', 'taxa_conversao', 'ticket_medio']):
                    conversion_params_list.append({
                        'fator_elasticidade': cargo.get('fator_elasticidade', 1.0),
                        'taxa_agendamento': cargo.get('taxa_agendamento', 30.0),
                        'taxa_comparecimento': cargo.get('taxa_comparecimento', 70.0),
                        'taxa_conversao': cargo.get('taxa_conversao', 45.0),
                        'ticket_medio': cargo.get('ticket_medio', 2400.0),
                        'quantidade': cargo.get('quantidade', 1)
                    })
        
        return conversion_params_list

    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Marketing model is always active now (remove checkbox)
        st.session_state['premissas_receitas']['modelo_marketing'] = True
        
        # Only show marketing model description
        st.info("**Modelo de Marketing**: Calcula receitas com base nos dados da equipe e conversões.")
        
        st.write("### Configurações de Repasse")
        st.write("Defina o percentual de repasse bruto aplicado às receitas.")
        
        # Keep the Repasse Bruto slider
        st.session_state['premissas_receitas']['repasse_bruto'] = st.slider(
            "Repasse Bruto (%)",
            min_value=1.0,
            max_value=100.0,
            value=st.session_state['premissas_receitas'].get('repasse_bruto', 85.0),
            step=0.5,
            format="%.1f",
            help="Percentual do valor da venda que é efetivamente recebido pela empresa."
        )
        
        # Always render marketing model
        self._render_marketing_model()
        
        # Show summary of parameters
        st.write("### Resumo das Premissas de Receitas")
        
        params = [
            ["Modelo", "Marketing"],
            ["Repasse Bruto (%)", st.session_state['premissas_receitas']['repasse_bruto']],
            ["Fonte de Dados", "Equipe definida em 'Premissas Despesas'"]
        ]
        
        df_summary = pd.DataFrame(params, columns=["Parâmetro", "Valor"])
        st.dataframe(df_summary, use_container_width=True)

    def _get_conversion_parameters_from_equipe_propria(self):
        """Get conversion parameters from team members in PremissasDespesas"""
        # Get premissas_receitas from session_state
        p_receitas = st.session_state.get('premissas_receitas', {})
        
        default_params = {
            'fator_elasticidade': 1.0,
            'taxa_agendamento': 30.0,
            'taxa_comparecimento': 70.0,
            'taxa_conversao': 45.0,
            'ticket_medio': 2400.0,
            'repasse_bruto': p_receitas.get('repasse_bruto', 85.0)  # Get repasse_bruto from p_receitas
        }
        
        # Get all conversion parameters
        conversion_params_list = self._get_all_conversion_parameters_from_equipe_propria()
        
        # If we have parameters, calculate the average values
        if conversion_params_list:
            total_params = {
                'fator_elasticidade': 0,
                'taxa_agendamento': 0,
                'taxa_comparecimento': 0,
                'taxa_conversao': 0,
                'ticket_medio': 0
            }
            
            total_quantity = sum(params.get('quantidade', 1) for params in conversion_params_list)
            
            # Calculate weighted average based on team member quantity
            for params in conversion_params_list:
                quantidade = params.get('quantidade', 1)
                weight = quantidade / total_quantity if total_quantity > 0 else 0
                
                total_params['fator_elasticidade'] += params.get('fator_elasticidade', 1.0) * weight
                total_params['taxa_agendamento'] += params.get('taxa_agendamento', 30.0) * weight
                total_params['taxa_comparecimento'] += params.get('taxa_comparecimento', 70.0) * weight
                total_params['taxa_conversao'] += params.get('taxa_conversao', 45.0) * weight
                total_params['ticket_medio'] += params.get('ticket_medio', 2400.0) * weight
            
            # Create averaged parameters
            avg_params = {
                'fator_elasticidade': total_params['fator_elasticidade'],
                'taxa_agendamento': total_params['taxa_agendamento'],
                'taxa_comparecimento': total_params['taxa_comparecimento'],
                'taxa_conversao': total_params['taxa_conversao'],
                'ticket_medio': total_params['ticket_medio'],
                'repasse_bruto': p_receitas.get('repasse_bruto', 85.0)  # Always get from p_receitas
            }
            
            return avg_params
        
        # If no parameters found, return defaults
        default_params['repasse_bruto'] = p_receitas.get('repasse_bruto', default_params['repasse_bruto'])
        return default_params
    
    def _render_marketing_model(self):
        st.write("### Canais de Vendas:")
        
        # Information message about using team data
        st.info("Os canais de venda são automaticamente gerados a partir dos membros de equipe marcados como 'Sujeito a Aumento de Receita' em 'Premissas Despesas'.")
        
        # Get team members configured in PremissasDespesas
        team_members = []
        if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
            team_members = [
                cargo for cargo in st.session_state['premissas_despesas']['equipe_propria'] 
                if cargo.get('sujeito_aumento_receita', False)
            ]
        
        if not team_members:
            st.warning("Nenhum membro de equipe marcado como 'Sujeito a Aumento de Receita'. Configure os membros de equipe em 'Premissas Despesas'.")
        else:
            # Generate sales channels based on team members
            st.session_state['premissas_receitas']['canais_venda'] = []
            
            for member in team_members:
                # Create a channel based on team member data
                canal = {
                    'descricao': f"Canal - {member['nome']}",
                    'gasto_mensal': member.get('gasto_mensal', 5000.0),
                    'cpl_base': member.get('cpl_base', 10.0),
                    'crescimento_vendas': member.get('crescimento_vendas', 'Linear'),
                    'periodicidade': member.get('periodicidade', 'Mensal'),
                    'tx_cresc_mensal': member.get('tx_cresc_mensal', 5.0),
                    'media_cresc_anual': member.get('media_cresc_anual', 15.0),
                    'fator_aceleracao_crescimento': member.get('fator_aceleracao_crescimento', 1.0),
                    'rpe_anual': member.get('rpe_anual', 125000.0),
                    'salario_medio': member.get('salario_medio', 60000.0),
                    'depreciacao': member.get('depreciacao', 1.5),
                    'conversion_params': {
                        'fator_elasticidade': member.get('fator_elasticidade', 1.0),
                        'taxa_agendamento': member.get('taxa_agendamento', 30.0),
                        'taxa_comparecimento': member.get('taxa_comparecimento', 70.0),
                        'taxa_conversao': member.get('taxa_conversao', 45.0),
                        'ticket_medio': member.get('ticket_medio', 2400.0)
                    }
                }
                
                st.session_state['premissas_receitas']['canais_venda'].append(canal)
            
            # Display generated sales channels
            st.write("#### Canais de Venda Configurados")
            
            # Calculate totals for display
            total_gasto = sum(canal['gasto_mensal'] for canal in st.session_state['premissas_receitas']['canais_venda'])
            total_cpl = sum(canal['cpl_base'] for canal in st.session_state['premissas_receitas']['canais_venda']) / len(st.session_state['premissas_receitas']['canais_venda']) if st.session_state['premissas_receitas']['canais_venda'] else 0
            
            # Display total metrics
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total Gasto nos Canais de Vendas", f"R$ {total_gasto:,.2f}")
            with col2:
                st.metric("Total Custo por Lead (média)", f"R$ {total_cpl:,.2f}")
            
            # Create a DataFrame for display
            canais_data = []
            for canal in st.session_state['premissas_receitas']['canais_venda']:
                canal_info = {
                    "Canal": canal['descricao'],
                    "Gasto Mensal (R$)": f"{canal['gasto_mensal']:,.2f}",
                    "Crescimento": canal['crescimento_vendas'],
                    "Custo por Lead (R$)": f"{canal.get('cpl_base', 0):,.2f}"
                }
                
                if canal['crescimento_vendas'] == "Linear":
                    canal_info["Taxa Crescimento"] = f"{canal['tx_cresc_mensal']}% ao mês"
                elif canal['crescimento_vendas'] in ["Não Linear S/ Downside", "Não Linear C/ Downside"]:
                    canal_info["Taxa Crescimento"] = f"{canal['media_cresc_anual']}% ao ano (média)"
                else:
                    canal_info["Taxa Crescimento"] = "N/A"
                    
                canais_data.append(canal_info)
            
            df_canais = pd.DataFrame(canais_data)
            st.dataframe(df_canais, use_container_width=True)
        
        # Outras Receitas Não Recorrentes
        self._render_outras_receitas()
    
    def _render_outras_receitas(self):
        # Expander for other non-recurring revenues
        with st.expander("Outras Receitas Não Recorrentes", expanded=False):
            # Inside _render_outras_receitas method within PremissasReceitas class
            with st.form("adicionar_outra_receita"):
                st.write("#### Adicionar Nova Receita Adicional")
                
                descricao = st.text_input("Descrição", key="other_rev_desc", help="Descrição da receita adicional")
                valor = st.number_input("Valor (R$)", min_value=0.0, step=1000.0, format="%.2f", help="Valor da receita")
                mes_inflow = st.number_input("Mês de Entrada", min_value=0, max_value=59, step=1, help="Mês em que a receita entra no fluxo de caixa")
                
                col1, col2 = st.columns(2)
                with col1:
                    periodicidade_ativa = st.checkbox("Receita Recorrente", key="other_rev_periodic", 
                                                    help="Marque se esta receita se repete periodicamente")
                
                with col2:
                    if periodicidade_ativa:
                        periodicidade = st.number_input("Periodicidade (meses)", min_value=1, max_value=60, value=1, step=1, 
                                                      help="De quantos em quantos meses esta receita se repete")
                    else:
                        periodicidade = 0
                
                # Add growth type selection
                forma_crescimento = st.selectbox(
                    "Forma de Crescimento",
                    ["Linear", "Composto", "Exponencial", "Sigmoidal"],
                    index=0,
                    help="Linear: cresce em valor absoluto fixo. Composto: cresce em percentual sobre o valor anterior."
                )
                
                # Modified growth percentage field to be annual mean
                tx_crescimento_anual = st.number_input(
                    "Taxa Média de Crescimento Anual (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=0.0,
                    step=1.0,
                    format="%.2f",
                    help="Taxa média anual de crescimento. Os valores mensais serão aleatórios seguindo uma distribuição, mas com média anual próxima a este valor."
                )
                
                # Add reajuste (inflation correction) option
                reajuste_ativo = st.checkbox("Aplicar Reajuste", key="other_rev_reajuste", 
                                            help="Marque para aplicar reajuste periódico ao valor da receita")
                
                if reajuste_ativo:
                    col1, col2 = st.columns(2)
                    with col1:
                        periodicidade_reajuste = st.selectbox(
                            "Periodicidade do Reajuste",
                            ["Mensal", "Trimestral", "Semestral", "Anual", "Bienal"],
                            index=3,
                            help="Periodicidade em que o valor será reajustado"
                        )
                    
                    with col2:
                        indice_reajuste = st.number_input(
                            "Índice de Reajuste (%)",
                            min_value=0.0,
                            max_value=100.0,
                            value=4.5,
                            step=0.1,
                            format="%.1f",
                            help="Índice percentual de reajuste periódico"
                        )
                else:
                    periodicidade_reajuste = "Anual"
                    indice_reajuste = 0.0
                
                prazo = st.number_input("Prazo (meses)", min_value=1, max_value=60, value=1, step=1, 
                                      help="Por quantos meses essa receita se estende")
                
                # Add field for variation intensity
                variacao_intensidade = st.slider(
                    "Intensidade da Variação",
                    min_value=0.0,
                    max_value=5.0,
                    value=1.0,
                    step=0.1,
                    format="%.1f",
                    help="Controla a intensidade da variação aleatória aplicada ao crescimento (0 = sem variação, 5 = alta variação)"
                )
                
                submitted = st.form_submit_button("Adicionar Receita")
                
                if submitted and descricao and valor > 0:
                    # Add the revenue to the list with the new fields
                    receita_data = {
                        'descricao': descricao,
                        'valor': valor,
                        'mes_inflow': mes_inflow,
                        'periodicidade_ativa': periodicidade_ativa,
                        'periodicidade': periodicidade,
                        'forma_crescimento': forma_crescimento,
                        'tx_crescimento_anual': tx_crescimento_anual,  # Changed from tx_crescimento to tx_crescimento_anual
                        'variacao_intensidade': variacao_intensidade,
                        'reajuste_ativo': reajuste_ativo,
                        'periodicidade_reajuste': periodicidade_reajuste,
                        'indice_reajuste': indice_reajuste,
                        'prazo': prazo
                    }
                    
                    st.session_state['premissas_receitas']['outras_receitas'].append(receita_data)
                    st.success(f"Receita '{descricao}' adicionada com sucesso!")
                    st.rerun()
            
            # Display existing other revenues
            if st.session_state['premissas_receitas']['outras_receitas']:
                st.write("#### Receitas Adicionais Configuradas")
                
                # Create a DataFrame for display
                receitas_data = []
                
                for receita in st.session_state['premissas_receitas']['outras_receitas']:
                    receita_info = {
                        "Descrição": receita['descricao'],
                        "Valor (R$)": f"{receita['valor']:,.2f}",
                        "Mês de Entrada": receita['mes_inflow']
                    }
                    
                    if receita['periodicidade_ativa']:
                        receita_info["Recorrência"] = f"A cada {receita['periodicidade']} meses"
                    else:
                        receita_info["Recorrência"] = "Única"
                    
                    # Add growth information - Updated to show annual growth rate
                    if receita.get('tx_crescimento_anual', 0) > 0:
                        receita_info["Crescimento"] = f"{receita.get('forma_crescimento', 'Linear')} ({receita.get('tx_crescimento_anual', 0):.1f}% anual)"
                    else:
                        receita_info["Crescimento"] = "Sem crescimento"
                    
                    if receita.get('reajuste_ativo', False):
                        receita_info["Reajuste"] = f"{receita.get('indice_reajuste', 0):.1f}% ({receita.get('periodicidade_reajuste', 'Anual')})"
                    else:
                        receita_info["Reajuste"] = "Não"
                        
                    receita_info["Prazo"] = f"{receita['prazo']} meses"
                    
                    receitas_data.append(receita_info)
                
                df_receitas = pd.DataFrame(receitas_data)
                st.dataframe(df_receitas, use_container_width=True)
                
                # NEW: Add option to select and remove individual revenue
                descricoes = [receita['descricao'] for receita in st.session_state['premissas_receitas']['outras_receitas']]
                if descricoes:
                    receita_para_remover = st.selectbox("Selecione uma receita para remover", descricoes)
                    
                    if st.button(f"Remover '{receita_para_remover}'"):
                        idx_to_remove = next((i for i, receita in enumerate(st.session_state['premissas_receitas']['outras_receitas']) 
                                              if receita['descricao'] == receita_para_remover), None)
                        
                        if idx_to_remove is not None:
                            st.session_state['premissas_receitas']['outras_receitas'].pop(idx_to_remove)
                            st.success(f"Receita '{receita_para_remover}' removida com sucesso!")
                            st.rerun()
                
                # Add button to remove all revenues
                if st.button("Remover Todas as Receitas Adicionais"):
                    st.session_state['premissas_receitas']['outras_receitas'] = []
                    st.success("Todas as receitas adicionais foram removidas!")
                    st.rerun()
    
    
    def _show_summary(self):
        # Create a DataFrame with all parameters for display
        params = []
        
        # Add model type
        params.append(["Tipo de Modelo", "Marketing" if st.session_state['premissas_receitas']['modelo_marketing'] else "Financeiro"])
        
        # Add relevant parameters based on model type
        if st.session_state['premissas_receitas']['modelo_marketing']:
            # Marketing model parameters
            params.append(["Número de Canais de Venda", len(st.session_state['premissas_receitas']['canais_venda'])])
            
            if st.session_state['premissas_receitas']['canais_venda']:
                total_gasto = sum(canal['gasto_mensal'] for canal in st.session_state['premissas_receitas']['canais_venda'])
                params.append(["Total Gasto nos Canais (R$)", f"{total_gasto:,.2f}"])
            
            params.append(["Fator de Elasticidade (β)", st.session_state['premissas_receitas'].get('fator_elasticidade', 1.0)])
            params.append(["Taxa de Agendamento (%)", st.session_state['premissas_receitas'].get('taxa_agendamento', 30.0)])
            params.append(["Taxa de Comparecimento (%)", st.session_state['premissas_receitas'].get('taxa_comparecimento', 70.0)])
            params.append(["Taxa de Conversão (%)", st.session_state['premissas_receitas'].get('taxa_conversao', 45.0)])
            params.append(["Ticket-Médio (R$)", f"{st.session_state['premissas_receitas'].get('ticket_medio', 2400.0):,.2f}"])
        else:
            # Financial model parameters
            params.append(["Receita Inicial (R$)", f"{st.session_state['premissas_receitas']['receita_inicial']:,.2f}"])
            params.append(["Valor Unitário (R$)", f"{st.session_state['premissas_receitas']['valor_unitario']:,.2f}"])
            params.append(["Tipo de Crescimento", st.session_state['premissas_receitas']['crescimento_receita']])
            
            if st.session_state['premissas_receitas']['crescimento_receita'] == "Linear":
                params.append(["Taxa de Crescimento Mensal (%)", st.session_state['premissas_receitas']['tx_cresc_mensal']])
            elif st.session_state['premissas_receitas']['crescimento_receita'] in ["Não Linear S/ Downside", "Não Linear C/ Downside"]:
                params.append(["Média de Crescimento Anual (%)", st.session_state['premissas_receitas']['media_cresc_anual']])
            
            params.append(["Número de Fontes de Receita", len(st.session_state['premissas_receitas']['fontes_primarias'])])
        
        # Common parameters
        params.append(["Repasse Bruto (%)", st.session_state['premissas_receitas']['repasse_bruto']])
        params.append(["Número de Receitas Adicionais", len(st.session_state['premissas_receitas']['outras_receitas'])])
        
         # Display the summary
        df_summary = pd.DataFrame(params, columns=pd.Index(["Parâmetro", "Valor"]))
        st.dataframe(df_summary, use_container_width=True)

class Receitas(Page):
    def __init__(self):
        self.graph_types = ["Gráfico de Linhas", "Gráfico de Pizza", "Gráfico de Barras"]
        self.time_frames = ["Mensal", "Trimestral", "Anual"]
    
    @property
    def title(self) -> str:
        return "Receitas"
    
    @property
    def icon(self) -> str:
        return "📈"
    
    def _guess_params(self, payroll, i, rpe_annual=125000.0, avg_salary=60000.0, depreciation=0.0):
        """
        Quick-and-dirty parameter guesses (A, B, C, δ).
        """
        payroll = np.asarray(payroll, dtype=float)
        x0 = max(payroll[0], 1.0)  # Ensure x0 is at least 1.0 to avoid division by zero
        
        # ---- productivity factors ~ 1 in the first month ------------------------
        payroll_mean = payroll.mean() if payroll.size > 0 else 1.0
        B = 1.0 / max(payroll_mean, 1.0)  # Ensure we never divide by zero
        C = 0.0 if i <= 0 else 1.0 / i    # cap-ex term - avoid division by zero
        δ = max(0, depreciation / 100)     # Convert from percentage to decimal, ensure non-negative
        
        # ---- derive A from the month-0 revenue target ---------------------------
        emp0 = x0 * 12 / max(avg_salary, 1.0)  # employees implied by payroll, avoid division by zero
        r0_target = emp0 * rpe_annual / 12     # $ revenue in month 0
        
        numerator = r0_target * (1 + B * x0 + C * i)
        denominator = max(B * x0 - δ * C * i, 1e-9)  # Ensure denominator is never too close to zero
        A = numerator / denominator
        return A, B, C, δ
    
    def _monthly_revenue(self, payroll, i, A, B, C, δ=0.0, months=60):
        """
        Original formula but returns the inverse of the result to produce an ascending curve
        """
        payroll = np.asarray(payroll, dtype=float)
        if months >= len(payroll):
            # Extend payroll array if needed
            extended_payroll = np.zeros(months)
            extended_payroll[:len(payroll)] = payroll
            
            # Extend using the last value
            if len(payroll) > 0:
                extended_payroll[len(payroll):] = payroll[-1]
            
            payroll = extended_payroll
        
        t = np.arange(months, dtype=float)
        decay = np.exp(-δ * t)
        cum_pay = np.cumsum(payroll)
    
        # Calculate using original formula
        num = A * (B * payroll - δ * C * i * decay)
        den = 1.0 + B * cum_pay + C * i * decay
        base_result = num / den
        
        # Invert the trend by using the starting value plus the difference from the starting value
        start_value = base_result[0] if len(base_result) > 0 else 0
        inversion_factor = 2 * start_value
        return inversion_factor - base_result  # This will invert the curve
    
    def _get_payroll_data(self, months=60):
        """
        Retrieve the planned payroll data for the specified number of months.
        """
        payroll_data = np.zeros(months)
        
        try:
            if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
                equipe = Equipe()
                df_equipe = equipe._gerar_dataframe_custos_equipe()
                
                if df_equipe is not None and not df_equipe.empty:
                    # Look for team costs in each month
                    for mes in range(min(months, df_equipe.shape[1])):
                        try:
                            # Check for "Total Custos de Equipe" at mes+1 (since df_equipe columns start at 1)
                            if mes+1 in df_equipe.columns and ("TOTAL", "Total Custos de Equipe") in df_equipe.index:
                                payroll_data[mes] = df_equipe.loc[("TOTAL", "Total Custos de Equipe"), mes+1]
                        except Exception:
                            pass
        except Exception as e:
            st.warning(f"Erro ao obter dados de folha de pagamento: {e}")
        
        # If first month is zero, use a reasonable default
        if payroll_data[0] == 0:
            payroll_data[0] = 10000.0  # Default value to avoid division by zero
        
        return payroll_data
    
    def _get_initial_investment(self):
        """
        Retrieve the initial investment amount from PremissasInvestimentos.
        """
        initial_investment = 0.0
        
        try:
            if 'premissas_investimentos' in st.session_state:
                investimentos = Investimentos()
                initial_investment = investimentos.get_total_investimento()
        except Exception as e:
            st.warning(f"Erro ao obter dados de investimento inicial: {e}")
        
        return initial_investment
    
    def _add_growth_variation(self, base_value, std_dev_percent=0.001, allow_downside=False):
        """
        Add random variation to a growth value based on a distribution that can include downside.
        
        Args:
            base_value (float): The base value (mean of the distribution)
            std_dev_percent (float): Standard deviation as percentage of mean
            allow_downside (bool): Whether to allow values below the mean
            
        Returns:
            float: Value with random variation
        """
        # Handle edge case
        if base_value == 0:
            return 0
            
        # Calculate standard deviation
        std_dev = abs(base_value) * (std_dev_percent / 100.0)
        
        if allow_downside:
            # Use normal distribution for symmetric variation (allowing downside)
            # But with slightly positive skew by adding a smaller chi-square component
            chi2_component = np.random.chisquare(2) - 2  # Centered chi-square
            
            # Scale by std_dev and add to base_value
            return base_value + chi2_component * std_dev
        else:
            # Original chi-square method for positive skew only
            df = 2  # degrees of freedom
            chi2_value = np.random.chisquare(df)
            centered = (chi2_value - df) / np.sqrt(2 * df)
            return base_value + centered * std_dev
    
    def _calcular_crescimento(self, valor_base, mes, tipo_crescimento, tx_cresc_mensal=None, 
                             media_cresc_anual=None, com_downside=True, periodicidade="Mensal", 
                             fator_aceleracao=1.0, rpe_anual=15000.0, salario_medio=60000.0, depreciacao=1.5):
        """Calcula o valor com crescimento conforme o tipo selecionado e periodicidade"""
        # Check for Produtividade model
        if tipo_crescimento == "Produtividade":
            # Get payroll data
            payroll_data = self._get_payroll_data(months=61)
            
            # Much more aggressive dampening for early growth to reduce explosiveness
            for i in range(min(18, len(payroll_data))):
                # Start at 15% for month 0 and very gradually increase to 80% by month 6
                # This creates a much smoother growth curve over the first 2 years
                dampen_factor = 0.15 + 0.6 * (i / 18)
                payroll_data[i] *= dampen_factor
            
            # Slow down later decay
            i = self._get_initial_investment()
            
            # Further reduce depreciation rate to slow down decay in later years
            adjusted_depreciation = depreciacao * 0.8 

            # Calculate parameters with adjusted depreciation rate
            A, B, C, δ = self._guess_params(
                payroll=payroll_data,
                i=i,
                rpe_annual=rpe_anual,
                avg_salary=salario_medio,
                depreciation=adjusted_depreciation
            )
            
            # Calculate revenues for all months - this is now our MEAN curve
            revenues = self._monthly_revenue(payroll_data, i, A, B, C, δ, months=61)
            
            # Return the value for the requested month with significant random variation
            if mes < len(revenues):
                # Increased standard deviation for more variation
                return self._add_growth_variation(revenues[mes], 8.0, allow_downside=com_downside)
            
            return valor_base
    
    
    def _calcular_cpl_ajustado(self, gasto, cpl_base, elasticidade=1.0):
        """Calcula o CPL ajustado pelo fator de elasticidade
        
        CPL(Spend) = CPL_base * (Spend/Base_Spend)^(β-1)
        Onde β é o fator de elasticidade
        
        β = 1: CPL constante (retorno linear)
        β > 1: CPL aumenta com gasto (retorno decrescente)
        β < 1: CPL diminui com gasto (retorno crescente)
        """
        # Definir um gasto base de referência
        gasto_base = 10000  # R$ 10.000 como referência
        
        # Se o gasto for zero, retornar o CPL base
        if gasto == 0:
            return cpl_base
        
        # Calcular CPL ajustado
        return cpl_base * ((gasto / gasto_base) ** (elasticidade - 1))
    
    # Add this method to the Receitas class
    def _calcular_equipe_comercial(self, df_faturamento):
        """Calcula a equipe comercial necessária (SDR e Closer) com base nos leads e atendimentos"""
        if df_faturamento is None or "Total de Leads" not in df_faturamento.index or "Total de Comparecimento" not in df_faturamento.index:
            return pd.DataFrame()
        
        meses = df_faturamento.columns
        
        # Criar DataFrame para equipe comercial com salários
        df_equipe = pd.DataFrame(index=pd.Index(["SDR Necessários", "Closer Necessários", "Salário SDR", "Salário Closer"]), columns=meses)
        
        # Obter capacidades da equipe a partir de PremissasDespesas
        capacidade_sdr_leads = 750  # Valor padrão
        capacidade_closer_atendimentos = 90  # Valor padrão
        
        # Obter quantidade inicial de SDR e Closer de PremissasDespesas
        sdr_inicial = 0
        closer_inicial = 0
        sdr_salario = 0
        closer_salario = 0
        
        # Buscar dados de quantidade e salário da equipe das premissas
        if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
            for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                if 'nome' in cargo:
                    if cargo['nome'] == 'SDR':
                        sdr_inicial = cargo.get('quantidade', 0)
                        sdr_salario = cargo.get('salario', 0)
                        capacidade_sdr_leads = cargo.get('capacidade_leads', 750)
                    elif cargo['nome'] == 'Closer':
                        closer_inicial = cargo.get('quantidade', 0)
                        closer_salario = cargo.get('salario', 0)
                        capacidade_closer_atendimentos = cargo.get('capacidade_atendimentos', 90)
        
        for mes in meses:
            # Calcular SDRs necessários com base no volume de leads
            total_leads = pd.to_numeric(df_faturamento.loc["Total de Leads", mes], errors='coerce') or 0
            sdr_necessarios = max(math.ceil(total_leads / capacidade_sdr_leads) if capacidade_sdr_leads > 0 else 0, sdr_inicial)
            df_equipe.loc["SDR Necessários", mes] = sdr_necessarios
            df_equipe.loc["Salário SDR", mes] = sdr_necessarios * sdr_salario
            
            # Calcular Closers necessários com base no volume de comparecimentos
            total_comparecimento = pd.to_numeric(df_faturamento.loc["Total de Comparecimento", mes], errors='coerce') or 0
            closer_necessarios = max(math.ceil(total_comparecimento / capacidade_closer_atendimentos) if capacidade_closer_atendimentos > 0 else 0, closer_inicial)
            df_equipe.loc["Closer Necessários", mes] = closer_necessarios
            df_equipe.loc["Salário Closer", mes] = closer_necessarios * closer_salario
        
        return df_equipe
    
    # In the Receitas class, update the _gerar_dataframe_faturamento method
    def _gerar_dataframe_faturamento(self):
        """Gera o dataframe de faturamento para 60 meses (5 anos)"""
        if 'premissas_receitas' not in st.session_state:
            return None
        
        # Check if we have team data in PremissasDespesas
        if ('premissas_despesas' not in st.session_state or 
            'equipe_propria' not in st.session_state['premissas_despesas'] or
            not st.session_state['premissas_despesas']['equipe_propria']):
            st.warning("Nenhum membro de equipe configurado. Configure a equipe em 'Premissas Despesas'.")
            return None
        
        # Criar dataframe para 60 meses
        meses = range(60)
        
        # Always use marketing model with team data
        return self._gerar_df_faturamento_marketing(meses)
    
    def _gerar_df_faturamento_marketing(self, meses):
        """Gera o dataframe de faturamento para modelo de marketing"""
        p_receitas = st.session_state['premissas_receitas']
        
        # Find SDR and Closer in the team
        sdr_data = None
        closer_data = None
        
        if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
            for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                if cargo['nome'] == 'SDR':
                    sdr_data = cargo
                elif cargo['nome'] == 'Closer':
                    closer_data = cargo
        
        # If we don't have both SDR and Closer data, we can't proceed
        if not sdr_data or not closer_data:
            st.warning("É necessário configurar os cargos SDR e Closer na equipe.")
            return None
        
        # Create indices for dataframe
        indices = [
            "Leads por Canal",
            "Total de Leads", 
            "Total de Agendamentos", 
            "Total de Comparecimento", 
            "Conversões", 
            "Faturamento Bruto",
            "Repasse Bruto",
            "ROAS",
            "Inside Sales",  # Only one channel now
            "Total"
        ]
        
        # Create DataFrame
        df = pd.DataFrame(index=pd.Index(indices), columns=meses)
        df = df.fillna(0.0).astype(float)
        
        # Calculate initial values based on team inputs
        inicial_leads = sdr_data.get('estimativa_leads', 200)
        taxa_agendamento = sdr_data.get('taxa_agendamento', 30.0) / 100
        taxa_comparecimento = sdr_data.get('taxa_comparecimento', 70.0) / 100
        taxa_conversao = closer_data.get('taxa_conversao', 45.0) / 100
        valor_unitario = closer_data.get('valor_unitario', 2400.0)
        quantidade_closer = closer_data.get('quantidade', 1)
        produtos_por_lead = closer_data.get('produtos_por_lead', 10)
        taxa_cancelamento = closer_data.get('taxa_cancelamento', 5.0) / 100
        
        # Calculate initial revenue
        inicial_agendamentos = inicial_leads * taxa_agendamento
        inicial_comparecimentos = inicial_agendamentos * taxa_comparecimento
        inicial_conversoes = inicial_comparecimentos * taxa_conversao
        
        # Calculate values for both revenue models
        valor_inflow_mensal = valor_unitario * produtos_por_lead  # Monthly subscription model
        valor_inflow_antecipado = (valor_unitario * 12 * produtos_por_lead) - ((valor_unitario * 12 * produtos_por_lead) * 0.2)  # Upfront model with 20% discount
        
        inicial_faturamento_mensal = (inicial_conversoes / 2) * valor_inflow_mensal * quantidade_closer
        inicial_faturamento_antecipado = (inicial_conversoes / 2) * valor_inflow_antecipado * quantidade_closer
        inicial_faturamento = inicial_faturamento_mensal + inicial_faturamento_antecipado
        
        # Calculate RPE (Revenue Per Employee) for productivity model
        rpe_annual = (inicial_faturamento * 12) / quantidade_closer
        
        # Get payroll data for productivity model
        payroll_data = self._get_payroll_data(months=60)
        initial_investment = self._get_initial_investment()
        
        # Create arrays to track new conversions by month for each model
        conversions_by_month_mensal = np.zeros(60)
        conversions_by_month_antecipado = np.zeros(60)
        
        # Calculate base revenue growth using productivity model
        for mes in meses:
            # For month 0, use inicial values
            if mes == 0:
                total_leads = inicial_leads
                total_agendamentos = inicial_agendamentos
                total_comparecimento = inicial_comparecimentos
                total_conversoes = inicial_conversoes
                
                # Split conversions between models (50/50)
                conversions_by_month_mensal[mes] = total_conversoes / 2
                conversions_by_month_antecipado[mes] = total_conversoes / 2
            else:
                # Use _calcular_crescimento to get the growth factor for this month
                scaling_factor = self._calcular_crescimento(
                    inicial_faturamento,
                    mes,
                    tipo_crescimento="Produtividade",
                    com_downside=False,
                    rpe_anual=rpe_annual,
                    salario_medio=closer_data.get('salario', 2400.0) * 12,
                    depreciacao=1.5  # Default depreciation
                ) / inicial_faturamento if inicial_faturamento > 0 else 1
                
                # Calculate conversion metrics independently for each month based on the scaled initial values
                total_leads = inicial_leads * scaling_factor
                total_agendamentos = total_leads * taxa_agendamento
                total_comparecimento = total_agendamentos * taxa_comparecimento
                total_conversoes = total_comparecimento * taxa_conversao
                
                # Split conversions between models (50/50)
                conversions_by_month_mensal[mes] = total_conversoes / 2
                conversions_by_month_antecipado[mes] = total_conversoes / 2
            
            # Store conversion metrics in DataFrame
            df.loc["Leads por Canal", mes] = total_leads
            df.loc["Total de Leads", mes] = total_leads
            df.loc["Total de Agendamentos", mes] = total_agendamentos
            df.loc["Total de Comparecimento", mes] = total_comparecimento
            df.loc["Conversões", mes] = total_conversoes
            
            # Calculate revenue considering both models
            total_faturamento = 0.0
            
            # MODEL 1: Monthly subscription with cancellation after 3 months
            for past_month in range(mes + 1):
                months_active = mes - past_month
                past_conversions_mensal = conversions_by_month_mensal[past_month]
                
                if months_active == 0:
                    # Revenue from conversions this month
                    revenue_from_past_mensal = past_conversions_mensal * valor_inflow_mensal * quantidade_closer
                elif months_active <= 2:
                    # Revenue from first 3 months (no cancellation)
                    revenue_from_past_mensal = past_conversions_mensal * valor_inflow_mensal * quantidade_closer
                else:
                    # Revenue after 3 months (apply cancellation rate)
                    remaining_factor = (1 - taxa_cancelamento) ** (months_active - 2)
                    revenue_from_past_mensal = past_conversions_mensal * valor_inflow_mensal * quantidade_closer * remaining_factor
                
                total_faturamento += revenue_from_past_mensal
            
            # MODEL 2: Upfront payment (12 months anticipated with discount)
            for past_month in range(mes + 1):
                if mes == past_month:  # Only add upfront revenue in the conversion month
                    past_conversions_antecipado = conversions_by_month_antecipado[past_month]
                    revenue_from_past_antecipado = past_conversions_antecipado * valor_inflow_antecipado * quantidade_closer
                    total_faturamento += revenue_from_past_antecipado
            
            df.loc["Faturamento Bruto", mes] = total_faturamento
            
            # Apply repasse bruto
            repasse_bruto = p_receitas.get('repasse_bruto', 85.0) / 100
            df.loc["Repasse Bruto", mes] = total_faturamento * repasse_bruto
            
            # Calculate ROAS (Return On Ad Spend) - use a default marketing spend if not available
            marketing_spend = 5000.0  # Default value
            df.loc["ROAS", mes] = total_faturamento / marketing_spend if marketing_spend > 0 else 0
            
            # Store channel and total revenue
            df.loc["Inside Sales", mes] = total_faturamento
            df.loc["Total", mes] = total_faturamento
        
        # Add other non-recurring revenues
        df = self._adicionar_outras_receitas(df, meses)
        
        return df
    
    def _gerar_dataframe_arrecadacao(self, df_faturamento):
        """Gera o dataframe de arrecadação para 60 meses (5 anos)"""
        if df_faturamento is None or 'premissas_receitas' not in st.session_state:
            return None
        
        p_receitas = st.session_state['premissas_receitas']
        repasse_bruto = p_receitas.get('repasse_bruto', 85.0) / 100
        
        # Criar DataFrame para arrecadação
        meses = range(60)
        df = pd.DataFrame(index=pd.Index(["Faturamento Bruto", "Arrecadação Líquida"]), columns=meses)
        
        # Copiar valores de faturamento bruto
        df.loc["Faturamento Bruto"] = df_faturamento.loc["Total"]
        
        # Calcular arrecadação líquida
        for mes in meses:
            df.loc["Arrecadação Líquida", mes] = df.loc["Faturamento Bruto", mes] * repasse_bruto
        
        return df
    
    def _agrupar_por_periodo(self, df, periodo):
        """Agrupa o dataframe por período (trimestral ou anual)"""
        if df is None or df.empty:
            return None
            
        if periodo == "Mensal":
            return df
        
        # Definir o tamanho do grupo
        if periodo == "Trimestral":
            tamanho_grupo = 3
        else:  # Anual
            tamanho_grupo = 12
        
        # Criar novo DataFrame
        periodos = [f"{periodo} {i//tamanho_grupo + 1}" for i in range(0, 60, tamanho_grupo)]
        df_agrupado = pd.DataFrame(index=df.index, columns=periodos)
        
        # Agrupar os dados
        for i, nome_periodo in enumerate(periodos):
            inicio = i * tamanho_grupo
            fim = min((i + 1) * tamanho_grupo, 60)
            
            # Calcular a soma para o período
            for idx in df.index:
                df_agrupado.loc[idx, nome_periodo] = df.loc[idx, inicio:fim-1].sum()
        
        return df_agrupado
    
    # Inside the _adicionar_outras_receitas method of the Receitas class
    def _adicionar_outras_receitas(self, df, meses):
        """Adiciona outras receitas não recorrentes ao DataFrame"""
        p_receitas = st.session_state['premissas_receitas']
        outras_receitas = p_receitas.get('outras_receitas', [])
        
        if not outras_receitas:
            return df
        
        # Criar um DataFrame temporário para as outras receitas
        df_outras = pd.DataFrame(0, index=pd.Index(["Outras Receitas"]), columns=meses)
        
        # Processar cada receita adicional
        for receita in outras_receitas:
            valor_base = receita['valor']
            mes_inicial = receita['mes_inflow']
            prazo = receita['prazo']
            periodicidade_ativa = receita['periodicidade_ativa']
            periodicidade = receita['periodicidade'] if periodicidade_ativa else 0
            
            # Get growth parameters if available
            forma_crescimento = receita.get('forma_crescimento', 'Linear')
            tx_crescimento_anual = receita.get('tx_crescimento_anual', 0.0)  # Get annual growth rate
            variacao_intensidade = receita.get('variacao_intensidade', 1.0)
            
            # Get reajuste parameters if available
            reajuste_ativo = receita.get('reajuste_ativo', False)
            periodicidade_reajuste = receita.get('periodicidade_reajuste', 'Anual')
            indice_reajuste = receita.get('indice_reajuste', 0.0) / 100
            
            # Map reajuste periodicity to number of months
            periodicidade_map = {
                'Mensal': 1,
                'Trimestral': 3,
                'Semestral': 6,
                'Anual': 12,
                'Bienal': 24
            }
            meses_reajuste = periodicidade_map.get(periodicidade_reajuste, 12)
            
            # Generate random monthly growth rates based on chi-squared distribution
            # that will approximate the annual growth rate on average
            if tx_crescimento_anual > 0:
                # Convert annual rate to equivalent monthly rate
                if forma_crescimento == "Linear":
                    # For linear growth, divide annual rate by 12
                    tx_mensal_media = tx_crescimento_anual / 12
                elif forma_crescimento == "Composto":
                    # For compound growth, use this formula: (1+annual_rate)^(1/12) - 1
                    tx_mensal_media = ((1 + tx_crescimento_anual/100) ** (1/12) - 1) * 100
                elif forma_crescimento == "Exponencial":
                    # For exponential growth, use this formula: e^(annual_rate/100)^(1/12) - 1
                    tx_mensal_media = (np.exp(tx_crescimento_anual/100) ** (1/12) - 1) * 100
                else: 
                    # For sigmoidal growth, use a logistic function approximation
                    tx_mensal_media = (1 / (1 + np.exp(-0.1 * (tx_crescimento_anual - 50)))) * 100

                # Pre-generate monthly growth rates for the entire period
                np.random.seed(42)  # Use fixed seed for reproducibility
                monthly_rates = []
                
                # Scale factor for the chi-square distribution to create appropriate variability
                # Higher variacao_intensidade = more variability
                scale_factor = 0.5 * variacao_intensidade
                
                # Generate monthly rates using chi-squared distribution
                for i in range(prazo):
                    # Use chi-square distribution with 3 degrees of freedom for positive skewness and high kurtosis
                    chi2_val = np.random.chisquare(2)
                    # Adjust the chi2 value to be centered around the mean monthly rate
                    adjustment = chi2_val - 2  # chi-square with df=2 has mean=2
                    # Scale the adjustment according to intensity and mean
                    monthly_rate = tx_mensal_media * (1 + scale_factor * adjustment / 2)
                    # Ensure no negative growth rates for the chi-squared model
                    # monthly_rate = max(0, monthly_rate)
                    monthly_rates.append(monthly_rate / 100)  # Convert to decimal
                
                # Normalize rates to ensure they average correctly over a year
                if len(monthly_rates) >= 12:
                    for year in range(len(monthly_rates) // 12):
                        start_idx = year * 12
                        end_idx = start_idx + 12
                        year_rates = monthly_rates[start_idx:end_idx]
                        
                        # Calculate current compound annual rate
                        if forma_crescimento == "Linear":
                            current_annual = sum(year_rates) * 100  # Convert back to percentage
                        else:  # "Composto"
                            current_annual = ((np.prod([1 + r for r in year_rates])) - 1) * 100
                        
                        # Calculate adjustment factor
                        if current_annual > 0:  # Avoid division by zero
                            adjustment_factor = tx_crescimento_anual / current_annual
                            
                            # Apply adjustment to monthly rates
                            for j in range(start_idx, end_idx):
                                monthly_rates[j] *= adjustment_factor
            else:
                # If no growth, all rates are zero
                monthly_rates = [0] * prazo
                
            if periodicidade_ativa and periodicidade > 0:
                # Receita recorrente periódica
                for mes_ocorrencia in range(mes_inicial, 60, periodicidade):
                    if mes_ocorrencia >= 60:
                        break
                    
                    # Calculate how many periods have passed for growth calculation
                    periodos_passados = (mes_ocorrencia - mes_inicial) // periodicidade
                    
                    # Apply growth based on selected method
                    valor_atual = valor_base
                    if periodos_passados > 0 and tx_crescimento_anual > 0:
                        # Use our pre-calculated monthly rates
                        if forma_crescimento == "Linear":
                            # For linear growth, apply cumulative sum of rates
                            cumulative_rate = sum(monthly_rates[:periodos_passados])
                            valor_atual = valor_base * (1 + cumulative_rate)
                        else:  # "Composto"
                            # For compound growth, apply rates compounded
                            for i in range(periodos_passados):
                                valor_atual *= (1 + monthly_rates[i % len(monthly_rates)])
                    
                    # Calculate reajusted value if applicable
                    if reajuste_ativo and indice_reajuste > 0:
                        # Calculate how many reajuste periods have passed
                        periodos_reajuste = (mes_ocorrencia - mes_inicial) // meses_reajuste
                        if periodos_reajuste > 0:
                            valor_atual = valor_atual * ((1 + indice_reajuste) ** periodos_reajuste)
                    
                    # Add the value to the specific month only
                    df_outras.loc["Outras Receitas", mes_ocorrencia] += valor_atual
            else:
                # Receita não recorrente
                for i in range(prazo):
                    if mes_inicial + i < 60:
                        # Calculate value with growth for this month
                        valor_atual = valor_base
                        if i > 0 and tx_crescimento_anual > 0:
                            # Use our pre-calculated monthly rates
                            if forma_crescimento == "Linear":
                                # For linear growth, apply cumulative sum of rates
                                cumulative_rate = sum(monthly_rates[:i])
                                valor_atual = valor_base * (1 + cumulative_rate)
                            else:  # "Composto"
                                # For compound growth, apply rates compounded
                                for j in range(i):
                                    valor_atual *= (1 + monthly_rates[j % len(monthly_rates)])
                        
                        # Apply reajuste for later months if applicable
                        if reajuste_ativo and indice_reajuste > 0:
                            # Calculate how many reajuste periods have passed
                            periodos_reajuste = i // meses_reajuste
                            if periodos_reajuste > 0:
                                valor_atual = valor_atual * ((1 + indice_reajuste) ** periodos_reajuste)
                        
                        df_outras.loc["Outras Receitas", mes_inicial + i] += valor_atual
        
        # Adicionar a linha de outras receitas ao DataFrame principal
        if not df_outras.empty and "Outras Receitas" not in df.index:
            df = pd.concat([df, df_outras])
        
        # Atualizar o total
        for mes in meses:
            # Convert both values to numeric to ensure proper addition
            total_atual = pd.to_numeric(df.loc["Total", mes], errors='coerce') or 0
            outras_receitas = pd.to_numeric(df_outras.loc["Outras Receitas", mes], errors='coerce') or 0
            df.loc["Total", mes] = total_atual + outras_receitas
     
        return df
    
    def _obter_capacidades_equipe(self):
        """Obtém as capacidades de leads e atendimentos configuradas para a equipe"""
        capacidade_sdr_leads = 750  # Valor padrão
        capacidade_closer_atendimentos = 90  # Valor padrão
        
        # Buscar dados de capacidade da equipe das premissas
        if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
            for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                # Verificar cada cargo na equipe própria
                if 'nome' in cargo:
                    nome_cargo = cargo['nome'].lower()
                    # Para SDRs
                    if 'sdr' in nome_cargo:
                        if 'capacidade_leads' in cargo:
                            capacidade_sdr_leads = cargo['capacidade_leads']
                    # Para Closers
                    elif 'closer' in nome_cargo or 'vendedor' in nome_cargo:
                        if 'capacidade_atendimentos' in cargo:
                            capacidade_closer_atendimentos = cargo['capacidade_atendimentos']
        
        return {
            'sdr': capacidade_sdr_leads,
            'closer': capacidade_closer_atendimentos
        }
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        if 'premissas_receitas' not in st.session_state:
            st.error("Premissas de receitas não definidas. Por favor, defina as premissas na página 'Premissas Receitas'.")
            return
        
        # Gerar os dataframes
        df_faturamento = self._gerar_dataframe_faturamento()
        
        if df_faturamento is None:
            st.error("Não foi possível gerar os dados de faturamento. Verifique os parâmetros nas premissas.")
            return
        
        df_arrecadacao = self._gerar_dataframe_arrecadacao(df_faturamento)
        
        # Criar tabs para Faturamento e Arrecadação
        tab1, tab2 = st.tabs(["Faturamento", "Arrecadação"])
        
        # Tab Faturamento
        with tab1:
            st.write("### Faturamento")
            
            # Opções de visualização
            col1, col2 = st.columns(2)
            
            with col1:
                time_frame = st.selectbox(
                    "Período", 
                    self.time_frames, 
                    index=0,
                    key="faturamento_time_frame"
                )
            
            with col2:
                plot_type = st.selectbox(
                    "Tipo de Gráfico", 
                    self.graph_types, 
                    index=0,
                    key="faturamento_plot_type"
                )
            
            # Agrupar dados conforme o período selecionado
            df_display = self._agrupar_por_periodo(df_faturamento, time_frame)
            
            if df_display is not None:
                # Mostrar o dataframe
                st.write(f"#### Faturamento - Visão {time_frame}")
                st.dataframe(df_display.style.format("{:.2f}"), use_container_width=True)
                
                # Seleção de categorias para visualização
                # Filtrar índices para remover categorias de cálculo interno
                filtro_indices = ["Total"]
                p_receitas = st.session_state['premissas_receitas']
                
                if p_receitas.get('modelo_marketing', True):
                    # Para modelo de marketing, adicionar as linhas de métricas de performance
                    metricas_indices = [
                        "Leads por Canal", 
                        "Total de Leads",
                        "Total de Agendamentos", 
                        "Total de Comparecimento", 
                        "Conversões",
                        "Faturamento Bruto", 
                        "ROAS"
                    ]
                    filtro_indices.extend(metricas_indices)
                else:
                    # Para modelo financeiro
                    filtro_indices.append("Faturamento Bruto")
                
                # Adicionar canais/fontes de receita
                for idx in df_display.index:
                    if idx not in filtro_indices:
                        filtro_indices.append(idx)
                
                # Seleção múltipla de categorias
                selected_categories = st.multiselect(
                    "Selecione as categorias para visualizar", 
                    filtro_indices,
                    default=["Total"],
                    key="faturamento_categories"
                )
                
                # Se nenhuma categoria foi selecionada, mostrar o total
                if not selected_categories:
                    selected_categories = ["Total"]
                
                # Preparar dados para o gráfico
                if selected_categories:
                    df_plot = df_display.loc[selected_categories].T  # Transpor para ter tempo no eixo x
                    
                    if plot_type == "Gráfico de Linhas":
                        fig = px.line(
                            df_plot,
                            y=selected_categories,
                            title=f"Evolução do Faturamento ({time_frame})",
                            markers=True
                        )
                        fig.update_layout(
                            xaxis_title="Período",
                            yaxis_title="Valor (R$)",
                            legend_title="Categoria"
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    
                    elif plot_type == "Gráfico de Pizza":
                        # Para gráfico de pizza, usamos a soma total de cada categoria
                        df_pie = pd.DataFrame({
                            'Categoria': selected_categories,
                            'Valor': [df_display.loc[cat].sum() for cat in selected_categories]
                        })
                        
                        fig = px.pie(
                            df_pie,
                            values='Valor',
                            names='Categoria',
                            title=f"Distribuição do Faturamento - Total ({time_frame})"
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    
                    elif plot_type == "Gráfico de Barras":
                        fig = px.bar(
                            df_plot,
                            y=selected_categories,
                            title=f"Evolução do Faturamento ({time_frame})"
                        )
                        fig.update_layout(
                            xaxis_title="Período",
                            yaxis_title="Valor (R$)",
                            legend_title="Categoria"
                        )
                        st.plotly_chart(fig, use_container_width=True)
        
        # After displaying the faturamento dataframe, add:
        if st.session_state['premissas_receitas']['modelo_marketing']:
            # Calculate commercial team data
            df_equipe_comercial = self._calcular_equipe_comercial(df_faturamento)
            if not df_equipe_comercial.empty:
                st.write("### Equipe Comercial Necessária")
                st.write("Com base no volume de leads e comparecimentos, a equipe comercial necessária seria:")
                
                # Show the capacities being used for calculation
                capacidades = self._obter_capacidades_equipe()
                st.info(f"Capacidades utilizadas: {capacidades['sdr']} leads/mês por SDR e {capacidades['closer']} atendimentos/mês por Closer.")
                
                # Store original monthly data before aggregation
                df_equipe_mensal = df_equipe_comercial.copy()
                
                # Convert to selected timeframe for display table and line chart
                df_equipe_display = self._agrupar_por_periodo(df_equipe_comercial, time_frame)
                
                # Display as integer values
                if df_equipe_display is not None:
                    st.dataframe(df_equipe_display.style.format("{:.0f}"), use_container_width=True)
                else:
                    st.info("Não há dados de equipe comercial para exibir.")
                
                # Create a line chart using the selected time frame data
                if df_equipe_display is not None:
                    df_plot = df_equipe_display.T
                    fig = px.line(
                        df_plot,
                        title=f"Evolução da Equipe Comercial ({time_frame})",
                        markers=True
                    )
                    fig.update_layout(
                        xaxis_title=f"{'Ano' if time_frame == 'Anual' else 'Período'}",
                        yaxis_title="Quantidade de Profissionais",
                        legend_title="Função"
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("Não há dados de equipe comercial para exibir.")
                
                # Create a bar chart always using monthly data
                if df_equipe_mensal is not None:
                    df_plot = df_equipe_mensal.T.reset_index()
                    df_plot.columns = ['Mês', 'SDR Necessários', 'Closer Necessários', 'Salário SDR', 'Salário Closer']
                    
                    fig = px.bar(
                        df_plot,
                        x='Mês',
                        y=['SDR Necessários', 'Closer Necessários'],
                        title="Evolução da Equipe Comercial (Mensal)",
                        barmode='group'
                    )
                    fig.update_layout(
                        xaxis_title="Mês",
                        yaxis_title="Quantidade",
                        legend_title="Cargo"
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("Não há dados suficientes para visualizar a evolução da equipe comercial.")
       
        # Tab Arrecadação
        # Inside the Receitas.render method, in the tab2 (Arrecadação) section
        
        # Tab Arrecadação
        with tab2:
            st.write("### Arrecadação")
            
            # Opções de visualização
            col1, col2 = st.columns(2)
            
            with col1:
                time_frame_arr = st.selectbox(
                    "Período", 
                    self.time_frames, 
                    index=0,
                    key="arrecadacao_time_frame"
                )
            
            with col2:
                plot_type_arr = st.selectbox(
                    "Tipo de Gráfico", 
                    self.graph_types, 
                    index=0,
                    key="arrecadacao_plot_type"
                )
            
            # Agrupar dados conforme o período selecionado
            df_display_arr = self._agrupar_por_periodo(df_arrecadacao, time_frame_arr)
            
            if df_display_arr is not None:
                st.dataframe(df_display_arr.style.format("{:.2f}"), use_container_width=True)
                
                # Visualização gráfica
                st.write("### Visualização Gráfica")
                
                # Prepare data for plotting - transpose for correct orientation
                df_plot_arr = df_display_arr.T
                
                # Prepare data for plotting
                if df_display_arr is not None:
                    if plot_type_arr == "Gráfico de Linhas":
                        fig_arr = px.line(
                            df_plot_arr, 
                            title=f"Evolução da Arrecadação ({time_frame_arr})",
                            markers=True
                        )
                        fig_arr.update_layout(
                            xaxis_title=f"{'Ano' if time_frame_arr == 'Anual' else 'Período'}",
                            yaxis_title="Valor (R$)",
                            legend_title="Tipo"
                        )
                        st.plotly_chart(fig_arr, use_container_width=True)
                        
                    elif plot_type_arr == "Gráfico de Pizza":
                        # Pie chart for total distribution
                        data_pie = {
                            'Categoria': df_display_arr.index.tolist(),
                            'Valor': [df_display_arr.loc[cat].sum() for cat in df_display_arr.index]
                        }
                        df_pie = pd.DataFrame(data_pie)
                        
                        fig_arr = px.pie(
                            df_pie,
                            values='Valor',
                            names='Categoria',
                            title=f"Distribuição da Arrecadação - Total ({time_frame_arr})"
                        )
                        st.plotly_chart(fig_arr, use_container_width=True)
                        
                    elif plot_type_arr == "Gráfico de Barras":
                        # Bar chart with proper formatting
                        df_plot_arr = df_display_arr.T.reset_index()  # Reset index to make it a column
                        
                        fig_arr = px.bar(
                            df_plot_arr,
                            x='index',  # Use the reset index column as x-axis
                            y=df_display_arr.index.tolist(),  # Use the original dataframe's index as y values
                            title=f"Evolução da Arrecadação ({time_frame_arr})",
                            barmode='group'
                        )
                        
                        fig_arr.update_layout(
                            xaxis_title=f"{'Ano' if time_frame_arr == 'Anual' else 'Período'}",
                            yaxis_title="Valor (R$)",
                            legend_title="Tipo"
                        )
                        st.plotly_chart(fig_arr, use_container_width=True)
                else:
                    st.info("Não há dados de arrecadação para exibir.")
      
        # Add metrics section
        st.write("### Métricas de Receita")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Calculate total revenue across all months
            total_faturamento = df_faturamento.loc["Total"].sum()
            st.metric(
                label="Faturamento Total (60 meses)", 
                value=f"R$ {total_faturamento:,.2f}"
            )
        
        with col2:
            # Calculate total net revenue across all months
            if df_arrecadacao is not None:
                total_arrecadacao = df_arrecadacao.loc["Arrecadação Líquida"].sum()
                st.metric(
                    label="Arrecadação Líquida Total (60 meses)", 
                    value=f"R$ {total_arrecadacao:,.2f}"
                )
        with col3:
            # Calculate percentage of revenue collected
            if df_arrecadacao is not None and isinstance(total_faturamento, (int, float)) and total_faturamento > 0:
                total_arrecadacao = df_arrecadacao.loc["Arrecadação Líquida"].sum()
                percentual_arrecadacao = (total_arrecadacao / total_faturamento * 100)
                st.metric(
                    label="% Arrecadação/Faturamento", 
                    value=f"{percentual_arrecadacao:.1f}%"
                )
            else:
                st.metric(
                    label="% Arrecadação/Faturamento", 
                    value="0.0%"
                )

class PremissasTributos(Page):

    def __init__(self):
        # Default values for tax parameters
        default_params = {
            # Tax regime
            'regime_tributario': 'Simples Nacional',
            'regime_apos_desenquadramento': 'Lucro Presumido',
            'ano_desenquadramento': None,
            
            # Tax types
            'pis_ativo': False,
            'cofins_ativo': False,
            'issqn_ativo': False,
            'icms_ativo': False,
            'ie_ativo': False,
            'ii_ativo': False,
            'ipi_ativo': False,
            'csll_ativo': False,
            'irpj_ativo': False,
            
            # Tax rates - standard
            'valor_pis': 0.65,
            'valor_cofins': 3.0,
            'valor_issqn': 5.0,
            'valor_icms': 18.0,
            'valor_ie': 2.0,
            'valor_ii': 10.0,
            'valor_ipi': 10.0,
            'valor_csll': 9.0,
            'valor_irpj': 15.0,
            
            # Tax rates - Simples Nacional
            'valor_pis_simples': 0.65,
            'valor_cofins_simples': 3.0,
            'valor_issqn_simples': 5.0,
            'valor_csll_simples': 9.0,
            'valor_irpj_simples': 15.0,
            
            # Additional tax calculation parameters
            'base_calculo_icms': 100.0,
            'valor_mercadoria_ie': 0.0,
            'quantidade_mercadoria_ie': 0,
            'valor_aduaneiro_ii': 0.0,
            'quantidade_mercadoria_ii': 0,
        }
        
        # Initialize or update session state
        if 'premissas_tributos' not in st.session_state:
            st.session_state['premissas_tributos'] = default_params
        else:
            # Check for any missing keys and add them with default values
            for key, value in default_params.items():
                if key not in st.session_state['premissas_tributos']:
                    st.session_state['premissas_tributos'][key] = value
    
    @property
    def title(self) -> str:
        return "Premissas Tributos"
    
    @property
    def icon(self) -> str:
        return "📋"
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Regime Tributário selection
        st.write("### Regime Tributário")
        st.session_state['premissas_tributos']['regime_tributario'] = st.selectbox(
            "Selecione o Regime Tributário",
            ["Simples Nacional", "Lucro Real", "Lucro Presumido"],
            index=["Simples Nacional", "Lucro Real", "Lucro Presumido"].index(
                st.session_state['premissas_tributos']['regime_tributario']
            )
        )
        
        # Check if the company exceeds the Simples Nacional limit
        excede_limite = False
        ano_desenquadramento = None
        
        if st.session_state['premissas_tributos']['regime_tributario'] == "Simples Nacional":
            try:
                # Try to get revenue data from Receitas class
                receitas = Receitas()
                df_faturamento = receitas._gerar_dataframe_faturamento()
                
                if df_faturamento is not None and "Faturamento" in df_faturamento.index:
                    # Check each year for the 4.8 million limit
                    for ano in range(1, 6):  # 5 years
                        inicio = (ano - 1) * 12
                        fim = ano * 12 - 1
                        faturamento_anual = df_faturamento.loc["Faturamento", inicio:fim].sum()
                        # Convert to scalar value if it's a pandas Series
                        if isinstance(faturamento_anual, pd.Series):
                            faturamento_anual = faturamento_anual.sum()
                        if float(faturamento_anual) > 4800000:  # 4.8 million limit
                            excede_limite = True
                            ano_desenquadramento = ano
                            st.session_state['premissas_tributos']['ano_desenquadramento'] = ano
                            break
                 
                    if excede_limite:
                        if ano_desenquadramento is not None:
                            st.warning(f"A empresa está enquadrada no regime fiscal escolhido até {ano_desenquadramento-1}, CUIDADO!")
                        else:
                            st.warning("Ano de desenquadramento não definido. CUIDADO!")
                     
                        # Option to select a new regime after exceeding the limit
                        st.write("### Regime fiscal a partir do desenquadramento")
                        st.session_state['premissas_tributos']['regime_apos_desenquadramento'] = st.selectbox(
                            "Selecione o Regime após desenquadramento",
                            ["Lucro Real", "Lucro Presumido"],
                            index=["Lucro Real", "Lucro Presumido"].index(
                                st.session_state['premissas_tributos'].get('regime_apos_desenquadramento', "Lucro Presumido")
                            )
                        )
                    else:
                           # Check total annual revenue against the limit
                        faturamento_total = df_faturamento.loc["Faturamento"].sum()
                        # Convert faturamento_total to float if it's a pandas Series
                        if isinstance(faturamento_total, pd.Series):
                            faturamento_total = float(faturamento_total.sum())
                        else:
                            faturamento_total = float(faturamento_total)
                            
                        if faturamento_total > 4800000 * 5:  # Simple check for 5 years
                            st.warning("CUIDADO! Empresa desenquadrada do regime tributário selecionado, considerando faturamento anual apurado!")
            except Exception as e:
                st.warning(f"Erro ao verificar o faturamento para validação do regime tributário: {e}")
      
        st.write("### Tributos em Espécie")
        st.write("Selecione os tributos que incidem sobre a sua empresa e defina as alíquotas correspondentes.")
        
        # Dictionary of taxes with their display names
        tributos = {
            'pis': 'PIS',
            'cofins': 'COFINS',
            'issqn': 'ISSQN',
            'icms': 'ICMS',
            'ie': 'IE',
            'ii': 'II',
            'ipi': 'IPI',
            'csll': 'CSLL',
            'irpj': 'IRPJ'
        }
        
        # Create columns for better layout
        col1, col2 = st.columns(2)
        
        count = 0
        for codigo, nome in tributos.items():
            # Alternate between columns
            current_col = col1 if count % 2 == 0 else col2
            
            with current_col:
                # Create an expander for each tax
                with st.expander(f"{nome}"):
                    # Checkbox to activate the tax
                    st.session_state['premissas_tributos'][f'{codigo}_ativo'] = st.checkbox(
                        f"Aplicar {nome}",
                        value=st.session_state['premissas_tributos'][f'{codigo}_ativo'],
                        key=f"checkbox_{codigo}",
                        help=f"Ativar ou desativar o cálculo do tributo {nome}."
                    )
                    
                    # Only show the rate inputs if the tax is active
                    if st.session_state['premissas_tributos'][f'{codigo}_ativo']:
                        # Special handling for Simples Nacional taxes
                        if codigo in ['pis', 'cofins', 'issqn', 'csll', 'irpj'] and st.session_state['premissas_tributos']['regime_tributario'] == "Simples Nacional":
                            # Show both Simples Nacional rate and post-disqualification rate
                            st.session_state['premissas_tributos'][f'valor_{codigo}_simples'] = st.number_input(
                                f"Valor {nome} no Simples Nacional (%)",
                                min_value=0.0,
                                max_value=100.0,
                                value=st.session_state['premissas_tributos'][f'valor_{codigo}_simples'],
                                step=0.01,
                                format="%.2f",
                                key=f"valor_{codigo}_simples",
                                help="Alíquota aplicada enquanto a empresa está no Simples Nacional."
                            )
                            
                            st.session_state['premissas_tributos'][f'valor_{codigo}'] = st.number_input(
                                f"Valor {nome} após desenquadramento (%)",
                                min_value=0.0,
                                max_value=100.0,
                                value=st.session_state['premissas_tributos'][f'valor_{codigo}'],
                                step=0.01,
                                format="%.2f",
                                key=f"valor_{codigo}_apos",
                                help="Alíquota aplicada após o desenquadramento do Simples Nacional, caso aplicável."
                            )
                        else:
                            # Standard tax rate input
                            st.session_state['premissas_tributos'][f'valor_{codigo}'] = st.number_input(
                                f"Valor {nome} (%)",
                                min_value=0.0,
                                max_value=100.0,
                                value=st.session_state['premissas_tributos'][f'valor_{codigo}'],
                                step=0.01,
                                format="%.2f",
                                key=f"valor_{codigo}",
                                help=f"Alíquota aplicada para o tributo {nome}."
                            )
                        
                        # Additional fields for specific taxes
                        if codigo == 'icms':
                            st.session_state['premissas_tributos']['base_calculo_icms'] = st.number_input(
                                "Base de Cálculo ICMS (unitário)",
                                min_value=0.0,
                                value=st.session_state['premissas_tributos'].get('base_calculo_icms', 100.0),
                                step=10.0,
                                format="%.2f",
                                key="base_calculo_icms",
                                help="Valor unitário da mercadoria para cálculo do ICMS."
                            )
                        
                        elif codigo == 'ie':
                            col1_ie, col2_ie = st.columns(2)
                            with col1_ie:
                                st.session_state['premissas_tributos']['valor_mercadoria_ie'] = st.number_input(
                                    "Valor unitário da mercadoria no local de embarque",
                                    min_value=0.0,
                                    value=st.session_state['premissas_tributos'].get('valor_mercadoria_ie', 0.0),
                                    step=10.0,
                                    format="%.2f",
                                    key="valor_mercadoria_ie",
                                    help="Valor unitário da mercadoria para cálculo do IE (Imposto de Exportação)."
                                )
                            with col2_ie:
                                st.session_state['premissas_tributos']['quantidade_mercadoria_ie'] = st.number_input(
                                    "Quantidade da mercadoria no local de embarque",
                                    min_value=0,
                                    value=st.session_state['premissas_tributos'].get('quantidade_mercadoria_ie', 0),
                                    step=1,
                                    key="quantidade_mercadoria_ie",
                                    help="Quantidade da mercadoria para cálculo do IE (Imposto de Exportação)."
                                )
                        
                        elif codigo == 'ii':
                            col1_ii, col2_ii = st.columns(2)
                            with col1_ii:
                                st.session_state['premissas_tributos']['valor_aduaneiro_ii'] = st.number_input(
                                    "Valor Aduaneiro",
                                    min_value=0.0,
                                    value=st.session_state['premissas_tributos'].get('valor_aduaneiro_ii', 0.0),
                                    step=10.0,
                                    format="%.2f",
                                    key="valor_aduaneiro_ii",
                                    help="Valor aduaneiro da mercadoria para cálculo do II (Imposto de Importação)."
                                )
                            with col2_ii:
                                st.session_state['premissas_tributos']['quantidade_mercadoria_ii'] = st.number_input(
                                    "Quantidade da mercadoria na aduana",
                                    min_value=0,
                                    value=st.session_state['premissas_tributos'].get('quantidade_mercadoria_ii', 0),
                                    step=1,
                                    key="quantidade_mercadoria_ii",
                                    help="Quantidade da mercadoria para cálculo do II (Imposto de Importação)."
                                )
            
            count += 1
        
        # Show summary of selected taxes
        st.write("### Resumo dos Tributos Selecionados")
        
        tributos_selecionados = []
        for codigo, nome in tributos.items():
            if st.session_state['premissas_tributos'][f'{codigo}_ativo']:
                if codigo in ['pis', 'cofins', 'issqn', 'csll', 'irpj'] and st.session_state['premissas_tributos']['regime_tributario'] == "Simples Nacional":
                    # Show both rates for Simples Nacional taxes
                    tributos_selecionados.append({
                        "Tributo": nome,
                        "Alíquota Simples Nacional (%)": f"{st.session_state['premissas_tributos'][f'valor_{codigo}_simples']:.2f}",
                        "Alíquota após desenquadramento (%)": f"{st.session_state['premissas_tributos'][f'valor_{codigo}']:.2f}"
                    })
                else:
                    # Show standard rate
                    tributos_selecionados.append({
                        "Tributo": nome,
                        "Alíquota (%)": f"{st.session_state['premissas_tributos'][f'valor_{codigo}']:.2f}"
                    })
                    
                    # Show additional parameters for special taxes
                    if codigo == 'icms':
                        tributos_selecionados[-1]["Base de Cálculo (un.)"] = f"R$ {st.session_state['premissas_tributos']['base_calculo_icms']:.2f}"
                    elif codigo == 'ie':
                        tributos_selecionados[-1]["Valor Unitário"] = f"R$ {st.session_state['premissas_tributos']['valor_mercadoria_ie']:.2f}"
                        tributos_selecionados[-1]["Quantidade"] = f"{st.session_state['premissas_tributos']['quantidade_mercadoria_ie']}"
                    elif codigo == 'ii':
                        tributos_selecionados[-1]["Valor Aduaneiro"] = f"R$ {st.session_state['premissas_tributos']['valor_aduaneiro_ii']:.2f}"
                        tributos_selecionados[-1]["Quantidade"] = f"{st.session_state['premissas_tributos']['quantidade_mercadoria_ii']}"
        
        if tributos_selecionados:
            df_tributos = pd.DataFrame(tributos_selecionados)
            st.dataframe(df_tributos, use_container_width=True)
        else:
            st.info("Nenhum tributo selecionado.")

class Tributos(Page):
    def __init__(self):
        self.graph_types = ["Gráfico de Barras", "Gráfico de Linhas", "Gráfico de Pizza"]
        self.time_frames = ["Mensal", "Anual"]
    
    @property
    def title(self) -> str:
        return "Tributos"
    
    @property
    def icon(self) -> str:
        return "💰"
    
    def _gerar_dataframe_tributos(self):
        """Gera o dataframe de tributos para 60 meses (5 anos)"""
        if 'premissas_tributos' not in st.session_state or 'premissas_receitas' not in st.session_state:
            return None
        
        # Get tax parameters and regime information
        p_tributos = st.session_state['premissas_tributos']
        regime_tributario = p_tributos['regime_tributario']
        regime_apos_desenquadramento = p_tributos.get('regime_apos_desenquadramento', "Lucro Presumido")
        ano_desenquadramento = p_tributos.get('ano_desenquadramento', None)
        
        # Get revenue data from Receitas class
        receitas = Receitas()
        df_faturamento = receitas._gerar_dataframe_faturamento()
        df_arrecadacao = receitas._gerar_dataframe_arrecadacao(df_faturamento)
        
        if df_arrecadacao is None or df_faturamento is None:
            return None
        
        # Get costs data from CustosTecnologia
        try:
            tecnologia = CustosTecnologia()
            df_tecnologia = tecnologia._gerar_dataframe_custos_tecnologia()
        except Exception as e:
            st.warning(f"Erro ao obter dados de custos de tecnologia: {e}")
            df_tecnologia = None
        
        # Get admin expenses from DespesasAdm
        try:
            despesas_adm = DespesasAdm()
            df_despesas = despesas_adm._gerar_dataframe_despesas()
        except Exception as e:
            st.warning(f"Erro ao obter dados de despesas administrativas: {e}")
            df_despesas = None
        
        # Get team expenses from Equipe
        try:
            equipe = Equipe()
            df_equipe = equipe._gerar_dataframe_custos_equipe()
        except Exception as e:
            st.warning(f"Erro ao obter dados de equipe: {e}")
            df_equipe = None
        
        # Get sales volume data if needed for ICMS calculations
        try:
            # Replace GeracaoDeCaixa with Receitas and derive sales volume from faturamento
            vendas_liquidas = None
            if df_faturamento is not None and "Total" in df_faturamento.index:
                # Estimate number of sales based on faturamento and ticket médio
                ticket_medio = st.session_state.get('premissas_receitas', {}).get('ticket_medio', 2400.0)
                if ticket_medio > 0:
                    vendas_liquidas = df_faturamento.loc["Total"] / ticket_medio
        except Exception as e:
            st.warning(f"Erro ao obter dados de vendas: {e}")
            vendas_liquidas = None
        
        # Create DataFrame for 60 months
        meses = range(60)
        
        # Determine which taxes are active
        tributos_ativos = []
        if p_tributos['pis_ativo']:
            tributos_ativos.append('PIS')
        if p_tributos['cofins_ativo']:
            tributos_ativos.append('COFINS')
        if p_tributos['issqn_ativo']:
            tributos_ativos.append('ISSQN')
        if p_tributos['icms_ativo']:
            tributos_ativos.append('ICMS')
        if p_tributos['ie_ativo']:
            tributos_ativos.append('IE')
        if p_tributos['ii_ativo']:
            tributos_ativos.append('II')
        if p_tributos['ipi_ativo']:
            tributos_ativos.append('IPI')
        if p_tributos['csll_ativo']:
            tributos_ativos.append('CSLL')
        if p_tributos['irpj_ativo']:
            tributos_ativos.append('IRPJ')
      
        # Add the total always
        all_tributos = tributos_ativos + ["Tributo Total"]
        
        # Create the DataFrame
        df = pd.DataFrame(index=pd.Index(all_tributos), columns=meses)
        df.fillna(0, inplace=True)
        
        # Fill values for each month
        for mes in meses:
            # Determine current tax regime based on the month/year
            regime_atual = regime_tributario
            if ano_desenquadramento is not None:
                ano_atual = mes // 12 + 1
                if ano_atual >= ano_desenquadramento:
                    regime_atual = regime_apos_desenquadramento
            
            # Get necessary values for calculations
            arrecadacao = pd.to_numeric(df_arrecadacao.loc["Arrecadação Líquida", mes], errors='coerce') or 0
            faturamento = pd.to_numeric(df_faturamento.loc["Total", mes], errors='coerce') or 0
            
            # Get sales volume for ICMS calculation - use derived value from faturamento
            vendas_liquidas_valor = 0
            if vendas_liquidas is not None:
                vendas_liquidas_valor = pd.to_numeric(vendas_liquidas[mes], errors='coerce') or 0
            
            # Initialize tributo_total for this month
            tributo_total = 0
            
            # Variables for tax calculation that depend on other taxes
            pis_valor = 0
            cofins_valor = 0
            issqn_valor = 0
            irpj_presumido_valor = 0
            
            # Calculate revenue-based taxes first
            
            # Calculate PIS
            if p_tributos['pis_ativo']:
                if regime_atual == "Simples Nacional":
                    pis_valor = arrecadacao * (p_tributos['valor_pis_simples'] / 100)
                else:
                    pis_valor = arrecadacao * (p_tributos['valor_pis'] / 100)
                    
                df.loc['PIS', mes] = pis_valor
                tributo_total += pis_valor
            
            # Calculate COFINS
            if p_tributos['cofins_ativo']:
                if regime_atual == "Simples Nacional":
                    cofins_valor = arrecadacao * (p_tributos['valor_cofins_simples'] / 100)
                else:
                    cofins_valor = arrecadacao * (p_tributos['valor_cofins'] / 100)
                    
                df.loc['COFINS', mes] = cofins_valor
                tributo_total += cofins_valor
            
            # Calculate ISSQN
            if p_tributos['issqn_ativo']:
                if regime_atual == "Simples Nacional":
                    issqn_valor = arrecadacao * (p_tributos['valor_issqn_simples'] / 100)
                else:
                    issqn_valor = arrecadacao * (p_tributos['valor_issqn'] / 100)
                    
                df.loc['ISSQN', mes] = issqn_valor
                tributo_total += issqn_valor
            
            # Calculate IRPJ for Lucro Presumido (revenue-based)
            irpj_presumido_valor = 0
            if p_tributos['irpj_ativo'] and regime_atual == "Lucro Presumido":
                irpj_presumido_valor = faturamento * (p_tributos['valor_irpj'] / 100)
                df.loc['IRPJ', mes] = irpj_presumido_valor
                tributo_total += irpj_presumido_valor
            
            # Calculate base for profit taxes (CSLL and IRPJ for non-Presumido)
            # Base = Revenue - Revenue taxes - Costs - Admin expenses - Team expenses
            
            # Get technology costs
            custos_producao = 0
            if df_tecnologia is not None and "Total" in df_tecnologia.index:
                custos_producao = pd.to_numeric(df_tecnologia.loc["Total", mes], errors='coerce') or 0
            
            # Add dummy production costs if needed
            custos_producao_adicionais = 0
            
            # Get admin expenses
            despesas_administrativas = 0
            if df_despesas is not None and mes < len(df_despesas):
                if "Total" in df_despesas.columns:
                    despesas_administrativas = pd.to_numeric(df_despesas.loc[mes, "Total"], errors='coerce') or 0
            
            # Get team expenses
            despesas_equipe = 0
            if df_equipe is not None:
                for idx in df_equipe.index:
                    if idx[0] == "TOTAL" and idx[1] == "Total Custos de Equipe":
                        if mes+1 <= df_equipe.shape[1]:  # +1 because df_equipe is 1-indexed
                            despesas_equipe = pd.to_numeric(df_equipe.loc[idx, mes+1], errors='coerce') or 0
                        break
            
            # Calculate profit base for CSLL and IRPJ
            base_lucro = (
                arrecadacao 
                - (pis_valor + cofins_valor + issqn_valor + (irpj_presumido_valor if regime_atual == "Lucro Presumido" else 0))
                - (custos_producao + custos_producao_adicionais)
                - despesas_administrativas
                - despesas_equipe
            )
            
            # Calculate CSLL based on profit
            if p_tributos['csll_ativo']:
                if regime_atual == "Simples Nacional":
                    csll_valor = max(0, base_lucro * (p_tributos['valor_csll_simples'] / 100))
                else:
                    csll_valor = max(0, base_lucro * (p_tributos['valor_csll'] / 100))
                    
                df.loc['CSLL', mes] = csll_valor
                tributo_total += csll_valor
            
            # Calculate IRPJ for Lucro Real or Simples Nacional (profit-based)
            if p_tributos['irpj_ativo'] and regime_atual != "Lucro Presumido":
                if regime_atual == "Simples Nacional":
                    irpj_valor = max(0, base_lucro * (p_tributos['valor_irpj_simples'] / 100))
                else:  # Lucro Real
                    irpj_valor = max(0, base_lucro * (p_tributos['valor_irpj'] / 100))
                
                df.loc['IRPJ', mes] = irpj_valor
                tributo_total += irpj_valor
            
            # Calculate ICMS
            if p_tributos['icms_ativo']:
                base_calculo = p_tributos.get('base_calculo_icms', 100.0)
                icms_valor = (base_calculo * (p_tributos['valor_icms'] / 100)) * vendas_liquidas_valor
                df.loc['ICMS', mes] = icms_valor
                tributo_total += icms_valor
            
            # Calculate IE
            if p_tributos['ie_ativo']:
                valor_mercadoria = p_tributos.get('valor_mercadoria_ie', 0.0)
                quantidade = p_tributos.get('quantidade_mercadoria_ie', 0)
                ie_valor = (valor_mercadoria * (p_tributos['valor_ie'] / 100)) * quantidade
                df.loc['IE', mes] = ie_valor
                tributo_total += ie_valor
            
            # Calculate II
            if p_tributos['ii_ativo']:
                valor_aduaneiro = p_tributos.get('valor_aduaneiro_ii', 0.0)
                quantidade = p_tributos.get('quantidade_mercadoria_ii', 0)
                ii_valor = (valor_aduaneiro * (p_tributos['valor_ii'] / 100)) * quantidade
                df.loc['II', mes] = ii_valor
                tributo_total += ii_valor
            
            # IPI placeholder
            if p_tributos['ipi_ativo']:
                df.loc['IPI', mes] = 0
            
            # Store total tax for this month
            df.loc['Tributo Total', mes] = tributo_total
        
        return df
    
    def _gerar_dataframe_anual(self, df_mensal):
        """Converte o dataframe mensal para anual"""
        if df_mensal is None:
            return None
            
        anos = 5
        df_anual = pd.DataFrame(index=df_mensal.index, columns=range(1, anos + 1))
        
        for ano in range(1, anos + 1):
            mes_inicio = (ano - 1) * 12
            mes_fim = ano * 12 - 1
            
            # Para cada categoria, somar os valores dos meses do ano
            for categoria in df_mensal.index:
                df_anual.loc[categoria, ano] = df_mensal.loc[categoria, mes_inicio:mes_fim].sum()
        
        return df_anual
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Check for prerequisites
        if 'premissas_tributos' not in st.session_state:
            st.error("Premissas de tributos não definidas. Por favor, defina as premissas na página 'Premissas Tributos'.")
            return
        
        # Verificar se há tributos selecionados
        p_tributos = st.session_state['premissas_tributos']
        tributos_ativos = any([
            p_tributos['pis_ativo'], 
            p_tributos['cofins_ativo'], 
            p_tributos['issqn_ativo'],
            p_tributos['icms_ativo'],
            p_tributos['ie_ativo'],
            p_tributos['ii_ativo'],
            p_tributos['ipi_ativo'],
            p_tributos['csll_ativo'],
            p_tributos['irpj_ativo']
        ])
        
        if not tributos_ativos:
            st.warning("Nenhum tributo foi selecionado nas premissas. Por favor, selecione pelo menos um tributo para visualizar os resultados.")
            return
        
        # Generate tax DataFrame
        df_mensal = self._gerar_dataframe_tributos()
        
        if df_mensal is None:
            st.error("Não foi possível gerar os dados de tributos. Verifique se os dados de receitas estão disponíveis.")
            return
        
        # Visualization options
        col1, col2 = st.columns(2)
        
        with col1:
            time_frame = st.selectbox("Período", self.time_frames, index=0)
        
        with col2:
            plot_type = st.selectbox("Tipo de Gráfico", self.graph_types, index=0)
        
        # Convert to annual if selected
        df_display = df_mensal.copy()
        if time_frame == "Anual":
            df_display = self._gerar_dataframe_anual(df_mensal)
            if df_display is not None:
                df_display.columns = [f"Ano {ano}" for ano in df_display.columns]
        else:
            if df_display is not None:
                df_display.columns = [f"Mês {mes}" for mes in df_display.columns]
        
        # Display the tax DataFrame
        periodo = "Anual" if time_frame == "Anual" else "Mensal"
        st.write(f"### Tributos - Visão {periodo}")
        if df_display is not None:
            st.dataframe(df_display.style.format("{:.2f}"), use_container_width=True)
        
        # Select taxes for visualization
        tributos = df_display.index.tolist() if df_display is not None else []
       
        # Default selection is the Total category
        selected_tributos = st.multiselect(
            "Selecione os tributos para visualizar", 
            tributos,
            default=["Tributo Total"] if "Tributo Total" in tributos else []
        )
        
        # If no selection, use the total by default
        if not selected_tributos:
            if "Tributo Total" in tributos:
                selected_tributos = ["Tributo Total"]
            elif len(tributos) > 0:
                selected_tributos = [tributos[0]]
        
        # Prepare data for plotting
        if selected_tributos and df_display is not None:
            if plot_type == "Gráfico de Barras":
                # Bar chart for tax comparison
                df_plot = df_display.loc[selected_tributos].T
                
                df_plot = df_plot.reset_index()
                index_name = df_plot.columns[0]
                
                fig = px.bar(
                    df_plot,
                    x=index_name,
                    y=selected_tributos,
                    title=f"Comparação de Tributos ({periodo})",
                    barmode='group'
                )
                
                fig.update_layout(
                    xaxis_title=f"{'Ano' if time_frame == 'Anual' else 'Mês'}",
                    yaxis_title="Valor (R$)",
                    legend_title="Tributo"
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
            elif plot_type == "Gráfico de Linhas":
                # Line chart for tax evolution over time
                df_plot = df_display.loc[selected_tributos].T
                
                fig = px.line(
                    df_plot,
                    y=selected_tributos,
                    title=f"Evolução dos Tributos ({periodo})",
                    markers=True
                )
                
                fig.update_layout(
                    xaxis_title=f"{'Ano' if time_frame == 'Anual' else 'Mês'}",
                    yaxis_title="Valor (R$)",
                    legend_title="Tributo"
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
            else:  # Pizza chart
                # If annual, use sum of years; if monthly, use last month
                if time_frame == "Anual":
                    data_pie = {
                        'Categoria': selected_tributos,
                        'Valor': [df_display.loc[tributo].sum() for tributo in selected_tributos]
                    }
                    title_suffix = f"Total ({periodo})"
                else:
                    # Use the last month for pie chart
                    last_month = df_display.columns[-1]
                    data_pie = {
                        'Categoria': selected_tributos,
                        'Valor': [df_display.loc[tributo, last_month] for tributo in selected_tributos]
                    }
                    title_suffix = f"Último Mês ({last_month})"
                
                df_pie = pd.DataFrame(data_pie)
                
                # Ensure all values are positive for the pie chart
                df_pie['Valor'] = df_pie['Valor'].abs()
                
                fig = px.pie(
                    df_pie,
                    values='Valor',
                    names='Categoria',
                    title=f"Distribuição dos Tributos - {title_suffix}"
                )
                
                st.plotly_chart(fig, use_container_width=True)
            
            # Main metrics section remains unchanged
            st.write("### Métricas Principais")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                total_tributos = df_mensal.loc["Tributo Total"].sum() if "Tributo Total" in df_mensal.index else 0
                st.metric(
                    label="Total de Tributos (60 meses)", 
                    value=f"R$ {total_tributos:,.2f}"
                )
            
            with col2:
                if time_frame == "Anual":
                    media_tributos = df_display.loc["Tributo Total"].mean() if "Tributo Total" in df_display.index else 0
                    st.metric(
                        label="Média Anual de Tributos", 
                        value=f"R$ {media_tributos:,.2f}"
                    )
                else:
                    media_tributos = df_mensal.loc["Tributo Total"].mean() if "Tributo Total" in df_mensal.index else 0
                    st.metric(
                        label="Média Mensal de Tributos", 
                        value=f"R$ {media_tributos:,.2f}"
                    )
            
            with col3:
                try:
                    receitas = Receitas()
                    df_faturamento = receitas._gerar_dataframe_faturamento()
                    df_arrecadacao = receitas._gerar_dataframe_arrecadacao(df_faturamento)
                    
                    if df_arrecadacao is not None and "Arrecadação Líquida" in df_arrecadacao.index:
                        arrecadacao_total = df_arrecadacao.loc["Arrecadação Líquida"].sum()
                        # Convert to scalar if it's a pandas Series
                        if isinstance(arrecadacao_total, pd.Series):
                            arrecadacao_total = float(arrecadacao_total.sum())
                        else:
                            arrecadacao_total = float(arrecadacao_total)
                        carga_tributaria = (total_tributos / arrecadacao_total * 100) if arrecadacao_total > 0 else 0
                        st.metric(
                            label="Carga Tributária", 
                            value=f"{carga_tributaria:.2f}%"
                        )
                    else:
                        st.metric(label="Carga Tributária", value="N/D")
                except Exception:
                    st.metric(label="Carga Tributária", value="N/D")

class PremissasProjeções(Page):
    def __init__(self):
        # Default values for projection parameters
        default_params = {
            # Loans
            'emprestimos': [],
            # New parameters for DRE
            'custos_cobranca': 0.0,
            'custos_juridicos': 0.0,
            
        }
        
        # Initialize or update session state
        if 'premissas_projecoes' not in st.session_state:
            st.session_state['premissas_projecoes'] = default_params
        else:
            # Check for any missing keys and add them with default values
            for key, value in default_params.items():
                if key not in st.session_state['premissas_projecoes']:
                    st.session_state['premissas_projecoes'][key] = value
    
    @property
    def title(self) -> str:
        return "Premissas Projeções"
    
    @property
    def icon(self) -> str:
        return "📊"
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Create tabs for cash flow and DRE
        tab1, tab2 = st.tabs(["Fluxo de Caixa", "DRE"])
        
        with tab1:
            st.write("### Configurações para Projeção de Fluxo de Caixa")
            
            # --- Empréstimos --- (now at top level within the tab)
            st.write("#### Empréstimos")
            emp_expander = st.expander("Adicionar/Editar Empréstimos")
            with emp_expander:
                # Display existing loans
                if st.session_state['premissas_projecoes']['emprestimos']:
                    st.write("Empréstimos configurados:")
                    for idx, emp in enumerate(st.session_state['premissas_projecoes']['emprestimos']):
                        st.write(f"**{idx+1}. {emp['nome']}**")
                        cols = st.columns(2)
                        with cols[0]:
                            st.write(f"Valor: R$ {emp['valor']:,.2f}")
                            st.write(f"Mês Inflow: {emp['mes_inflow']}")
                            st.write(f"Taxa de Juros: {emp['taxa_juros']}%")
                        with cols[1]:
                            if emp.get('termo_caixa', False):
                                st.write("**Termo Caixa**")
                                st.write(f"Carência: {emp['periodo_carencia']} meses")
                            else:
                                st.write(f"Prazo: {emp['prazo']} meses")
                                if emp['carencia']:
                                    st.write(f"Carência: {emp['periodo_carencia']} meses")
                                else:
                                    st.write("Sem carência")
                        
                        # Delete button for this loan
                        if st.button(f"Remover empréstimo {emp['nome']}", key=f"del_emp_{idx}"):
                            st.session_state['premissas_projecoes']['emprestimos'].pop(idx)
                            st.rerun()
                        
                        st.markdown("---")
             
                with emp_expander:
                    with st.form("novo_emprestimo"):
                        nome_emp = st.text_input("Nome do Empréstimo", key="emp_nome", help="Nome ou identificador do empréstimo.")
                        
                        # Add checkbox for "Termo Caixa"
                        termo_caixa = st.checkbox("Termo Caixa", key="emp_termo_caixa", 
                                                 help="Quando marcado, o empréstimo terá um único pagamento após o período de carência.")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            valor_emp = st.number_input(
                                "Valor Líquido (R$)", 
                                min_value=0.0, 
                                step=1000.0, 
                                key="emp_valor", 
                                help="Valor líquido do empréstimo que será recebido."
                            )
                            
                            if not termo_caixa:
                                prazo_emp = st.number_input(
                                    "Prazo (meses)", 
                                    min_value=1, 
                                    max_value=360, 
                                    value=12, 
                                    step=1, 
                                    key="emp_prazo", 
                                    help="Prazo total do empréstimo em meses."
                                )
                            
                            mes_inflow_emp = st.number_input(
                                "Mês Inflow", 
                                min_value=0, 
                                max_value=59, 
                                step=1, 
                                key="emp_mes", 
                                help="Mês em que o empréstimo entra no fluxo de caixa (0 a 59)."
                            )
                        
                        with col2:
                            taxa_juros_emp = st.number_input(
                                "Taxa de Juros (% a.m.)", 
                                min_value=0.0, 
                                max_value=100.0, 
                                value=1.0, 
                                step=0.1, 
                                key="emp_taxa", 
                                help="Taxa de juros mensal do empréstimo."
                            )
                            
                            if not termo_caixa:
                                regime_amortizacao = st.selectbox(
                                    "Regime de Amortização",
                                    options=["Sistema de Amortização Constante (SAC)", 
                                             "Sistema Price (tabela Price)", 
                                             "Sistema de Amortização Misto (SAM)"],
                                    index=0,
                                    key="emp_regime",
                                    help="No SAC, a prestação é calculada dividindo o valor do financiamento pelo número de prestações, e os juros são calculados sobre o saldo devedor decrescente. "
                                         "Já no Price, as prestações são fixas e calculadas com a fórmula P(Price)=valordoemprestimo∗(1+i)**n∗i/(1+i)**n−i, onde i é a taxa de juros e n é o número de parcelas. "
                                         "O Sistema de Amortização Misto (SAM) é uma combinação dos sistemas de amortização SAC e Price, onde cada prestação é a média aritmética entre os valores das prestações desses dois sistemas."
                                )
                            
                            tipo_carencia = st.selectbox(
                                "Possui Carência?",
                                options=["Sem Carência", "Carência com Juros", "Carência Total"],
                                index=0,
                                key="emp_tipo_carencia",
                                help="Sem Carência: Pagamento começa no mês seguinte ao Inflow. "
                                     "Carência com Juros: Durante o período de carência, paga-se apenas os juros, sem amortização do principal. "
                                     "Carência Total: Nenhum pagamento (nem juros nem principal) durante o período de carência."
                            )
                            
                            periodo_carencia = st.number_input(
                                "Período de Carência (meses)", 
                                min_value=0, 
                                max_value=36, 
                                value=0, 
                                step=1, 
                                key="emp_carencia", 
                                help="Número de meses de carência, caso exista."
                            )
                            
                        submitted = st.form_submit_button("Adicionar Empréstimo")
                        
                        if submitted and nome_emp and valor_emp > 0 and (termo_caixa or prazo_emp > 0):
                            # Convert regime_amortizacao to a simpler code for storage
                            regime_code = "SAC"  # Default
                            if not termo_caixa:
                                if regime_amortizacao == "Sistema Price (tabela Price)":
                                    regime_code = "PRICE"
                                elif regime_amortizacao == "Sistema de Amortização Misto (SAM)":
                                    regime_code = "SAM"
                                    
                            # Determine carência settings based on selection
                            carencia = tipo_carencia != "Sem Carência"
                            carencia_total = tipo_carencia == "Carência Total"
                            
                            # Create empréstimo dictionary
                            novo_emprestimo = {
                                'nome': nome_emp,
                                'valor': valor_emp,
                                'taxa_juros': taxa_juros_emp,
                                'prazo': prazo_emp if not termo_caixa else 1,  # For Termo Caixa, use 1 as prazo
                                'mes_inflow': mes_inflow_emp,
                                'carencia': carencia,
                                'carencia_total': carencia_total,
                                'periodo_carencia': periodo_carencia,
                                'regime_amortizacao': regime_code,
                                'termo_caixa': termo_caixa  # Add this new field
                            }
                            
                            st.session_state['premissas_projecoes']['emprestimos'].append(novo_emprestimo)
                            st.success(f"Empréstimo '{nome_emp}' adicionado com sucesso!")
                            st.rerun()

        with tab2:
            st.write("### Configurações para Projeção de DRE")
            
            # Add the DRE inputs
            col1, col2 = st.columns(2)
            with col1:
                st.session_state['premissas_projecoes']['custos_cobranca'] = st.number_input(
                    "Custos de cobrança/mês (R$)",
                    min_value=0.0,
                    value=st.session_state['premissas_projecoes'].get('custos_cobranca', 0.0),
                    step=100.0,
                    format="%.2f",
                    help="Custos mensais relacionados à cobrança de clientes (ex.: taxas de cartão, tarifas bancárias)."
                )
            
            with col2:
                st.session_state['premissas_projecoes']['custos_juridicos'] = st.number_input(
                    "Custos Jurídicos e arbitrais/mês (R$)",
                    min_value=0.0,
                    value=st.session_state['premissas_projecoes'].get('custos_juridicos', 0.0),
                    step=100.0,
                    format="%.2f",
                    help="Custos mensais relacionados a serviços jurídicos e arbitrais."
                )
        
            # Existing information message
            st.info("Defina os parâmetros para a projeção de DRE.")
            
class ProjeçãodeFluxodeCaixa(Page):
    def __init__(self):
        self.graph_types = ["Gráfico de Linhas", "Gráfico de Pizza", "Gráfico de Barras"]
        self.time_frames = ["Mensal", "Anual"]
    
    @property
    def title(self) -> str:
        return "Projeção de Fluxo de Caixa"
    
    @property
    def icon(self) -> str:
        return "💰"
    
    def _calcular_emprestimos(self):
        """Calcula os fluxos de empréstimos para 60 meses"""
        if 'premissas_projecoes' not in st.session_state:
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        meses = range(60)
        
        # Cria DataFrames para inflows, pagamentos de juros e amortizações
        inflows = pd.DataFrame(0.0, index=pd.Index(['Empréstimos']), columns=meses)
        juros = pd.DataFrame(0.0, index=pd.Index(['Taxas/Juros de Empréstimos']), columns=meses)
        amortizacoes = pd.DataFrame(0.0, index=pd.Index(['Amortização de empréstimos']), columns=meses)
        
        # Processa cada empréstimo
        for emprestimo in st.session_state['premissas_projecoes']['emprestimos']:
            mes_inflow = emprestimo['mes_inflow']
            valor = emprestimo['valor']
            taxa_mensal = emprestimo['taxa_juros'] / 100
            prazo = emprestimo['prazo']
            carencia = emprestimo['periodo_carencia'] if emprestimo['carencia'] else 0
            carencia_total = emprestimo.get('carencia_total', False)
            regime_amortizacao = emprestimo.get('regime_amortizacao', 'SAC')
            termo_caixa = emprestimo.get('termo_caixa', False)  # Get the new field
            
            # Adiciona o inflow do empréstimo
            if mes_inflow < 60:
                inflows.loc['Empréstimos', mes_inflow] += valor
            
            # Se estiver fora do período de projeção, pula para o próximo empréstimo
            if mes_inflow + 1 >= 60:
                continue
                
            # Special handling for "Termo Caixa"
            if termo_caixa:
                # Calculate the payment month (inflow month + carência period)
                mes_pagamento = mes_inflow + carencia
                
                # Check if payment month is within projection period
                if mes_pagamento < 60:
                    # Calculate interest over the carência period
                    montante = valor * (1 + taxa_mensal) ** carencia
                    juros_total = montante - valor
                    
                    # Record the payments
                    amortizacoes.loc['Amortização de empréstimos', mes_pagamento] -= valor
                    juros.loc['Taxas/Juros de Empréstimos', mes_pagamento] -= juros_total
                
                # Skip the rest of the processing for this loan
                continue
            
            # Regular loan processing (existing code)
            # Calcula valores de acordo com o regime de amortização
            saldo_devedor = valor
            prazo_efetivo = prazo - carencia if prazo > carencia else 0
            
            # Pré-calcular valores de amortização e parcelas para os diferentes sistemas
            # SAC - Amortização constante
            if prazo_efetivo > 0:
                amortizacao_sac = valor / prazo_efetivo
            else:
                amortizacao_sac = 0
                
            # PRICE - Prestação fixa
            if taxa_mensal > 0 and prazo_efetivo > 0:
                parcela_price = valor * (taxa_mensal * (1 + taxa_mensal) ** prazo_efetivo) / (((1 + taxa_mensal) ** prazo_efetivo) - 1)
            else:
                parcela_price = valor / prazo_efetivo if prazo_efetivo > 0 else 0
                
            # Aplica os pagamentos
            for mes in range(mes_inflow + 1, 60):
                # Índice para controle do mês relativo do empréstimo (0-indexed)
                mes_rel = mes - mes_inflow - 1
                
                # Durante carência
                if mes_rel < carencia:
                    # Cálculo de juros durante carência
                    juro = saldo_devedor * taxa_mensal
                    
                    if not carencia_total:  # Se não for carência total, paga juros
                        juros.loc['Taxas/Juros de Empréstimos', mes] -= juro
                    # Na carência total, não paga nem juros nem amortização
                    continue
                
                # Após carência, até o fim do prazo
                elif mes_rel < carencia + prazo_efetivo:
                    # Cálculo de juros em qualquer regime (sempre sobre o saldo devedor)
                    juro = saldo_devedor * taxa_mensal
                    
                    # Cálculo da amortização de acordo com o regime
                    if regime_amortizacao == "SAC":
                        # SAC: amortização constante
                        amortizacao = amortizacao_sac
                        parcela = amortizacao + juro
                    
                    elif regime_amortizacao == "PRICE":
                        # PRICE: parcela fixa
                        parcela = parcela_price
                        amortizacao = parcela - juro
                    
                    else:  # SAM - média aritmética entre SAC e PRICE
                        # Calcular parcela SAC para o mês atual
                        parcela_sac_atual = amortizacao_sac + juro
                        
                        # A parcela SAM é a média entre a parcela SAC e a parcela Price
                        parcela = (parcela_sac_atual + parcela_price) / 2
                        amortizacao = parcela - juro
                    
                    # Registrar valores nas tabelas de fluxo
                    current_juros = pd.to_numeric(juros.loc['Taxas/Juros de Empréstimos', mes], errors='coerce') or 0
                    current_amortizacao = pd.to_numeric(amortizacoes.loc['Amortização de empréstimos', mes], errors='coerce') or 0
                    
                    juros.loc['Taxas/Juros de Empréstimos', mes] = current_juros - juro
                    amortizacoes.loc['Amortização de empréstimos', mes] = current_amortizacao - amortizacao
                    
                    # Atualizar saldo devedor
                    saldo_devedor -= amortizacao
                    
                    # Verificar se o saldo devedor zerou (com margem de erro)
                    if abs(saldo_devedor) < 0.01:
                        saldo_devedor = 0
                        
                    if saldo_devedor <= 0:
                        break  # Empréstimo quitado
        
        return inflows, juros, amortizacoes
    
    def _calcular_outras_receitas(self):
        """Calcula os fluxos de outras receitas para 60 meses"""
        meses = range(60)
        outras_receitas = pd.DataFrame(0.0, index=pd.Index(['Outras Receitas']), columns=meses)
        
        # Get data from Receitas class
        try:
            receitas = Receitas()
            df_faturamento = receitas._gerar_dataframe_faturamento()
            
            if df_faturamento is not None and "Outras Receitas" in df_faturamento.index:
                # Copy the "Outras Receitas" row directly from the faturamento dataframe
                outras_receitas.loc['Outras Receitas'] = df_faturamento.loc["Outras Receitas"]
        except Exception as e:
            st.warning(f"Erro ao obter dados de outras receitas: {e}")
        
        return outras_receitas
    
    def _calcular_investimentos_socios(self):
        """Calcula os fluxos de investimentos dos sócios para 60 meses"""
        if 'premissas_investimentos' not in st.session_state:
            return pd.DataFrame()
        
        meses = range(60)
        investimentos = pd.DataFrame(0.0, index=pd.Index(['Capital Próprio Investido na Empresa']), columns=meses)
       
        for inv in st.session_state['premissas_investimentos']['investimentos_socios']:
            valor = inv['valor']
            mes_inicial = inv['mes_inflow']
            periodicidade = inv['periodicidade'] if inv['periodicidade_ativa'] else 0
            
            if periodicidade > 0:
                # Investimento periódico
                for mes in range(mes_inicial, 60, periodicidade):
                    investimentos.loc['Capital Próprio Investido na Empresa', mes] += valor
            else:
                # Investimento único
                if mes_inicial < 60:
                    investimentos.loc['Capital Próprio Investido na Empresa', mes_inicial] += valor
        
        return investimentos
    
    def _calcular_investimentos_futuros(self):
        """Calcula os fluxos de investimentos futuros para 60 meses"""
        if 'premissas_investimentos' not in st.session_state:
            return pd.DataFrame()
        
        meses = range(60)
        investimentos = pd.DataFrame(0.0, index=pd.Index(['Ampliações & Melhorias - Investimentos Futuros']), columns=meses)
        
        for inv in st.session_state['premissas_investimentos']['investimentos_futuros']:
            valor = -inv['valor']  # Valor negativo pois é uma saída
            mes_inicial = inv['mes_outflow']
            periodicidade = inv['periodicidade'] if inv['periodicidade_ativa'] else 0
            
            if periodicidade > 0:
                # Investimento periódico
                for mes in range(mes_inicial, 60, periodicidade):
                    investimentos.loc['Ampliações & Melhorias - Investimentos Futuros', mes] += valor
            else:
                # Investimento único
                if mes_inicial < 60:
                    mes = mes_inicial  # Define mes with mes_inicial value
                    investimentos.loc['Ampliações & Melhorias - Investimentos Futuros', mes] += valor
       
        return investimentos
    
    def _gerar_dataframe_entradas(self):
        """Gera o dataframe de entradas (inflows) para 60 meses"""
        meses = range(60)
        
        # Criar DataFrame base
        df = pd.DataFrame(0.0, index=pd.Index(["Receita de Vendas / Serviços", "Empréstimos", "Outras Receitas", 
                                      "Capital Próprio Investido na Empresa", "Total Entradas"]), columns=meses)
        
        # Receita de Vendas / Serviços (da classe Receitas)
        try:
            receitas = Receitas()
            df_faturamento = receitas._gerar_dataframe_faturamento()
            df_arrecadacao = receitas._gerar_dataframe_arrecadacao(df_faturamento)
            
            if df_arrecadacao is not None and "Arrecadação Líquida" in df_arrecadacao.index:
                for mes in meses:
                    df.loc["Receita de Vendas / Serviços", mes] = df_arrecadacao.loc["Arrecadação Líquida", mes]
        except Exception as e:
            st.warning(f"Erro ao obter dados de arrecadação: {e}")
        
        # Empréstimos
        df_emprestimos, _, _ = self._calcular_emprestimos()
        if not df_emprestimos.empty:
            df.loc["Empréstimos"] = df_emprestimos.loc["Empréstimos"]
        
        # Outras Receitas
        df_outras_receitas = self._calcular_outras_receitas()
        if not df_outras_receitas.empty:
            df.loc["Outras Receitas"] = df_outras_receitas.loc["Outras Receitas"]
        
        # Capital Próprio Investido na Empresa
        df_investimentos = self._calcular_investimentos_socios()
        if not df_investimentos.empty:
            df.loc["Capital Próprio Investido na Empresa"] = df_investimentos.loc["Capital Próprio Investido na Empresa"]
        
          # Calcular total
        for mes in meses:
            # Convert each value to numeric using pandas to_numeric function with errors='coerce'
            # This will convert non-numeric values to NaN and then replace NaN with 0
            receitas = pd.to_numeric(df.loc["Receita de Vendas / Serviços", mes], errors='coerce') or 0
            emprestimos = pd.to_numeric(df.loc["Empréstimos", mes], errors='coerce') or 0
            outras_receitas = pd.to_numeric(df.loc["Outras Receitas", mes], errors='coerce') or 0
            capital_proprio = pd.to_numeric(df.loc["Capital Próprio Investido na Empresa", mes], errors='coerce') or 0
            
            # Sum the numeric values
            df.loc["Total Entradas", mes] = receitas + emprestimos + outras_receitas + capital_proprio
      
        return df
    
    def _gerar_dataframe_saidas(self):
        """Gera o dataframe de saídas (outflows) para 60 meses"""
        meses = range(60)
        
        # Criar índices multinível
        idx = pd.MultiIndex.from_tuples([
            ("Investimentos Fixos", "Implantação - Investimento Inicial"),
            ("Investimentos Fixos", "Ampliações & Melhorias - Investimentos Futuros"),
            ("Investimentos Fixos", "Aquisição de Equipamentos"), 
            ("Despesas Administrativas", "Total Despesas Administrativas"),
            ("Remuneração da Equipe", "Total Remuneração da Equipe"),
            ("Custos diretos", "Despesas de Produção"),
            ("Custos diretos", "Comissões e taxas sobre vendas"),
            ("Despesas Tributárias", "Impostos a Pagar"),
            ("Despesas Financeiras", "Taxas/Juros de Empréstimos"),
            ("Despesas Financeiras", "Amortização de empréstimos"),
            ("Dividendos (Distribuição de lucros)", "Total Dividendos"),
            ("TOTAL", "Total Saídas")
        ])
        
        # Criar DataFrame base
        df = pd.DataFrame(0.0, index=idx, columns=meses)
        
         # 1. Investimentos Fixos: Implantação - Investimento Inicial
        try:
            investimentos = Investimentos()
            df_gastos = investimentos.get_total_investimento()
            if df_gastos is not None:
                if isinstance(df_gastos, pd.DataFrame) and 'Total' in df_gastos.columns:
                    total_implantacao = -df_gastos['Total'].sum()  # Valor negativo pois é saída
                elif isinstance(df_gastos, (int, float)):
                    total_implantacao = -df_gastos  # Valor negativo pois é saída
                else:
                    total_implantacao = 0
                df.loc[("Investimentos Fixos", "Implantação - Investimento Inicial"), 0] = total_implantacao
        except Exception as e:
            st.warning(f"Erro ao obter dados de investimento inicial: {e}")
       
        # 2. Investimentos Fixos: Ampliações & Melhorias - Investimentos Futuros
        df_investimentos_futuros = self._calcular_investimentos_futuros()
        if not df_investimentos_futuros.empty:
            df.loc[("Investimentos Fixos", "Ampliações & Melhorias - Investimentos Futuros")] = df_investimentos_futuros.loc["Ampliações & Melhorias - Investimentos Futuros"]
        
        # 3. Investimentos Fixos: Aquisição de Equipamentos
        try:
            # Calculate depreciation expenses
            # 3. Investimentos Fixos: Aquisição de Equipamentos
            if 'premissas_despesas' in st.session_state and 'equipamentos' in st.session_state['premissas_despesas']:
                for equip in st.session_state['premissas_despesas']['equipamentos']:
                    mes_aquisicao = equip.get('mes_aquisicao', 0)
                    if mes_aquisicao < 60:
                        valor_total = -equip['valor'] * equip['quantidade']  # Negative as it's an outflow
                        df.loc[("Investimentos Fixos", "Aquisição de Equipamentos"), mes_aquisicao] += valor_total
                    
        except Exception as e:
            st.warning(f"Erro ao obter dados de aquisição de equipamentos: {e}")
        
        # 3.1 Depreciation outflows - should only appear once per year
        try:
            # Get depreciation data from CustosTecnologia
            tecnologia = CustosTecnologia()
            df_tecnologia = tecnologia._gerar_dataframe_custos_tecnologia()
            
            if df_tecnologia is not None and "Depreciação de Equipamentos" in df_tecnologia.index:
                # Create the index if it doesn't exist
                if ("Investimentos Fixos", "Depreciação de Equipamentos") not in df.index:
                    df.loc[("Investimentos Fixos", "Depreciação de Equipamentos")] = 0
                
                # Copy the values for each month from the CustosTecnologia DataFrame
                for mes in meses:
                    if mes < len(df_tecnologia.columns):
                        # Get the depreciation value (already negative in CustosTecnologia) and maintain the sign
                        valor_depreciacao = pd.to_numeric(df_tecnologia.loc["Depreciação de Equipamentos", mes], errors='coerce') or 0
                        df.loc[("Investimentos Fixos", "Depreciação de Equipamentos"), mes] = valor_depreciacao
        except Exception as e:
            st.warning(f"Erro ao calcular outflows de depreciação de equipamentos: {str(e)}")
        
        # 3. Despesas Administrativas
        try:
            despesas_adm = DespesasAdm()
            df_despesas = despesas_adm._gerar_dataframe_despesas()
            if not df_despesas.empty:
                for mes in meses:
                    if mes < len(df_despesas):
                        # Convert to numeric value first, then negate
                        valor = pd.to_numeric(df_despesas.loc[mes, "Total"], errors='coerce') or 0
                        df.loc[("Despesas Administrativas", "Total Despesas Administrativas"), mes] = -valor
        except Exception as e:
            st.warning(f"Erro ao obter dados de despesas administrativas: {e}")
      
        # 4. Remuneração da Equipe
        try:
            # Get data from Receitas._calcular_equipe_comercial
            receitas = Receitas()
            df_faturamento = receitas._gerar_dataframe_faturamento()
            df_equipe_comercial = receitas._calcular_equipe_comercial(df_faturamento)
            
            # Get data from ComissaoVendas
            comissao_vendas = ComissaoVendas()
            df_comissoes = comissao_vendas._calcular_comissoes()
            
            # Calculate total team compensation
            for mes in meses:
                total_remuneracao = 0
                
                # Add SDR and Closer salaries
                if not df_equipe_comercial.empty and mes < len(df_equipe_comercial.columns):
                    salario_sdr = pd.to_numeric(df_equipe_comercial.loc["Salário SDR", mes], errors='coerce') or 0
                    salario_closer = pd.to_numeric(df_equipe_comercial.loc["Salário Closer", mes], errors='coerce') or 0
                    total_remuneracao += salario_sdr + salario_closer
                
                # Add commissions
                if df_comissoes is not None and not df_comissoes.empty and mes < len(df_comissoes) and "Total" in df_comissoes.columns:
                    comissoes_total = pd.to_numeric(df_comissoes.loc[mes, "Total"], errors='coerce') or 0
                    total_remuneracao += comissoes_total
                
                # Add other employee salaries from PremissasDespesas
                if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
                    for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                        if 'nome' in cargo and cargo['nome'] not in ['SDR', 'Closer']:  # Exclude SDR and Closer
                            salario = cargo.get('salario', 0)
                            quantidade = cargo.get('quantidade', 0)
                            total_remuneracao += salario * quantidade
                
                # Make value negative since it's an outflow
                df.loc[("Remuneração da Equipe", "Total Remuneração da Equipe"), mes] = -total_remuneracao
        except Exception as e:
            st.warning(f"Erro ao calcular remuneração da equipe: {e}")
        
        # 5. Custos diretos: Despesas de Produção
        try:
            tecnologia = CustosTecnologia()
            df_tecnologia = tecnologia._gerar_dataframe_custos_tecnologia()
            if not df_tecnologia.empty:
                for mes in meses:
                     if mes < len(df_tecnologia.columns):
                        # Convert to numeric first, then negate
                        valor = pd.to_numeric(df_tecnologia.loc["Total", mes], errors='coerce') or 0
                        df.loc[("Custos diretos", "Despesas de Produção"), mes] = -valor
        except Exception as e:
            st.warning(f"Erro ao obter dados de custos de tecnologia: {e}")
        
        # 6. Custos diretos: Comissões e taxas sobre vendas
        try:
            comissao_vendas = ComissaoVendas()
            df_comissoes = comissao_vendas._calcular_comissoes()
            
            if df_comissoes is not None and not df_comissoes.empty:
                for mes in meses:
                    if mes < len(df_comissoes) and "Total" in df_comissoes.columns:
                        # Get the Total column value directly - it already includes all commissions
                        comissoes_total = pd.to_numeric(df_comissoes.loc[mes, "Total"], errors='coerce') or 0
                        # Make value negative since it's an outflow
                        df.loc[("Custos diretos", "Comissões e taxas sobre vendas"), mes] = -comissoes_total
        except Exception as e:
            st.warning(f"Erro ao obter dados de comissões: {e}")
            # Initialize with zeros in case of error
            for mes in meses:
                df.loc[("Custos diretos", "Comissões e taxas sobre vendas"), mes] = 0
        
        # Despesas Tributárias: Impostos a Pagar
        try:
            tributos = Tributos()
            df_tributos = tributos._gerar_dataframe_tributos()
            
            if df_tributos is not None and "Tributo Total" in df_tributos.index:
                # Add this code to copy values from df_tributos to the outflows dataframe
                for mes in meses:
                    df.loc[("Despesas Tributárias", "Impostos a Pagar"), mes] = -df_tributos.loc["Tributo Total", mes]
            
        except Exception as e:
            st.warning(f"Erro ao obter dados de tributos: {e}")
       
        # 8 e 9. Despesas Financeiras
        _, df_juros, df_amortizacoes = self._calcular_emprestimos()
        if not df_juros.empty:
            df.loc[("Despesas Financeiras", "Taxas/Juros de Empréstimos")] = df_juros.loc["Taxas/Juros de Empréstimos"]
        if not df_amortizacoes.empty:
            df.loc[("Despesas Financeiras", "Amortização de empréstimos")] = df_amortizacoes.loc["Amortização de empréstimos"]
        
         # 10. Dividendos (Distribuição de lucros)
        # Initialize df_equipe before usage to avoid "possibly unbound" error
        df_equipe = pd.DataFrame()
        try:
            equipe = Equipe()
            df_equipe = equipe._gerar_dataframe_custos_equipe()
            if not df_equipe.empty:
                dividendos_total = 0
                for mes in meses:
                     if mes < len(df_equipe.columns):
                        if ("Bônus dos Lucros", "Bônus CEO") in df_equipe.index:
                            bonus_ceo = pd.to_numeric(df_equipe.loc[("Bônus dos Lucros", "Bônus CEO"), mes+1], errors='coerce') or 0
                            dividendos_total += -bonus_ceo
                        if ("Bônus dos Lucros", "Bônus CFO") in df_equipe.index:
                                bonus_cfo = pd.to_numeric(df_equipe.loc[("Bônus dos Lucros", "Bônus CFO"), mes+1], errors='coerce') or 0
                                dividendos_total += -bonus_cfo
                        if ("Bônus dos Lucros", "Bônus Head de Vendas") in df_equipe.index:
                            bonus_head = pd.to_numeric(df_equipe.loc[("Bônus dos Lucros", "Bônus Head de Vendas"), mes+1], errors='coerce') or 0
                            dividendos_total += -bonus_head
                       
                            df.loc[("Dividendos (Distribuição de lucros)", "Total Dividendos"), mes] = dividendos_total
        except Exception as e:
            st.warning(f"Erro ao obter dados de dividendos: {e}")
        
         # Calcular total
        for mes in meses:
            # Convert each value to float using pd.to_numeric with errors='coerce' (converts invalid values to NaN)
            # Then replace NaN with 0 using the or 0 pattern
            investimentos_impl = pd.to_numeric(df.loc[("Investimentos Fixos", "Implantação - Investimento Inicial"), mes], errors='coerce') or 0
            investimentos_futuros = pd.to_numeric(df.loc[("Investimentos Fixos", "Ampliações & Melhorias - Investimentos Futuros"), mes], errors='coerce') or 0
            equipamentos = pd.to_numeric(df.loc[("Investimentos Fixos", "Aquisição de Equipamentos"), mes], errors='coerce') or 0  # Add this line
            despesas_adm = pd.to_numeric(df.loc[("Despesas Administrativas", "Total Despesas Administrativas"), mes], errors='coerce') or 0
            remuneracao = pd.to_numeric(df.loc[("Remuneração da Equipe", "Total Remuneração da Equipe"), mes], errors='coerce') or 0
            despesas_producao = pd.to_numeric(df.loc[("Custos diretos", "Despesas de Produção"), mes], errors='coerce') or 0
            comissoes = pd.to_numeric(df.loc[("Custos diretos", "Comissões e taxas sobre vendas"), mes], errors='coerce') or 0
            impostos = pd.to_numeric(df.loc[("Despesas Tributárias", "Impostos a Pagar"), mes], errors='coerce') or 0
            juros = pd.to_numeric(df.loc[("Despesas Financeiras", "Taxas/Juros de Empréstimos"), mes], errors='coerce') or 0
            amortizacao = pd.to_numeric(df.loc[("Despesas Financeiras", "Amortização de empréstimos"), mes], errors='coerce') or 0
            dividendos = pd.to_numeric(df.loc[("Dividendos (Distribuição de lucros)", "Total Dividendos"), mes], errors='coerce') or 0
            
            # Sum the numeric values
             # Sum values individually, ensuring each is treated as a numeric value
            total_saidas = 0.0
            for value in [investimentos_impl, investimentos_futuros, equipamentos, 
                         despesas_adm, remuneracao, despesas_producao, comissoes,
                         impostos, juros, amortizacao, dividendos]:
                # Convert to float and handle any conversion errors
                try:
                    if isinstance(value, (pd.Series, pd.DataFrame)):
                        value_float = float(value.iloc[0])
                    else:
                        value_float = float(value)
                    total_saidas += value_float
                except (ValueError, TypeError):
                    # If conversion fails, try to get a numeric representation or use 0
                    try:
                        total_saidas += float(pd.to_numeric(value, errors='coerce') or 0)
                    except:
                        pass  # If all else fails, don't add this value
            
            df.loc[("TOTAL", "Total Saídas"), mes] = total_saidas
      
        return df
    
    def _gerar_dataframe_anual(self, df_mensal):
        """Converte o dataframe mensal para anual"""
        if df_mensal.empty:
            return pd.DataFrame()
            
        anos = 5
        df_anual = pd.DataFrame(index=df_mensal.index, columns=range(1, anos + 1))
        
        for ano in range(1, anos + 1):
            mes_inicio = (ano - 1) * 12
            mes_fim = ano * 12 - 1
            
            # Para cada categoria, somar os valores dos meses do ano
            for categoria in df_mensal.index:
                df_anual.loc[categoria, ano] = df_mensal.loc[categoria, mes_inicio:mes_fim].sum()
        
        return df_anual
    
    def _calcular_fluxo_periodo(self, df_entradas, df_saidas):
        """Calcula o fluxo do período como Entradas - Saídas"""
        meses = range(60)
        df_fluxo = pd.DataFrame(0.0, index=pd.Index(["Fluxo do Período (E - S)"]), columns=meses)
        
        for mes in meses:
            df_fluxo.loc["Fluxo do Período (E - S)", mes] = (
                df_entradas.loc["Total Entradas", mes] + 
                df_saidas.loc[("TOTAL", "Total Saídas"), mes]  # Saídas já são negativas
            )
        
        return df_fluxo
    
    def _calcular_saldo_acumulado(self, df_fluxo):
        """Calcula o saldo acumulado ao longo do tempo"""
        meses = range(60)
        df_saldo = pd.DataFrame(0.0, index=pd.Index(["Saldo Acumulado"]), columns=meses)
        
        saldo_acumulado = 0
        for mes in meses:
            saldo_acumulado += df_fluxo.loc["Fluxo do Período (E - S)", mes]
            df_saldo.loc["Saldo Acumulado", mes] = saldo_acumulado
        
        return df_saldo
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        if 'premissas_projecoes' not in st.session_state:
            st.error("Premissas de projeções não definidas. Por favor, defina as premissas na página 'Premissas Projeções'.")
            return
        
        # Criar as abas para diferentes visualizações
        tab1, tab2, tab3 = st.tabs(["Entradas", "Saídas", "Resumo"])
        
        # Gerar dataframes
        df_entradas = self._gerar_dataframe_entradas()
        df_saidas = self._gerar_dataframe_saidas()
        df_fluxo = self._calcular_fluxo_periodo(df_entradas, df_saidas)
        df_saldo = self._calcular_saldo_acumulado(df_fluxo)
        
        # Combinar df_fluxo e df_saldo
        df_resumo = pd.concat([df_fluxo, df_saldo])
        
        # Tab Entradas
        with tab1:
            st.write("### Entradas de Caixa")
            
            # Opções de visualização
            col1, col2 = st.columns(2)
            
            with col1:
                time_frame_entradas = st.selectbox(
                    "Período", 
                    self.time_frames, 
                    index=0,
                    key="entradas_time_frame"
                )
            
            with col2:
                plot_type_entradas = st.selectbox(
                    "Tipo de Gráfico", 
                    self.graph_types, 
                    index=0,
                    key="entradas_plot_type"
                )
            
            # Converter para anual se selecionado
            df_display_entradas = df_entradas.copy()
            if time_frame_entradas == "Anual":
                df_display_entradas = self._gerar_dataframe_anual(df_entradas)
                df_display_entradas.columns = [f"Ano {ano}" for ano in df_display_entradas.columns]
            else:
                df_display_entradas.columns = [f"Mês {mes}" for mes in df_display_entradas.columns]
            
            # Mostrar o dataframe
            periodo_entradas = "Anual" if time_frame_entradas == "Anual" else "Mensal"
            st.write(f"#### Entradas de Caixa - Visão {periodo_entradas}")
            st.dataframe(df_display_entradas.style.format("{:.2f}"), use_container_width=True)
            
            # Seleção de categorias para visualização
            categorias_entradas = df_display_entradas.index.tolist()  # Excluindo o Total
            
            # Seleção múltipla de categorias
            selected_categories_entradas = st.multiselect(
                "Selecione as categorias para visualizar", 
                categorias_entradas,
                default=categorias_entradas[:2] if len(categorias_entradas) >= 2 else categorias_entradas,
                key="entradas_categories"
            )
            
            # Se nenhuma categoria foi selecionada, mostrar as duas primeiras ou todas se menos de 2
            if not selected_categories_entradas:
                selected_categories_entradas = categorias_entradas[:2] if len(categorias_entradas) >= 2 else categorias_entradas
            
            # Preparar dados para o gráfico
            if selected_categories_entradas:
                df_plot_entradas = df_display_entradas.loc[selected_categories_entradas].T  # Transpor para ter períodos como índice
                
                if plot_type_entradas == "Gráfico de Pizza":
                    # Existing pizza chart code
                    data_pie_entradas = {
                        'Categoria': selected_categories_entradas,
                        'Valor': [df_display_entradas.loc[cat].sum() for cat in selected_categories_entradas]
                    }
                    df_pie_entradas = pd.DataFrame(data_pie_entradas)
                    
                    fig_entradas = px.pie(
                        df_pie_entradas,
                        values='Valor',
                        names='Categoria',
                        title=f"Distribuição de Entradas - Total ({periodo_entradas})"
                    )
                    st.plotly_chart(fig_entradas, use_container_width=True)
                elif plot_type_entradas == "Gráfico de Barras":
                    # New bar chart code
                    fig_entradas = px.bar(
                        df_plot_entradas,
                        title=f"Fluxo de Entradas ({periodo_entradas})",
                        barmode='group'
                    )
                    fig_entradas.update_layout(
                        xaxis_title=f"{'Ano' if time_frame_entradas == 'Anual' else 'Mês'}",
                        yaxis_title="Valor (R$)",
                        legend_title="Categoria"
                    )
                    st.plotly_chart(fig_entradas, use_container_width=True)
                else:  # Default to line chart
                    # Existing line chart code
                    fig_entradas = px.line(
                        df_plot_entradas,
                        title=f"Fluxo de Entradas ({periodo_entradas})",
                        markers=True
                    )
                    fig_entradas.update_layout(
                        xaxis_title=f"{'Ano' if time_frame_entradas == 'Anual' else 'Mês'}",
                        yaxis_title="Valor (R$)",
                        legend_title="Categoria"
                    )
                    st.plotly_chart(fig_entradas, use_container_width=True)
        
        # Tab Saídas
        with tab2:
            st.write("### Saídas de Caixa")
            
            # Opções de visualização
            col1, col2 = st.columns(2)
            
            with col1:
                time_frame_saidas = st.selectbox(
                    "Período", 
                    self.time_frames, 
                    index=0,
                    key="saidas_time_frame"
                )
            
            with col2:
                plot_type_saidas = st.selectbox(
                    "Tipo de Gráfico", 
                    self.graph_types, 
                    index=0,
                    key="saidas_plot_type"
                )
            
            # Converter para anual se selecionado
            df_display_saidas = df_saidas.copy()
            if time_frame_saidas == "Anual":
                df_display_saidas = self._gerar_dataframe_anual(df_saidas)
                df_display_saidas.columns = [f"Ano {ano}" for ano in df_display_saidas.columns]
            else:
                df_display_saidas.columns = [f"Mês {mes}" for mes in df_display_saidas.columns]
            
            # Mostrar o dataframe
            periodo_saidas = "Anual" if time_frame_saidas == "Anual" else "Mensal"
            st.write(f"#### Saídas de Caixa - Visão {periodo_saidas}")
            st.dataframe(df_display_saidas.style.format("{:.2f}"), use_container_width=True)
            
            # Seleção de categorias para visualização
            # Pegar os índices de primeiro nível excluindo o TOTAL
            categorias_nivel1 = [idx[0] for idx in df_display_saidas.index.tolist()]
            categorias_nivel1 = list(dict.fromkeys(categorias_nivel1))  # Remover duplicados
            
            # Seleção múltipla de categorias
            selected_categories_saidas = st.multiselect(
                "Selecione as categorias para visualizar", 
                categorias_nivel1,
                default=categorias_nivel1[:2] if len(categorias_nivel1) >= 2 else categorias_nivel1,
                key="saidas_categories"
            )
            
            # Se nenhuma categoria foi selecionada, mostrar as duas primeiras ou todas se menos de 2
            if not selected_categories_saidas:
                selected_categories_saidas = categorias_nivel1[:2] if len(categorias_nivel1) >= 2 else categorias_nivel1
            
            # Preparar dados para o gráfico - somando valores por categoria de primeiro nível
            if selected_categories_saidas:
                # Criar dataframe para o plot
                df_plot_saidas = pd.DataFrame()
                
                # Para cada categoria selecionada, somar os valores das subcategorias
                for categoria in selected_categories_saidas:
                    # Pegar todas as linhas dessa categoria
                    indices_categoria = [idx for idx in df_display_saidas.index.tolist() if idx[0] == categoria]
                    
                     # Somar valores para cada coluna (mês/ano)
                    for col in df_display_saidas.columns:
                        # Convert all values to numeric before summing
                        soma_categoria = sum(pd.to_numeric(df_display_saidas.loc[idx, col], errors='coerce') or 0 for idx in indices_categoria)
                        if categoria not in df_plot_saidas:
                            df_plot_saidas[categoria] = pd.Series(dtype='float64')
                        df_plot_saidas.at[col, categoria] = soma_categoria
               
                if plot_type_saidas == "Gráfico de Linhas":
                    # Gráfico de linha para evolução temporal
                    fig_saidas = px.line(
                        df_plot_saidas,
                        title=f"Evolução das Saídas de Caixa ({periodo_saidas})",
                        markers=True
                    )
                    fig_saidas.update_layout(
                        xaxis_title=f"{'Ano' if time_frame_saidas == 'Anual' else 'Mês'}",
                        yaxis_title="Valor (R$)",
                        legend_title="Categoria"
                    )
                    st.plotly_chart(fig_saidas, use_container_width=True)
                elif plot_type_saidas == "Gráfico de Barras":
                    # Gráfico de barras para comparação de categorias
                    fig_saidas = px.bar(
                        df_plot_saidas,
                        title=f"Fluxo de Saídas de Caixa ({periodo_saidas})",
                        barmode='group'
                    )
                    fig_saidas.update_layout(
                        xaxis_title=f"{'Ano' if time_frame_saidas == 'Anual' else 'Mês'}",
                        yaxis_title="Valor (R$)",
                        legend_title="Categoria"
                    )
                    st.plotly_chart(fig_saidas, use_container_width=True)
                else:
                    # Gráfico de pizza para composição
                    # Calcular a soma total de cada categoria para o gráfico de pizza
                    data_pie_saidas = {
                        'Categoria': selected_categories_saidas,
                        'Valor': [df_plot_saidas[cat].sum() for cat in selected_categories_saidas]
                    }
                    df_pie_saidas = pd.DataFrame(data_pie_saidas)
                    
                    # Usar valores absolutos para o gráfico de pizza
                    df_pie_saidas['Valor'] = df_pie_saidas['Valor'].abs()
                    
                    fig_pie_saidas = px.pie(
                        df_pie_saidas,
                        values='Valor',
                        names='Categoria',
                        title=f"Composição das Saídas de Caixa - Total {periodo_saidas}"
                    )
                    st.plotly_chart(fig_pie_saidas, use_container_width=True)
        
        # Tab Resumo
        with tab3:
            st.write("### Resumo do Fluxo de Caixa")
            
            # Opções de visualização
            time_frame_resumo = st.selectbox(
                "Período", 
                self.time_frames, 
                index=0,
                key="resumo_time_frame"
            )
            
            # Converter para anual se selecionado
            df_display_resumo = df_resumo.copy()
            df_display_entradas_resumo = df_entradas.loc[["Total Entradas"]].copy()
            df_display_saidas_resumo = df_saidas.loc[[("TOTAL", "Total Saídas")]].copy()
            
            if time_frame_resumo == "Anual":
                df_display_resumo = self._gerar_dataframe_anual(df_resumo)
                df_display_entradas_resumo = self._gerar_dataframe_anual(df_entradas.loc[["Total Entradas"]])
                df_display_saidas_resumo = self._gerar_dataframe_anual(df_saidas.loc[[("TOTAL", "Total Saídas")]])
                
                df_display_resumo.columns = [f"Ano {ano}" for ano in df_display_resumo.columns]
                df_display_entradas_resumo.columns = [f"Ano {ano}" for ano in df_display_entradas_resumo.columns]
                df_display_saidas_resumo.columns = [f"Ano {ano}" for ano in df_display_saidas_resumo.columns]
            else:
                df_display_resumo.columns = [f"Mês {mes}" for mes in df_display_resumo.columns]
                df_display_entradas_resumo.columns = [f"Mês {mes}" for mes in df_display_entradas_resumo.columns]
                df_display_saidas_resumo.columns = [f"Mês {mes}" for mes in df_display_saidas_resumo.columns]
            
            # Mostrar o dataframe
            periodo_resumo = "Anual" if time_frame_resumo == "Anual" else "Mensal"
            st.write(f"#### Fluxo de Caixa - Visão {periodo_resumo}")
            st.dataframe(df_display_resumo.style.format("{:.2f}"), use_container_width=True)
            
            # Gráfico de barras para Entradas vs Saídas
            df_plot_resumo = pd.DataFrame()
            
             # Renomear índices para exibição usando pd.Index
            df_display_entradas_resumo.index = pd.Index(["Entradas"])
            df_display_saidas_resumo.index = pd.Index(["Saídas"])
           
            # Criar dataframe para plot
            df_plot_resumo = pd.concat([df_display_entradas_resumo, df_display_saidas_resumo.abs()])
            
            # Transpor para ter meses/anos como índice
            df_plot_resumo = df_plot_resumo.T
            
            # Gráfico de barras para Entradas vs Saídas
            fig_barras = px.bar(
                df_plot_resumo,
                y=["Entradas", "Saídas"],
                title=f"Entradas vs Saídas ({periodo_resumo})",
                barmode='group'
            )
            fig_barras.update_layout(
                xaxis_title=f"{'Ano' if time_frame_resumo == 'Anual' else 'Mês'}",
                yaxis_title="Valor (R$)",
                legend_title="Categoria"
            )
            st.plotly_chart(fig_barras, use_container_width=True)
            
            # Gráfico de linha para evolução do saldo acumulado
            fig_saldo = px.line(
                df_display_resumo.T,  # Transpor para ter meses/anos como índice
                y="Saldo Acumulado",
                title=f"Evolução do Saldo Acumulado ({periodo_resumo})",
                markers=True
            )
            fig_saldo.update_layout(
                xaxis_title=f"{'Ano' if time_frame_resumo == 'Anual' else 'Mês'}",
                yaxis_title="Valor (R$)"
            )
            st.plotly_chart(fig_saldo, use_container_width=True)
            
            # Métricas principais
            st.write("### Métricas Principais")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                total_entradas = df_entradas.loc["Total Entradas"]
                if isinstance(total_entradas, pd.DataFrame):
                    total_entradas = total_entradas.astype(float).sum().sum()
                elif isinstance(total_entradas, pd.Series):
                    total_entradas = pd.to_numeric(total_entradas, errors='coerce').sum()
                else:
                    total_entradas = float(total_entradas) if total_entradas else 0
                    st.metric(label="Total de Entradas (60 meses)", value=f"R$ {total_entradas:,.2f}")
            
            with col2:
                 # Handle different types of return values
                total_saidas_value = df_saidas.loc[("TOTAL", "Total Saídas")]
                
                # Check if it's a Series that has a sum method
                if hasattr(total_saidas_value, 'sum'):
                    total_saidas = abs(total_saidas_value.sum())
                else:
                    # If it's a scalar value or any other type, convert to numeric
                    total_saidas = abs(pd.to_numeric(total_saidas_value, errors='coerce') or 0)
                    st.metric(label="Total de Saídas (60 meses)", value=f"R$ {total_saidas:,.2f}")
           
            with col3:
                  # Convert to scalar if it's a Series
                if isinstance(df_saldo.loc["Saldo Acumulado"].iloc[-1], pd.Series):
                    saldo_final = df_saldo.loc["Saldo Acumulado"].iloc[-1].iloc[0]
                else:
                    saldo_final = df_saldo.loc["Saldo Acumulado"].iloc[-1]
                
                # Convert to numeric to handle any remaining type issues
                saldo_final = pd.to_numeric(saldo_final, errors='coerce') or 0.0
                
                # Ensure saldo_final is a scalar value for comparison
                if isinstance(saldo_final, pd.Series):
                    saldo_final = float(saldo_final.iloc[0]) if not saldo_final.empty else 0.0
                else:
                    saldo_final = float(saldo_final)
                
                st.metric(
                    label="Saldo Final", 
                    value=f"R$ {saldo_final:,.2f}", 
                    delta=f"{saldo_final:,.2f}" if saldo_final >= 0 else f"-{abs(saldo_final):,.2f}",
                    delta_color="normal" if saldo_final >= 0 else "inverse"
                )

class ProjeçãoDRE(Page):
    def __init__(self):
        self.graph_types = ["Gráfico de Linhas", "Gráfico de Pizza", "Gráfico de Barras"]
        self.time_frames = ["Mensal", "Trimestral", "Semestral", "Anual"]
    
    @property
    def title(self) -> str:
        return "Projeção DRE"
    
    @property
    def icon(self) -> str:
        return "📈"
    
    def _gerar_dataframe_dre(self):
        """Gera o dataframe DRE para 60 meses (5 anos)"""
        if 'premissas_projecoes' not in st.session_state:
            st.error("Premissas de projeções não definidas. Por favor, defina as premissas.")
            return None
        
        meses = range(60)
        
        # Criar índices multinível para o DRE
        idx = pd.MultiIndex.from_tuples([
            ("1", "Receita Bruta de Vendas"),
            ("1.1", "(-) Impostos sobre faturamento"),
            ("1.2", "(-) Comissões e taxas sobre vendas"),
            ("2", "(=) Receita Líquida de Vendas"),
            ("3", "(-) Custos de Produção"),
            ("3.1", "Custo de cobrança"),
            ("3.2", "Custo de Produtos"),
            ("4", "(=) Lucro Bruto (Margem de Contribuição)"),
            ("5", "(-) Despesas Operacionais"),
            ("5.1", "Despesas Administrativas"),
            ("5.2", "Equipe Própria (com encargos)"),
            ("5.3", "Terceiros"),
            ("5.4", "Aquisição de Equipamentos"),  # Added equipment acquisition here
            ("6", "(=) Resultado Operacional (EBITDA/LAJIDA)"),
            ("7", "(+/-) Receitas/Despesas não Operacionais"),
            ("7.1", "(-) Custos Jurídicos e arbitrais"),
            ("7.2", "(-) Amortização Empréstimo principal"),
            ("7.3", "(-) Taxas/Juros de Financiamentos"),
            ("7.4", "(-) Depreciação"),  # Keep depreciation here
            ("8", "(=) Resultado Tributável"),
            ("8.1", "(-) Imposto sobre Lucro"),
            ("9", "(=) Lucro Líquido"),
            ("9.1", "(-) Dividendos (distribuídos/provisionados)"),
            ("10", "(=) Resultado do Exercício"),
            ("11", "Margem de Contribuição (%)"),
            ("12", "Ponto de Equilíbrio Financeiro"),
        ])
        
        # Criar DataFrame base
        df = pd.DataFrame(0.0, index=idx, columns=meses)
        
        # 1. Receita Bruta de Vendas (equivalente à Arrecadação)
        try:
            receitas = Receitas()
            df_faturamento = receitas._gerar_dataframe_faturamento()
            df_arrecadacao = receitas._gerar_dataframe_arrecadacao(df_faturamento)
            
            if df_arrecadacao is not None and "Arrecadação Líquida" in df_arrecadacao.index:
                for mes in meses:
                    df.loc[("1", "Receita Bruta de Vendas"), mes] = df_arrecadacao.loc["Arrecadação Líquida", mes]
        except Exception as e:
            st.warning(f"Erro ao obter dados de arrecadação: {e}")
        
        # Inside ProjeçãoDRE._gerar_dataframe_dre() method, replace the tax sections:
        
        # 1.1 (-) Impostos sobre faturamento (PIS, COFINS, ISSQN, and IRPJ if Lucro Presumido)
        try:
            tributos = Tributos()
            df_tributos = tributos._gerar_dataframe_tributos()
            
            if df_tributos is not None:
                for mes in meses:
                    impostos_faturamento = 0
                    # Sum PIS, COFINS and ISSQN
                    if "PIS" in df_tributos.index:
                        impostos_faturamento += pd.to_numeric(df_tributos.loc["PIS", mes], errors='coerce') or 0
                    if "COFINS" in df_tributos.index:
                        impostos_faturamento += pd.to_numeric(df_tributos.loc["COFINS", mes], errors='coerce') or 0
                    if "ISSQN" in df_tributos.index:
                        impostos_faturamento += pd.to_numeric(df_tributos.loc["ISSQN", mes], errors='coerce') or 0
                    
                    # Get current tax regime for this month
                    regime_tributario = st.session_state.get('premissas_tributos', {}).get('regime_tributario', 'Simples Nacional')
                    regime_apos_desenquadramento = st.session_state.get('premissas_tributos', {}).get('regime_apos_desenquadramento', 'Lucro Presumido')
                    ano_desenquadramento = st.session_state.get('premissas_tributos', {}).get('ano_desenquadramento', None)
                    
                    regime_atual = regime_tributario
                    if ano_desenquadramento is not None:
                        ano_atual = mes // 12 + 1
                        if ano_atual >= ano_desenquadramento:
                            regime_atual = regime_apos_desenquadramento
                        
                    # Add IRPJ if Lucro Presumido regime
                    if "IRPJ" in df_tributos.index and regime_atual == "Lucro Presumido":
                        impostos_faturamento += pd.to_numeric(df_tributos.loc["IRPJ", mes], errors='coerce') or 0
                    
                    df.loc[("1.1", "(-) Impostos sobre faturamento"), mes] = -impostos_faturamento
        except Exception as e:
            st.warning(f"Erro ao obter dados de tributos: {e}")
        
        # 1.2 (-) Comissões e taxas sobre vendas
        try:
            comissao_vendas = ComissaoVendas()
            df_comissoes = comissao_vendas._calcular_comissoes()
            
            if df_comissoes is not None and not df_comissoes.empty:
                for mes in meses:
                    if mes < len(df_comissoes) and "Total" in df_comissoes.columns:
                        # Get the Total column value directly - it already includes all commissions
                        comissoes_total = pd.to_numeric(df_comissoes.loc[mes, "Total"], errors='coerce') or 0
                        # Make value negative since it's an expense
                        df.loc[("1.2", "(-) Comissões e taxas sobre vendas"), mes] = -abs(comissoes_total)
        except Exception as e:
            st.warning(f"Erro ao obter dados de comissões: {e}")
        
          # 2. (=) Receita Líquida de Vendas = Receita Bruta - Impostos - Comissões
        for mes in meses:
            # Convert each value to numeric before addition to handle type issues
            receita_bruta = pd.to_numeric(df.loc[("1", "Receita Bruta de Vendas"), mes], errors='coerce') or 0
            impostos = pd.to_numeric(df.loc[("1.1", "(-) Impostos sobre faturamento"), mes], errors='coerce') or 0
            comissoes = pd.to_numeric(df.loc[("1.2", "(-) Comissões e taxas sobre vendas"), mes], errors='coerce') or 0
            
            # Now perform the addition with numeric values
            df.loc[("2", "(=) Receita Líquida de Vendas"), mes] = receita_bruta + impostos + comissoes
      
        # 3.1 Custo de cobrança (valor fixo das premissas de projeções)
        p_projecoes = st.session_state['premissas_projecoes']
        custo_cobranca = p_projecoes.get('custos_cobranca', 0.0)
        for mes in meses:
            df.loc[("3.1", "Custo de cobrança"), mes] = -custo_cobranca
        
        # 3.2 Custo de Produtos (Total de CustosTecnologia)
        try:
            tecnologia = CustosTecnologia()
            df_tecnologia = tecnologia._gerar_dataframe_custos_tecnologia()
            
            if df_tecnologia is not None:
                for mes in meses:
                    if mes < len(df_tecnologia.columns):
                         # Convert to numeric before negating to handle type issues
                        valor = pd.to_numeric(df_tecnologia.loc["Total", mes], errors='coerce') or 0
                        df.loc[("3.2", "Custo de Produtos"), mes] = -valor
        except Exception as e:
            st.warning(f"Erro ao obter dados de custos de tecnologia: {e}")
        
         # 3. (-) Custos de Produção = soma dos custos 3.1 e 3.2
        for mes in meses:
            # Convert values to numeric before addition to ensure proper type handling
            custo_cobranca = pd.to_numeric(df.loc[("3.1", "Custo de cobrança"), mes], errors='coerce') or 0
            custo_produtos = pd.to_numeric(df.loc[("3.2", "Custo de Produtos"), mes], errors='coerce') or 0
            
            # Add the numeric values
            df.loc[("3", "(-) Custos de Produção"), mes] = custo_cobranca + custo_produtos
       
         # 4. (=) Lucro Bruto (Margem de Contribuição) = Receita Líquida + Custos de Produção
        for mes in meses:
            receita_liquida = pd.to_numeric(df.loc[("2", "(=) Receita Líquida de Vendas"), mes], errors='coerce') or 0
            custos_producao = pd.to_numeric(df.loc[("3", "(-) Custos de Produção"), mes], errors='coerce') or 0
            df.loc[("4", "(=) Lucro Bruto (Margem de Contribuição)"), mes] = receita_liquida + custos_producao
       
        # 5.1 Despesas Administrativas (total de DespesasAdm)
        try:
            despesas_adm = DespesasAdm()
            df_despesas = despesas_adm._gerar_dataframe_despesas()
            
            if df_despesas is not None:
                for mes in meses:
                     if mes < len(df_despesas):
                        # Convert to numeric before negating to handle type issues
                        valor = pd.to_numeric(df_despesas.loc[mes, "Total"], errors='coerce') or 0
                        df.loc[("5.1", "Despesas Administrativas"), mes] = -valor
        except Exception as e:
            st.warning(f"Erro ao obter dados de despesas administrativas: {e}")
        
        # Replace the existing "Equipe Própria (com encargos)" section in _gerar_dataframe_dre method in ProjeçãoDRE
        
        # 5.2 Equipe Própria (com encargos)
        try:
            # Get data from Receitas._calcular_equipe_comercial
            receitas = Receitas()
            df_faturamento = receitas._gerar_dataframe_faturamento()
            df_equipe_comercial = receitas._calcular_equipe_comercial(df_faturamento)
            
            # Get data from ComissaoVendas
            comissao_vendas = ComissaoVendas()
            df_comissoes = comissao_vendas._calcular_comissoes()
            
            # Get encargos sociais percentage
            encargos_sociais_perc = 68.0  # Default
            if 'premissas_despesas' in st.session_state:
                encargos_sociais_perc = st.session_state['premissas_despesas'].get('encargos_sociais_perc', 68.0)
            
            # Calculate total team compensation with encargos
            for mes in meses:
                total_remuneracao = 0
                
                # Add SDR and Closer salaries
                if not df_equipe_comercial.empty and mes < len(df_equipe_comercial.columns):
                    salario_sdr = pd.to_numeric(df_equipe_comercial.loc["Salário SDR", mes], errors='coerce') or 0
                    salario_closer = pd.to_numeric(df_equipe_comercial.loc["Salário Closer", mes], errors='coerce') or 0
                    total_remuneracao += salario_sdr + salario_closer
                
                # Add other employee salaries from PremissasDespesas
                if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
                    for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                        if 'nome' in cargo and cargo['nome'] not in ['SDR', 'Closer']:  # Exclude SDR and Closer
                            salario = cargo.get('salario', 0)
                            quantidade = cargo.get('quantidade', 0)
                            total_remuneracao += salario * quantidade
                
                # Apply social security charges (encargos sociais)
                total_remuneracao_com_encargos = total_remuneracao * (1 + encargos_sociais_perc / 100)
                
                # Add commissions (they are already included in the compensation, no encargos needed)
                if df_comissoes is not None and not df_comissoes.empty and mes < len(df_comissoes) and "Total" in df_comissoes.columns:
                    comissoes_total = pd.to_numeric(df_comissoes.loc[mes, "Total"], errors='coerce') or 0
                    total_remuneracao_com_encargos += comissoes_total
                
                # Make value negative since it's an expense
                df.loc[("5.2", "Equipe Própria (com encargos)"), mes] = -total_remuneracao_com_encargos
        except Exception as e:
            st.warning(f"Erro ao calcular custos de equipe própria: {e}")
        
        # 5.4 Aquisição de Equipamentos (valor fixo das premissas de projeções)
        try:
                
            # 5.4 Aquisição de Equipamentos
            if 'premissas_despesas' in st.session_state and 'equipamentos' in st.session_state['premissas_despesas']:
                for mes in meses:
                    equipment_cost = 0
                    for equip in st.session_state['premissas_despesas']['equipamentos']:
                        mes_aquisicao = equip.get('mes_aquisicao', 0)
                        if mes == mes_aquisicao:
                            equipment_cost += equip['valor'] * equip['quantidade']
                    df.loc[("5.4", "Aquisição de Equipamentos"), mes] = -equipment_cost
            
            # Update the total Despesas Operacionais calculation to include equipment acquisition
            for mes in meses:
                # Convert each value to numeric before addition
                despesas_adm = pd.to_numeric(df.loc[("5.1", "Despesas Administrativas"), mes], errors='coerce') or 0
                equipe_propria = pd.to_numeric(df.loc[("5.2", "Equipe Própria (com encargos)"), mes], errors='coerce') or 0
                terceiros = pd.to_numeric(df.loc[("5.3", "Terceiros"), mes], errors='coerce') or 0
                equipamentos = pd.to_numeric(df.loc[("5.4", "Aquisição de Equipamentos"), mes], errors='coerce') or 0
                
                # Now perform the addition with numeric values
                df.loc[("5", "(-) Despesas Operacionais"), mes] = despesas_adm + equipe_propria + terceiros + equipamentos
        except Exception as e:
            st.warning(f"Erro ao obter dados de aquisição de equipamentos: {e}")
            
        # 5. (-) Despesas Operacionais = soma de 5.1, 5.2 e 5.3
        for mes in meses:
            # Convert each value to numeric before addition
            despesas_adm = pd.to_numeric(df.loc[("5.1", "Despesas Administrativas"), mes], errors='coerce') or 0
            equipe_propria = pd.to_numeric(df.loc[("5.2", "Equipe Própria (com encargos)"), mes], errors='coerce') or 0
            terceiros = pd.to_numeric(df.loc[("5.3", "Terceiros"), mes], errors='coerce') or 0
            
            # Now perform the addition with numeric values
            df.loc[("5", "(-) Despesas Operacionais"), mes] = despesas_adm + equipe_propria + terceiros
       
         # 6. (=) Resultado Operacional (EBITDA/LAJIDA) = Lucro Bruto + Despesas Operacionais
        for mes in meses:
            # Convert each value to numeric before addition to ensure proper type handling
            lucro_bruto = pd.to_numeric(df.loc[("4", "(=) Lucro Bruto (Margem de Contribuição)"), mes], errors='coerce') or 0
            despesas_operacionais = pd.to_numeric(df.loc[("5", "(-) Despesas Operacionais"), mes], errors='coerce') or 0
            
            # Perform the addition with numeric values
            df.loc[("6", "(=) Resultado Operacional (EBITDA/LAJIDA)"), mes] = lucro_bruto + despesas_operacionais
       
        # 7.1 (-) Custos Jurídicos e arbitrais (valor fixo das premissas de projeções)
        custos_juridicos = p_projecoes.get('custos_juridicos', 0.0)
        for mes in meses:
            df.loc[("7.1", "(-) Custos Jurídicos e arbitrais"), mes] = -custos_juridicos
        
        # 7.2 e 7.3 - Usar valores da projeção de fluxo de caixa
        try:
            projecao_fc = ProjeçãodeFluxodeCaixa()
            _, df_juros, df_amortizacoes = projecao_fc._calcular_emprestimos()
            
            if not df_juros.empty:
                # Get all negative values (expenses) for each month from df_juros
                for mes in meses:
                    juros_valor = pd.to_numeric(df_juros.loc["Taxas/Juros de Empréstimos", mes], errors='coerce') or 0
                    df.loc[("7.3", "(-) Taxas/Juros de Financiamentos"), mes] = juros_valor  # Already negative
            
            if not df_amortizacoes.empty:
                # Get all negative values (expenses) for each month from df_amortizacoes
                for mes in meses:
                    amortizacao_valor = pd.to_numeric(df_amortizacoes.loc["Amortização de empréstimos", mes], errors='coerce') or 0
                    df.loc[("7.2", "(-) Amortização Empréstimo principal"), mes] = amortizacao_valor  # Already negative
        except Exception as e:
            st.warning(f"Erro ao obter dados de empréstimos: {e}")
        
         # 7. (+/-) Receitas/Despesas não Operacionais = soma de 7.1, 7.2, 7.3 e 7.4
        for mes in meses:
            # Convert each value to numeric before addition to handle type issues
            custos_juridicos = pd.to_numeric(df.loc[("7.1", "(-) Custos Jurídicos e arbitrais"), mes], errors='coerce') or 0
            amortizacao = pd.to_numeric(df.loc[("7.2", "(-) Amortização Empréstimo principal"), mes], errors='coerce') or 0
            taxas_juros = pd.to_numeric(df.loc[("7.3", "(-) Taxas/Juros de Financiamentos"), mes], errors='coerce') or 0
            
            # 7.4 Depreciação
            depreciation_value = 0.0
            try:
                # Get depreciation data from CustosTecnologia
                tecnologia = CustosTecnologia()
                df_tecnologia = tecnologia._gerar_dataframe_custos_tecnologia()
                
                if df_tecnologia is not None and "Depreciação de Equipamentos" in df_tecnologia.index:
                    # Get the depreciation value from CustosTecnologia for this month
                    if mes < len(df_tecnologia.columns):
                        # Get the depreciation value (already negative in CustosTecnologia)
                        depreciation_value = pd.to_numeric(df_tecnologia.loc["Depreciação de Equipamentos", mes], errors='coerce') or 0
                        
                # Add depreciation value to the DRE
                df.loc[("7.4", "(-) Depreciação"), mes] = depreciation_value
            except Exception as e:
                st.warning(f"Erro ao calcular depreciação para DRE: {str(e)}")
                # Ensure the value is set to 0.0 in case of error
                df.loc[("7.4", "(-) Depreciação"), mes] = 0.0

            # Now perform the addition with numeric values
            df.loc[("7", "(+/-) Receitas/Despesas não Operacionais"), mes] = custos_juridicos + amortizacao + taxas_juros + depreciation_value
    
        # 8. (=) Resultado Tributável = Resultado Operacional + Receitas/Despesas não Operacionais
        for mes in meses:
            # Convert values to numeric before addition to ensure proper type handling
            resultado_operacional = pd.to_numeric(df.loc[("6", "(=) Resultado Operacional (EBITDA/LAJIDA)"), mes], errors='coerce') or 0
            receitas_despesas = pd.to_numeric(df.loc[("7", "(+/-) Receitas/Despesas não Operacionais"), mes], errors='coerce') or 0
            
            # Perform the addition with numeric values
            df.loc[("8", "(=) Resultado Tributável"), mes] = resultado_operacional + receitas_despesas
       
        # 8.1 (-) Imposto sobre Lucro (soma de IRPJ e CSLL)
        try:
            tributos = Tributos()
            df_tributos = tributos._gerar_dataframe_tributos()
            
            # 8.1 (-) Imposto sobre Lucro (soma de IRPJ e CSLL)
            try:
                # For impostos sobre lucro (profit taxes):
                if df_tributos is not None:
                    for mes in meses:
                        imposto_lucro = 0
                        
                        # Get current tax regime for this month
                        regime_tributario = st.session_state.get('premissas_tributos', {}).get('regime_tributario', 'Simples Nacional')
                        regime_apos_desenquadramento = st.session_state.get('premissas_tributos', {}).get('regime_apos_desenquadramento', 'Lucro Presumido')
                        ano_desenquadramento = st.session_state.get('premissas_tributos', {}).get('ano_desenquadramento', None)
                        regime_atual = regime_tributario
                        if ano_desenquadramento is not None:
                            ano_atual = mes // 12 + 1
                            if ano_atual >= ano_desenquadramento:
                                regime_atual = regime_apos_desenquadramento
            
                        # Always include CSLL
                        if "CSLL" in df_tributos.index:
                            imposto_lucro += pd.to_numeric(df_tributos.loc["CSLL", mes], errors='coerce') or 0

                        # Add IRPJ if not Lucro Presumido regime
                        if "IRPJ" in df_tributos.index and regime_atual != "Lucro Presumido":
                            imposto_lucro += pd.to_numeric(df_tributos.loc["IRPJ", mes], errors='coerce') or 0
            
                        df.loc[("8.1", "(-) Imposto sobre Lucro"), mes] = -imposto_lucro
            except Exception as e:
                st.warning(f"Erro ao obter dados de tributos sobre lucro: {e}")
        except Exception as e:
            st.warning(f"Erro ao obter dados de tributos: {e}")
        
         # 9. (=) Lucro Líquido = Resultado Tributável + Imposto sobre Lucro
        for mes in meses:
            # Convert each value to numeric before addition to handle type issues
            resultado_tributavel = pd.to_numeric(df.loc[("8", "(=) Resultado Tributável"), mes], errors='coerce') or 0
            imposto_lucro = pd.to_numeric(df.loc[("8.1", "(-) Imposto sobre Lucro"), mes], errors='coerce') or 0
            
            # Now perform the addition with numeric values
            df.loc[("9", "(=) Lucro Líquido"), mes] = resultado_tributavel + imposto_lucro
       
        # 9.1 (-) Dividendos (soma de bônus da classe Equipe)
        try:
            equipe = Equipe()
            df_equipe = equipe._gerar_dataframe_custos_equipe()
            
            if df_equipe is not None:
                for mes in meses:
                    if mes < len(df_equipe.columns):
                        dividendos_total = 0
                        for idx in df_equipe.index:
                            if idx[0] == "Bônus dos Lucros":
                                dividendos_total += pd.to_numeric(df_equipe.loc[idx, mes+1], errors='coerce') or 0
                       
                        df.loc[("9.1", "(-) Dividendos (distribuídos/provisionados)"), mes] = -dividendos_total
        except Exception as e:
            st.warning(f"Erro ao obter dados de dividendos: {e}")
        
         # 10. (=) Resultado do Exercício = Lucro Líquido + Dividendos
        for mes in meses:
            # Convert each value to numeric before addition to handle type issues
            lucro_liquido = pd.to_numeric(df.loc[("9", "(=) Lucro Líquido"), mes], errors='coerce') or 0
            dividendos = pd.to_numeric(df.loc[("9.1", "(-) Dividendos (distribuídos/provisionados)"), mes], errors='coerce') or 0
            
            # Now perform the addition with numeric values
            df.loc[("10", "(=) Resultado do Exercício"), mes] = lucro_liquido + dividendos
       
         # 11. Margem de Contribuição (%) = Lucro Bruto / Receita Bruta * 100
        for mes in meses:
            receita_bruta = pd.to_numeric(df.loc[("1", "Receita Bruta de Vendas"), mes], errors='coerce') or 0
            lucro_bruto = pd.to_numeric(df.loc[("4", "(=) Lucro Bruto (Margem de Contribuição)"), mes], errors='coerce') or 0
            
            if receita_bruta != 0:
                df.loc[("11", "Margem de Contribuição (%)"), mes] = (lucro_bruto / receita_bruta) * 100
            else:
                df.loc[("11", "Margem de Contribuição (%)"), mes] = 0
       
         # 12. Ponto de Equilíbrio Financeiro
        for mes in meses:
            margem_contribuicao = pd.to_numeric(df.loc[("11", "Margem de Contribuição (%)"), mes], errors='coerce') or 0
            lucro_bruto = pd.to_numeric(df.loc[("4", "(=) Lucro Bruto (Margem de Contribuição)"), mes], errors='coerce') or 0
            
            if margem_contribuicao != 0 and lucro_bruto > 0:
                # Ponto de Equilíbrio = -(Despesas Operacionais + Custos Jurídicos + Taxas/Juros) / Margem de Contribuição
                # Convert all values to numeric before performing arithmetic operations
                despesas_op = pd.to_numeric(df.loc[("5", "(-) Despesas Operacionais"), mes], errors='coerce') or 0
                custos_jur = pd.to_numeric(df.loc[("7.1", "(-) Custos Jurídicos e arbitrais"), mes], errors='coerce') or 0
                taxas_jur = pd.to_numeric(df.loc[("7.3", "(-) Taxas/Juros de Financiamentos"), mes], errors='coerce') or 0
                
                ponto_equilibrio = -(despesas_op + custos_jur + taxas_jur) / (margem_contribuicao / 100)
               
                df.loc[("12", "Ponto de Equilíbrio Financeiro"), mes] = ponto_equilibrio
            else:
                df.loc[("12", "Ponto de Equilíbrio Financeiro"), mes] = 0
        
        return df
    
    def _gerar_dataframe_anual(self, df_mensal):
        """Converte o dataframe mensal para anual"""
        if df_mensal is None or df_mensal.empty:
            return None
            
        anos = 5
        df_anual = pd.DataFrame(index=df_mensal.index, columns=range(1, anos + 1))
        
        for ano in range(1, anos + 1):
            mes_inicio = (ano - 1) * 12
            mes_fim = ano * 12 - 1
            
            # Para cada categoria, somar os valores dos meses do ano
            for categoria in df_mensal.index:
                df_anual.loc[categoria, ano] = df_mensal.loc[categoria, mes_inicio:mes_fim].sum()
        
        # Recalcular algumas métricas para o período anual (médias em vez de somas)
        for ano in range(1, anos + 1):
             # Margem de Contribuição
            receita_bruta = pd.to_numeric(df_anual.loc[("1", "Receita Bruta de Vendas"), ano], errors='coerce') or 0
            lucro_bruto = pd.to_numeric(df_anual.loc[("4", "(=) Lucro Bruto (Margem de Contribuição)"), ano], errors='coerce') or 0
            
            if receita_bruta != 0:
                df_anual.loc[("11", "Margem de Contribuição (%)"), ano] = (lucro_bruto / receita_bruta) * 100
            else:
                df_anual.loc[("11", "Margem de Contribuição (%)"), ano] = 0
           
             # Ponto de Equilíbrio
            margem_contribuicao = pd.to_numeric(df_anual.loc[("11", "Margem de Contribuição (%)"), ano], errors='coerce') or 0
            lucro_bruto = pd.to_numeric(df_anual.loc[("4", "(=) Lucro Bruto (Margem de Contribuição)"), ano], errors='coerce') or 0
            # Convert all values to numeric before performing arithmetic operations
            despesas_op = pd.to_numeric(df_anual.loc[("5", "(-) Despesas Operacionais"), ano], errors='coerce') or 0
            custos_jur = pd.to_numeric(df_anual.loc[("7.1", "(-) Custos Jurídicos e arbitrais"), ano], errors='coerce') or 0
            taxas_jur = pd.to_numeric(df_anual.loc[("7.3", "(-) Taxas/Juros de Financiamentos"), ano], errors='coerce') or 0
            
            ponto_equilibrio = -(despesas_op + custos_jur + taxas_jur) / (margem_contribuicao / 100)

            if margem_contribuicao != 0 and lucro_bruto > 0:    
                df_anual.loc[("12", "Ponto de Equilíbrio Financeiro"), ano] = ponto_equilibrio
            else:
                df_anual.loc[("12", "Ponto de Equilíbrio Financeiro"), ano] = 0
        
        return df_anual
    
    def _calculate_break_even_conversions(self):
        """Calculate the number of conversions needed to break-even in the first year"""
        # Get total outflows from ProjeçãoFluxodeCaixa
        try:
            projecao_fc = ProjeçãodeFluxodeCaixa()
            df_saidas = projecao_fc._gerar_dataframe_saidas()
            if df_saidas is not None and not df_saidas.empty and ("TOTAL", "Total Saídas") in df_saidas.index:
                # Calculate total outflows for the first 12 months (0-11)
                total_saidas_12meses = abs(sum(pd.to_numeric(df_saidas.loc[("TOTAL", "Total Saídas"), mes], errors='coerce') or 0 for mes in range(12)))
            else:
                return None, None
        except Exception as e:
            st.warning(f"Erro ao obter dados de saídas: {e}")
            return None, None
        
        # Get valor unitario and produtos por lead from premissas_despesas
        valor_unitario = 0
        produtos_por_lead = 0
        if 'premissas_despesas' in st.session_state and 'equipe_propria' in st.session_state['premissas_despesas']:
            for cargo in st.session_state['premissas_despesas']['equipe_propria']:
                if cargo.get('nome') == 'Closer':
                    valor_unitario = cargo.get('valor_unitario', 0)
                    produtos_por_lead = cargo.get('produtos_por_lead', 0)
                    break
        
        if valor_unitario <= 0 or produtos_por_lead <= 0:
            return None, None
        
        valor_inflow = (valor_unitario * 12) - ((valor_unitario * 12) * 0.2)

        # Calculate break-even conversions
        receita_por_conversao = valor_inflow * produtos_por_lead
        break_even_conversions = total_saidas_12meses / receita_por_conversao if receita_por_conversao > 0 else 0
        
        return break_even_conversions, total_saidas_12meses
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        if 'premissas_projecoes' not in st.session_state:
            st.error("Premissas de projeções não definidas. Por favor, defina as premissas na página 'Premissas Projeções'.")
            return
        
        # Gerar o DataFrame DRE
        df_mensal = self._gerar_dataframe_dre()
        
        if df_mensal is None:
            st.error("Não foi possível gerar a projeção DRE. Verifique os dados de entrada.")
            return
        
        # Opções de visualização
        col1, col2 = st.columns(2)
        
        with col1:
            time_frame = st.selectbox("Período", self.time_frames, index=1)  # Default para visualização anual
        
        with col2:
            plot_type = st.selectbox("Tipo de Gráfico", self.graph_types, index=0)
        
        # Converter para anual se selecionado
        df_display = df_mensal.copy()
        if time_frame == "Anual":
            df_display = self._gerar_dataframe_anual(df_mensal)
            if df_display is not None:
                df_display.columns = [f"Ano {ano}" for ano in df_display.columns]
        else:
            if df_display is not None:
                df_display.columns = [f"Mês {mes}" for mes in df_display.columns]
        
        # Mostrar o dataframe
        periodo = "Anual" if time_frame == "Anual" else "Mensal"
        st.write(f"### Demonstração de Resultado do Exercício - Visão {periodo}")
        
        if df_display is not None:
            # Formatar o DataFrame para exibição, aplicando formato monetário
            st.dataframe(df_display.style.format(lambda x: f"{x:,.2f}" if isinstance(x, (int, float)) else str(x)), 
                        use_container_width=True)
       
        # Selecionar linhas para visualização gráfica
        st.write("### Visualização Gráfica")
        
        # Categorias principais para visualizar (por exemplo, receitas, lucro, resultado)
        categorias_principais = [
            ("1", "Receita Bruta de Vendas"),
            ("2", "(=) Receita Líquida de Vendas"),
            ("4", "(=) Lucro Bruto (Margem de Contribuição)"),
            ("6", "(=) Resultado Operacional (EBITDA/LAJIDA)"),
            ("9", "(=) Lucro Líquido"),
            ("10", "(=) Resultado do Exercício"),
            ("TOTAL", "TOTAL")
        ]
        
        # Verificar quais categorias existem no DataFrame
        categorias_disponiveis = []
        if df_display is not None:
            categorias_disponiveis = [cat for cat in categorias_principais if cat in df_display.index]
      
         # Adicionar opção para métricas/indicadores
        metricas = []
        if df_display is not None:
            metricas = [
                ("11", "Margem de Contribuição (%)"),
                ("12", "Ponto de Equilíbrio Financeiro")
            ]
        
         # Separar visualizações para valores monetários e métricas
        tab1, tab2 = st.tabs(["Resultados Financeiros", "Métricas e Indicadores"])

        with tab1:
            # Seleção múltipla para categorias financeiras
            selected_categories = st.multiselect(
              "Selecione as categorias para visualizar", 
                categorias_disponiveis,
                default=[categorias_disponiveis[0], categorias_disponiveis[-1]] if len(categorias_disponiveis) > 1 else categorias_disponiveis,
                format_func=lambda x: x[1],
                key="dre_categories"
            )
            
            if not selected_categories:
                selected_categories = [categorias_disponiveis[0]] if categorias_disponiveis else []
            
            if selected_categories and df_display is not None:
                if plot_type == "Gráfico de Linhas":
                    # Preparar dados para o gráfico
                    df_plot = df_display.loc[selected_categories].T
                    
                    # Renomear as colunas para exibir apenas os nomes descritivos
                    df_plot.columns = [cat[1] for cat in selected_categories]
                    
                    # Criar gráfico de linhas
                    fig = px.line(
                        df_plot,
                        title=f"Evolução dos Resultados Financeiros ({periodo})",
                        markers=True
                    )
                    
                    fig.update_layout(
                        xaxis_title=f"{'Ano' if time_frame == 'Anual' else 'Mês'}",
                        yaxis_title="Valor (R$)",
                        legend_title="Categoria"
                    )
                    
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    # Para gráfico de pizza, usamos a soma total do período
                    data_pie = {
                        'Categoria': [cat[1] for cat in selected_categories],
                        'Valor': [df_display.loc[cat].sum() for cat in selected_categories]
                    }
                    
                    # Converter valores negativos para positivos para visualização
                    df_pie = pd.DataFrame(data_pie)
                    df_pie['Valor'] = df_pie['Valor'].abs()
                    
                    fig = px.pie(
                        df_pie,
                        values='Valor',
                        names='Categoria',
                        title=f"Composição dos Resultados Financeiros - Total {periodo}"
                    )
                    
                    st.plotly_chart(fig, use_container_width=True)
        
        with tab2:
            # Visualização de métricas
            metricas_disponiveis = [m for m in metricas if df_display is not None and m in df_display.index]

            if metricas_disponiveis and df_display is not None:
                # Preparar dados para o gráfico de métricas
                df_metricas = df_display.loc[metricas_disponiveis].T
               
                # Renomear as colunas para exibir apenas os nomes descritivos
                df_metricas.columns = [m[1] for m in metricas_disponiveis]
               
                # Criar gráfico de linhas para métricas
                fig_metricas = px.line(
                    df_metricas,
                    title=f"Evolução de Métricas e Indicadores ({periodo})",
                    markers=True
                )
                
                fig_metricas.update_layout(
                    xaxis_title=f"{'Ano' if time_frame == 'Anual' else 'Mês'}",
                    yaxis_title="Valor",
                    legend_title="Métrica"
                )
                
                st.plotly_chart(fig_metricas, use_container_width=True)
            else:
                 st.info("Não há métricas disponíveis para visualização.")
        
        # Métricas principais
        # In the ProjeçãoDRE class, modify the render method's metrics section
        # Look for "Métricas-Chave" section and update the col1 section
        
        # Métricas principais
        st.write("### Métricas-Chave")
        col1, col2, col3, col4, col5 = st.columns(5)  # Change from 4 columns to 5

        with col1:
            # Get DRE data to calculate "Faturamento Acumulado (12 meses)"
            try:
                dre = ProjeçãoDRE()
                df_dre = dre._gerar_dataframe_dre()
                if df_dre is not None and ("1", "Receita Bruta de Vendas") in df_dre.index:
                    # Sum the first 12 months (0 to 11)
                    faturamento_acumulado = df_dre.loc[("1", "Receita Bruta de Vendas"), :11].sum()
                    st.metric(
                        label="Faturamento Acumulado (12 meses)", 
                        value=f"R$ {faturamento_acumulado:,.2f}"
                    )
                else:
                    st.metric(label="Faturamento Acumulado (12 meses)", value="N/A")
            except Exception as e:
                st.warning(f"Erro ao calcular Faturamento Acumulado: {e}")
                st.metric(label="Faturamento Acumulado (12 meses)", value="N/A")
        
            with col2:
            # Get DRE data to calculate "Resultado do Exercício Acumulado (12 meses)"
                df_dre = None  # Ensure df_dre is always defined
            try:
                dre = ProjeçãoDRE()
                df_dre = dre._gerar_dataframe_dre()
                if df_dre is not None and ("10", "(=) Resultado do Exercício") in df_dre.index:
                    # Sum the first 12 months (0 to 11)
                    resultado_exercicio_acumulado = df_dre.loc[("10", "(=) Resultado do Exercício"), :11].sum()
                    st.metric(
                        label="Resultado do Exercício Acumulado (12 meses)", 
                        value=f"R$ {resultado_exercicio_acumulado:,.2f}"
                    )
                else:
                    st.metric(label="Resultado do Exercício Acumulado (12 meses)", value="N/A")
            except Exception as e:
                st.warning(f"Erro ao calcular Resultado Acumulado: {e}")
                st.metric(label="Resultado do Exercício Acumulado (12 meses)", value="N/A")
    
        with col3:
            # Calculate "EBITDA Acumulado (12 meses)"
            try:
                if df_dre is not None and ("6", "(=) Resultado Operacional (EBITDA/LAJIDA)") in df_dre.index:
                    # Sum the first 12 months (0 to 11)
                    ebitda_acumulado = df_dre.loc[("6", "(=) Resultado Operacional (EBITDA/LAJIDA)"), :11].sum()
                    st.metric(
                        label="EBITDA Acumulado (12 meses)", 
                        value=f"R$ {ebitda_acumulado:,.2f}"
                    )
                else:
                    st.metric(label="EBITDA Acumulado (12 meses)", value="N/A")
            except Exception as e:
                st.warning(f"Erro ao calcular EBITDA Acumulado: {e}")
                st.metric(label="EBITDA Acumulado (12 meses)", value="N/A")
        
        with col4:
            # Calculate "Margem Líquida Média (12 meses)"
            try:
                if df_dre is not None and ("1", "Receita Bruta de Vendas") in df_dre.index and ("10", "(=) Resultado do Exercício") in df_dre.index:
                    # Get first 12 months of revenue and results
                    receita_12_meses = df_dre.loc[("1", "Receita Bruta de Vendas"), :11].sum()
                    resultado_12_meses = df_dre.loc[("10", "(=) Resultado do Exercício"), :11].sum()
                    
                    if receita_12_meses.sum() > 0:
                        margem_liquida = (resultado_12_meses.sum() / receita_12_meses.sum()) * 100
                        st.metric(label="Margem Líquida Média (12 meses)", value=f"{margem_liquida:.2f}%")
                    else:
                        st.metric(label="Margem Líquida Média (12 meses)", value="N/A")
                else:
                    st.metric(label="Margem Líquida Média (12 meses)", value="N/A")
            except Exception as e:
                st.warning(f"Erro ao calcular Margem Líquida: {e}")
                st.metric(label="Margem Líquida Média (12 meses)", value="N/A")
        
        with col5:
            # Calculate Break-Even Conversions
            break_even_conversions, total_saidas_12meses = self._calculate_break_even_conversions()
            
            if break_even_conversions is not None and break_even_conversions > 0:
                st.metric(
                    label="Break-Even (12 meses)",
                    value=f"{int(break_even_conversions)} conversões"
                )
                if total_saidas_12meses is not None:
                    st.caption(f"Para cobrir R$ {total_saidas_12meses:,.2f} em saídas no primeiro ano")
            else:
                st.metric(
                    label="Break-Even (12 meses)",
                    value="N/A"
                )
                st.caption("Dados insuficientes para cálculo")

class PaginaAcompanhamento(Page):
    def __init__(self):
        self.time_frames = ["Mensal", "Trimestral", "Anual"]
        self.periods = {"Mensal": 1, "Trimestral": 3, "Anual": 12}
    
    @property
    def title(self) -> str:
        return "Página de Acompanhamento"
    
    @property
    def icon(self) -> str:
        return "📊"
    
    def _generate_user_inputs_report(self):
        """Collects all user inputs from session state and generates a report DataFrame"""
        # Initialize list for all inputs
        all_inputs = []
        
        # Collect data from each session state dictionary with inputs
        input_sources = [
            ('Investimentos', 'premissas_investimentos'),
            ('Comissões', 'premissas_comissao'),
            ('Despesas', 'premissas_despesas'),
            ('Receitas', 'premissas_receitas'), 
            ('Tributos', 'premissas_tributos'),
            ('Projeções', 'premissas_projecoes')
        ]
        
        # Iterate through each input source and collect data
        for category, state_key in input_sources:
            if state_key in st.session_state:
                self._flatten_dict(all_inputs, st.session_state[state_key], category, '')
        
        # Create DataFrame from collected inputs
        df = pd.DataFrame(all_inputs, columns=['Categoria', 'Parâmetro', 'Valor'])
        return df
    
    def _flatten_dict(self, result_list, d, category, prefix=''):
        """Recursively flattens a nested dictionary into a list of (category, parameter, value) tuples"""
        for k, v in d.items():
            param_name = f"{prefix}.{k}" if prefix else k
            
            if isinstance(v, dict):
                self._flatten_dict(result_list, v, category, param_name)
            elif isinstance(v, list):
                # Handle lists - convert to string representation
                result_list.append([category, param_name, str(v)])
            else:
                # Add simple values directly
                result_list.append([category, param_name, v])
    
    def _collect_all_dataframes(self):
        """Collects all generated dataframes from different pages for export"""
        all_dataframes = {}
        
        # Collect Investimentos dataframes
        try:
            investimentos = Investimentos()
            df_investimentos = investimentos._gerar_dataframe_investimentos()
            if df_investimentos is not None and not df_investimentos.empty:
                all_dataframes["Investimentos"] = df_investimentos
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Investimentos: {e}")
        
        # Collect Despesas Administrativas dataframes
        try:
            despesas_adm = DespesasAdm()
            df_despesas = despesas_adm._gerar_dataframe_despesas()
            if df_despesas is not None and not df_despesas.empty:
                all_dataframes["Despesas Administrativas"] = df_despesas
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Despesas Administrativas: {e}")
        
        # Collect Equipe dataframes
        try:
            equipe = Equipe()
            df_equipe = equipe._gerar_dataframe_custos_equipe()
            if df_equipe is not None and not df_equipe.empty:
                all_dataframes["Custos de Equipe"] = df_equipe
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Equipe: {e}")
        
        # Collect Tecnologia dataframes
        try:
            tecnologia = CustosTecnologia()
            df_tecnologia = tecnologia._gerar_dataframe_custos_tecnologia()
            if df_tecnologia is not None and not df_tecnologia.empty:
                all_dataframes["Custos de Tecnologia"] = df_tecnologia
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Tecnologia: {e}")
        
        # Collect Comissões dataframes
        try:
            comissoes = ComissaoVendas()
            df_comissoes = comissoes._calcular_comissoes()
            if df_comissoes is not None and not df_comissoes.empty:
                all_dataframes["Comissões"] = df_comissoes
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Comissões: {e}")
        
        # Collect Receitas dataframes
        try:
            receitas = Receitas()
            df_faturamento = receitas._gerar_dataframe_faturamento()
            if df_faturamento is not None and not df_faturamento.empty:
                all_dataframes["Faturamento"] = df_faturamento
            
            df_arrecadacao = receitas._gerar_dataframe_arrecadacao(df_faturamento)
            if df_arrecadacao is not None and not df_arrecadacao.empty:
                all_dataframes["Arrecadação"] = df_arrecadacao
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Receitas: {e}")
        
        # Collect Tributos dataframes
        try:
            tributos = Tributos()
            df_tributos = tributos._gerar_dataframe_tributos()
            if df_tributos is not None and not df_tributos.empty:
                all_dataframes["Tributos"] = df_tributos
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Tributos: {e}")
        
        # Collect Fluxo de Caixa dataframes
        try:
            fluxo_caixa = ProjeçãodeFluxodeCaixa()
            df_entradas = fluxo_caixa._gerar_dataframe_entradas()
            if df_entradas is not None and not df_entradas.empty:
                all_dataframes["Fluxo Caixa - Entradas"] = df_entradas
            
            df_saidas = fluxo_caixa._gerar_dataframe_saidas()
            if df_saidas is not None and not df_saidas.empty:
                all_dataframes["Fluxo Caixa - Saídas"] = df_saidas
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Fluxo de Caixa: {e}")
        
        # Collect DRE dataframes
        try:
            dre = ProjeçãoDRE()
            df_dre = dre._gerar_dataframe_dre()
            if df_dre is not None and not df_dre.empty:
                all_dataframes["DRE"] = df_dre
        except Exception as e:
            st.warning(f"Erro ao coletar dados de DRE: {e}")
        
        # Collect Metas dataframes
        try:
            metas = MetasColabs()
            df_metas = metas._calcular_metas_mensais()  # Default values
            if df_metas is not None and not df_metas.empty:
                all_dataframes["Metas"] = df_metas
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Metas: {e}")
        
                # Collect Projeçoes 12 Meses dataframes
        try:
            proj_inic = ProjeçãoInicial()
            df_projecao_inicial = proj_inic._get_receitas_data()
            if df_projecao_inicial is not None and isinstance(df_projecao_inicial, pd.DataFrame) and not df_projecao_inicial.empty:
                all_dataframes["Projeção Inicial"] = df_projecao_inicial
            elif df_projecao_inicial is not None and isinstance(df_projecao_inicial, pd.Series):
                all_dataframes["Projeção Inicial"] = pd.DataFrame(df_projecao_inicial)
        except Exception as e:
            st.warning(f"Erro ao coletar dados de Projeções 12 Meses: {e}")
    
        return all_dataframes
    
    def _export_dataframes_to_excel(self, dataframes):
        """Exports all dataframes to a single Excel file"""
        import io
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        if not dataframes:
            return None
        
        # Create a workbook and buffer
        buffer = io.BytesIO()
        wb = Workbook()
        
          # Remove default worksheet if it exists
        if wb.active is not None:
            wb.remove(wb.active)
      
        # Add each dataframe as a worksheet
        for sheet_name, df in dataframes.items():
            # Create a valid Excel sheet name (max 31 chars, no special chars)
            valid_sheet_name = sheet_name[:31].replace("[", "").replace("]", "").replace(":", "")
            
            # Add sheet
            ws = wb.create_sheet(title=valid_sheet_name)
            
            # Convert DataFrame to rows and add to worksheet
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)
            
            # Basic formatting
            for cell in ws[1]:
                cell.style = 'Headline 2'
        
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Create tabs for different analyses
        tabs = st.tabs(["Relatório Inputs", "Exportar Dados"])
        tab1 = tabs[0]  # Get the first tab object
        tab2 = tabs[1]  # Get the second tab object
    
        # Tab 1: Relatório Inputs
        with tab1:
            styled_title("Relatório Inputs", level=2)
    
            # Generate and display the report
            df_report = self._generate_user_inputs_report()
            st.dataframe(df_report)
            
            st.write("### Relatório de Parâmetros")
            st.write("Gere um relatório com todos os parâmetros configurados no sistema.")
            
            if st.button("Gerar Relatório CSV"):
                # Generate the report DataFrame
                df_report = self._generate_user_inputs_report()
                
                # Display a preview of the report
                st.write("#### Prévia do Relatório")
                st.dataframe(df_report, use_container_width=True)
                
                # Convert DataFrame to CSV
                csv = df_report.to_csv(index=False)
                
                # Create a download button
                st.download_button(
                    label="Baixar Relatório CSV",
                    data=csv,
                    file_name="relatorio_parametros.csv",
                    mime="text/csv",
                )
        
        # Tab 2: Export All Data
        with tab2:
            styled_title("Exportar Todos os Dados", level=2)
            st.write("Exporte todos os dados gerados pelo sistema em um único arquivo Excel.")
            
            if st.button("Coletar e Preparar Dados para Exportação"):
                # Collect all dataframes
                with st.spinner("Coletando dados de todas as páginas..."):
                    all_dataframes = self._collect_all_dataframes()
                    
                    if all_dataframes:
                        st.success(f"Dados coletados com sucesso! {len(all_dataframes)} conjuntos de dados encontrados.")
                        
                        # List the available dataframes
                        st.write("### Dados Disponíveis para Exportação:")
                        for i, sheet_name in enumerate(all_dataframes.keys(), 1):
                            st.write(f"{i}. {sheet_name}")
                        
                        # Create Excel file
                        with st.spinner("Preparando arquivo Excel..."):
                            excel_data = self._export_dataframes_to_excel(all_dataframes)
                            
                            if excel_data:
                                # Download button
                                st.download_button(
                                    label="Baixar Todos os Dados (Excel)",
                                    data=excel_data,
                                    file_name="TuaVita_dados_completos.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )
                            else:
                                st.error("Erro ao gerar o arquivo Excel.")
                    else:
                        st.warning("Nenhum conjunto de dados encontrado. Navegue pelo aplicativo para gerar dados primeiro.")

class MetasColabs(Page):
    def __init__(self):
        pass
    
    @property
    def title(self) -> str:
        return "Metas Mensais por Colaborador"
    
    @property
    def icon(self) -> str:
        return "🎯"
    
    def _calcular_metas_mensais(self):
        """Calcula as metas mensais por colaborador para SDRs e Closers"""
        # Get revenue data and commercial team data from Receitas
        receitas = Receitas()
        df_faturamento = receitas._gerar_dataframe_faturamento()
        df_equipe_comercial = receitas._calcular_equipe_comercial(df_faturamento)
        
        # Initialize DataFrame for targets
        meses = range(60)
        df_metas = pd.DataFrame(index=pd.Index([
            "Atendimentos Necessários por SDR", 
            "Conversões Necessárias por Closer"
        ]), columns=meses)
        
        if df_faturamento is not None and df_equipe_comercial is not None:
            # Calculate targets for each month
            for mes in meses:
                # Get total appointments and SDR count
                total_agendamentos = pd.to_numeric(df_faturamento.loc["Total de Agendamentos", mes], errors='coerce') or 0
                sdr_necessarios = pd.to_numeric(df_equipe_comercial.loc["SDR Necessários", mes], errors='coerce') or 1  # Default to 1 to avoid division by zero
                
                # Get total attendance and Closer count
                total_conversoes = pd.to_numeric(df_faturamento.loc["Conversões", mes], errors='coerce') or 0
                closer_necessarios = pd.to_numeric(df_equipe_comercial.loc["Closer Necessários", mes], errors='coerce') or 1  # Default to 1 to avoid division by zero
                
                # Calculate appointments per SDR and conversions per Closer
                if sdr_necessarios > 0:
                    df_metas.loc["Atendimentos Necessários por SDR", mes] = total_agendamentos / sdr_necessarios
                else:
                    df_metas.loc["Atendimentos Necessários por SDR", mes] = 0
                    
                if closer_necessarios > 0:
                    df_metas.loc["Conversões Necessárias por Closer", mes] = total_conversoes / closer_necessarios
                else:
                    df_metas.loc["Conversões Necessárias por Closer", mes] = 0
        
        return df_metas
    
    def render(self):
        styled_title(f"{self.icon} {self.title}")
        
        # Check if prerequisites are met
        if 'premissas_receitas' not in st.session_state:
            st.error("Premissas de receitas não definidas. Por favor, defina as premissas na página 'Premissas Receitas'.")
            return
        
        # Calculate monthly targets without needing valor_unitario input
        df_metas_mensais = self._calcular_metas_mensais()
        
        if df_metas_mensais is None or df_metas_mensais.empty:
            st.error("Não foi possível calcular as metas. Verifique se há dados de faturamento e equipe comercial disponíveis.")
            return
        
        # Display the monthly dataframe directly - no time period selection needed
        st.write("### Metas por Colaborador - Visão Mensal")
        st.dataframe(df_metas_mensais.style.format("{:.1f}"), use_container_width=True)
        
        # Add visualizations
        st.write("### Visualização Gráfica")
        
        # Line chart showing targets over time
        df_plot = df_metas_mensais.T.reset_index()
        df_plot.columns = ['Mês', 'Atendimentos por SDR', 'Conversões por Closer']
        
        fig = px.line(
            df_plot,
            x='Mês',
            y=['Atendimentos por SDR', 'Conversões por Closer'],
            title='Evolução das Metas por Colaborador',
            markers=True
        )
        
        fig.update_layout(
            xaxis_title="Mês",
            yaxis_title="Quantidade",
            legend_title="Métrica"
        )
        
        st.plotly_chart(fig, use_container_width=True)

class ProjeçãoInicial(Page):
    def __init__(self):
        pass
    
    @property
    def title(self) -> str:
        return "Projeção 12 meses"
    
    @property
    def icon(self) -> str:
        return "📊"
    
    def _get_receitas_data(self, metric="Arrecadação Líquida"):
        """Get the first 12 months of data for the selected metric"""
        try:
            receitas = Receitas()
            df_faturamento = receitas._gerar_dataframe_faturamento()
            df_arrecadacao = receitas._gerar_dataframe_arrecadacao(df_faturamento)
        
            if metric == "Arrecadação Líquida":
                if df_arrecadacao is not None and metric in df_arrecadacao.index:
                    # Get first 12 months
                    return pd.to_numeric(df_arrecadacao.loc[metric, :11], errors='coerce')
            else:
                if df_faturamento is not None and metric in df_faturamento.index:
                    # Get first 12 months
                    return pd.to_numeric(df_faturamento.loc[metric, :11], errors='coerce')
                    
            return None
        except Exception as e:
            st.warning(f"Erro ao obter dados de {metric}: {e}")
            return None
    
    def _calculate_regression(self, data):
        """Calculate OLS regression and 95% confidence interval"""
        
        if data is None or len(data) < 2:
            return None, None, None, None, None
        
        # Convert data to numpy arrays
        x = np.array(range(len(data)))
        y = np.array(data)
        
        # Calculate linear regression
        slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
        
        # Calculate predictions
        y_pred = intercept + slope * x
        
        # Calculate confidence interval
        n = len(data)
        mean_x = np.mean(x)
        ss_x = np.sum((x - mean_x)**2)
        residuals = y - y_pred
        ss_res = np.sum(residuals**2)
        se = np.sqrt(ss_res / (n - 2))
        se_line = se * np.sqrt(1/n + (x - mean_x)**2 / ss_x)
        t_value = stats.t.ppf(0.975, n - 2)
        
        # Confidence intervals
        ci_lower = y_pred - t_value * se_line
        ci_upper = y_pred + t_value * se_line
        
        return x, y_pred, ci_lower, ci_upper, r_value**2

    def render(self):
        
        styled_title(f"{self.icon} {self.title}")
        
        # Create selection box
        metric_options = [
            "Total de Agendamentos", 
            "Total de Comparecimento", 
            "Conversões", 
            "Arrecadação Líquida"
        ]
        
        selected_metric = st.selectbox(
            "Selecione a métrica para projeção",
            options=metric_options,
            index=0
        )
        
        # Get data based on selection
        data = self._get_receitas_data(selected_metric)
        
        if data is None or not hasattr(data, '__len__') or len(data) == 0:
            st.error(f"Não foi possível obter dados de {selected_metric}. Verifique se há dados disponíveis.")
            return
    
        # Calculate regression and confidence interval
        x_values, y_pred, ci_lower, ci_upper, r2 = self._calculate_regression(data)
        
        if x_values is None:
            st.error("Não foi possível calcular a regressão. Verifique se há dados suficientes.")
            return
        
        # Create a figure using Plotly Express
            # Ensure 'data' is a sequence before using len()
        if hasattr(data, '__len__'):
            data_length = len(data)
        else:
            data_length = 1 if data is not None else 0
            data = [data] if data is not None else []
            y_pred = [y_pred] if y_pred is not None else []

        months = [f"Mês {i+1}" for i in range(data_length)]
        
        # Create DataFrame for plotting
        df_plot = pd.DataFrame({
            'Mês': months,
            'Valor Original': data,
            'Linha de Tendência': y_pred,
        })
    
        # Create scatter plot with line
        fig = px.scatter(
            df_plot, 
            x='Mês', 
            y='Valor Original',
            title=f"Projeção de {selected_metric} - 12 Primeiros Meses"
        )
        
        # Add trend line
        fig.add_trace(
            px.line(df_plot, x='Mês', y='Linha de Tendência').data[0]
        )
        
        # Add confidence interval (filled area)
        fig.add_trace({
            'type': 'scatter',
            'x': months + months[::-1],
            'y': list(ci_upper if ci_upper is not None else []) + list(ci_lower if ci_lower is not None else [])[::-1],
            'fill': 'toself',
            'fillcolor': 'rgba(0,176,246,0.2)',
            'line': {'color': 'rgba(255,255,255,0)'},
            'hoverinfo': 'skip',
            'showlegend': True,
            'name': 'Intervalo de Confiança 95%'
        })
        
        # Update layout
        fig.update_layout(
            xaxis_title="Mês",
            yaxis_title=selected_metric,
            legend_title="Legenda"
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Display regression statistics
        st.write("### Estatísticas da Regressão")
        
            # Calculate RMSE (with type checking and conversion)
        if data is not None and y_pred is not None:
            # Convert both to numpy arrays if they aren't already
            data_array = np.array(data, dtype=float)
            y_pred_array = np.array(y_pred, dtype=float)
            
            # Ensure they have the same shape
            if data_array.shape == y_pred_array.shape:
                rmse = np.sqrt(np.mean((data_array - y_pred_array)**2))
            else:
                rmse = 0.0
                st.warning("Shapes of data and predictions don't match for RMSE calculation")
        else:
            rmse = 0.0
            st.warning("Missing data for RMSE calculation")
    
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("Coeficiente de Determinação (R²)", f"{r2:.4f}")
        
        with col2:
            st.metric("Erro Médio Quadrático (RMSE)", f"{rmse:.2f}")

class Dashboard:
    def __init__(self):
        # Group pages by category
        self.page_groups = {
            "Página Inicial": [
                PaginaAcompanhamento()
            ],
            "Despesas": [
                PremissasDespesas(),
                DespesasAdm(),
                Equipe(),
                CustosTecnologia()
            ],
            "Investimentos": [
                PremissasInvestimentos(), 
                Investimentos()
            ],
            "Comissão Vendas": [
                PremissasComissao(),
                ComissaoVendas()
            ],
            "Tributos": [
                PremissasTributos(),
                Tributos()
            ],
                "Receitas": [
                PremissasReceitas(),
                Receitas()
            ],
            "Projeções": [
                PremissasProjeções(),
                ProjeçãodeFluxodeCaixa(),
                ProjeçãoDRE(),
                MetasColabs(),
                ProjeçãoInicial()
            ],
        }
        
        # Initialize session state for navigation if not present
        if 'current_group' not in st.session_state:
            st.session_state.current_group = list(self.page_groups.keys())[0]
        if 'current_page' not in st.session_state:
            st.session_state.current_page = self.page_groups[st.session_state.current_group][0].title

    def render(self):
        # Display logo at the top of the sidebar
        try:
            logo_path = "C:\\Users\\Guilherme\\Documents\\Scratch\\TuaVitaLogo.jpg"
            st.sidebar.image(logo_path, width=300)
        except Exception as e:
            st.sidebar.error(f"Could not load logo: {e}")
        
        # Add a styled header for navigation
        st.sidebar.markdown("""
        <div style='background-color: #f0f2f6; padding: 10px; border-radius: 5px; margin-bottom: 15px'>
            <h2 style='text-align: center; color: #262730; margin: 0px'>Navegação</h2>
        </div>
        """, unsafe_allow_html=True)
        
        # Initialize session state variables if not present
        if 'nav_view' not in st.session_state:
            st.session_state.nav_view = "groups"
        if 'current_group' not in st.session_state:
            st.session_state.current_group = list(self.page_groups.keys())[0]
        if 'current_page' not in st.session_state:
            st.session_state.current_page = self.page_groups[st.session_state.current_group][0].title
        
        # Navigation UI
        if st.session_state.nav_view == "pages":
            # Back button when viewing pages
            if st.sidebar.button("← Voltar para Grupos", use_container_width=True):
                st.session_state.nav_view = "groups"
            
            # Divider
            st.sidebar.markdown("<hr>", unsafe_allow_html=True)
            
            # Page buttons for the selected group
            current_group = st.session_state.current_group
            st.sidebar.markdown(f"#### Páginas em {current_group}:")
            
            pages_in_group = self.page_groups[current_group]
            for page in pages_in_group:
                page_title = f"{page.icon} {page.title}"
                if st.sidebar.button(page_title, key=f"page_{page.title}", use_container_width=True):
                    st.session_state.current_page = page.title
        else:
            # Group buttons
            st.sidebar.markdown("#### Selecione um grupo:")
            
            for group_name in self.page_groups.keys():
                if st.sidebar.button(f"📁 {group_name}", key=f"group_{group_name}", use_container_width=True):
                    st.session_state.current_group = group_name
                    st.session_state.nav_view = "pages"
                    st.session_state.current_page = self.page_groups[group_name][0].title
        
        # Divider before footer
        st.sidebar.markdown("<hr style='margin-top: 15px;'>", unsafe_allow_html=True)
        
        # Render the current page
        current_group = st.session_state.current_group
        current_page = st.session_state.current_page
        
        for page in self.page_groups[current_group]: 
            if page.title == current_page:
                page.render()
                break
        
        # Footer
        st.sidebar.markdown("<div style='text-align: center; color: #888888; font-size: 12px'>TuaVita Dashboard v1.0</div>", unsafe_allow_html=True)

if __name__ == "__main__":
    dashboard = Dashboard()
    dashboard.render()